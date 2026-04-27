import csv
import json
from datetime import date, timedelta, datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count, Q, Prefetch
from django.db.models import Case, When, IntegerField, Value
from config.decorators import admin_required, researcher_required
from config.utils import log_action, generate_project_summary
from config.models import (
    CustomUser, ResearcherProfile, StudentProfile,
    Publication, Activity, Funding, Project,
    ProjectMember, Education, Recognition,
    SupervisionRecord, StudentNotification, StrategicObjective, SupervisionRecord, Recognition, Conference
)


# ─────────────────────────────────────────────
# PI Report
# ─────────────────────────────────────────────

@login_required
@researcher_required
def pi_report_view(request):
    from django.shortcuts import get_object_or_404
    from django.db.models import Prefetch, Case, When, IntegerField, Count, Sum, Q

    researcher = get_object_or_404(ResearcherProfile, user=request.user)

    # ── XLSX Export ───────────────────────────────────────────
    if request.GET.get('export') == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        RED   = 'C8102E'
        GREY  = 'F3F4F6'
        WHITE = 'FFFFFF'

        wb = Workbook()
        wb.remove(wb.active)

        def make_sheet(wb, title, headers):
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font      = Font(bold=True, color=WHITE, name='Arial', size=11)
                cell.fill      = PatternFill('solid', start_color=RED)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 20
            return ws

        def auto_width(ws, min_w=12, max_w=60):
            for col in ws.columns:
                length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in col
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

        def style_data_rows(ws):
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                bg = GREY if row_idx % 2 == 0 else WHITE
                for cell in row:
                    cell.fill      = PatternFill('solid', start_color=bg)
                    cell.font      = Font(name='Arial', size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # ── Sheet 1: Projects & Grants ────────────────────────
        ws = make_sheet(wb, 'Projects & Grants', [
            'Title', 'Organization', 'Program', 'Role', 'Status',
            'Currency', 'Total Funding', 'Awarded to IBME', 'Kept by IBME',
            'Start Date', 'End Date', 'Funding Type',
            'Description', 'Conception', 'Next Steps', 'IP Activities',
            'HQP', 'Collaborators & Partners',
        ])
        for p in Project.objects.filter(
            researcher=researcher, is_deleted=False
        ).exclude(status='rejected').prefetch_related(
            Prefetch(
                'tagged_members',
                queryset=CustomUser.objects.filter(user_type='student'),
                to_attr='hqp_students',
            ),
            'team_members',
        ).order_by('-start_date'):
            hqp     = '; '.join(u.get_full_name() for u in p.hqp_students)
            collabs = '; '.join(
                f"{m.name} ({m.get_partner_type_display()})" for m in p.team_members.all()
            )
            ws.append([
                p.title,
                p.funding_organization or '',
                p.program_name or '',
                p.get_role_display() or '',
                p.get_status_display() or '',
                p.currency or 'CAD',
                float(p.total_funding)       if p.total_funding       else '',
                float(p.funding_received)    if p.funding_received    else '',
                float(p.funding_kept_by_unb) if p.funding_kept_by_unb else '',
                str(p.start_date) if p.start_date else '',
                str(p.end_date)   if p.end_date   else '',
                p.funding_type  or '',
                p.description   or '',
                p.conception    or '',
                p.next_steps    or '',
                p.ip_activities or '',
                hqp,
                collabs,
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 2: Publications ─────────────────────────────
        ws = make_sheet(wb, 'Publications', [
            'Title', 'Type', 'Status', 'Authors', 'Journal / Venue',
            'Date', 'Year', 'Volume', 'Issue', 'Pages', 'DOI',
            'Refereed', 'Open Access', 'Invited', 'Language', 'Abstract',
        ])
        for p in Publication.objects.filter(
            researcher=researcher, is_active=True
        ).order_by('-publication_date'):
            ws.append([
                p.title,
                p.get_publication_type_display(),
                p.get_status_display() or '',
                p.authors  or '',
                p.journal  or '',
                str(p.publication_date)     if p.publication_date else '',
                str(p.publication_date)[:4] if p.publication_date else '',
                p.volume   or '',
                p.issue    or '',
                p.pages    or '',
                p.doi      or '',
                'Yes' if p.refereed    else ('No' if p.refereed    is False else ''),
                'Yes' if p.open_access else ('No' if p.open_access is False else ''),
                'Yes' if p.invited     else ('No' if p.invited     is False else ''),
                p.language or '',
                p.abstract or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 3: Activities ───────────────────────────────
        ws = make_sheet(wb, 'Activities', [
            'Title', 'Category', 'Type', 'Date', 'Location',
            'Invited', 'Keynote', 'Co-Presenters', 'Audience', 'Description',
        ])
        for a in Activity.objects.filter(
            researcher=researcher, is_active=True
        ).order_by('-date'):
            ws.append([
                a.title,
                a.get_category_display()      or '',
                a.get_activity_type_display() or '',
                str(a.date) if a.date else '',
                a.location      or '',
                'Yes' if a.invited else ('No' if a.invited is False else ''),
                'Yes' if a.keynote else ('No' if a.keynote is False else ''),
                a.co_presenters or '',
                a.audience      or '',
                a.description   or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 4: Supervision ──────────────────────────────
        ws = make_sheet(wb, 'Supervision', [
            'Student', 'Degree', 'Department', 'Role', 'Institution',
            'Start', 'End', 'Status', 'Thesis Title',
        ])
        for s in SupervisionRecord.objects.filter(
            researcher=researcher
        ).order_by('-start_date'):
            ws.append([
                s.student_name,
                s.get_degree_type_display() or '',
                s.department       or '',
                s.supervision_role or '',
                s.institution      or '',
                str(s.start_date) if s.start_date else '',
                str(s.end_date)   if s.end_date   else '',
                s.get_status_display() or '',
                s.thesis_title or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 5: Funding Breakdown ────────────────────────
        # grant_total = full project Total Funding from XML (blue bar)
        # amount      = researcher's allocated portion (may be 0 for Co-I)
        # amount_to_ibme = what actually came to IBME
        ws = make_sheet(wb, 'Funding Breakdown', [
            'Grant Title', 'Organization', 'Program', 'Role', 'Status',
            'Currency', 'Grant Total', 'Researcher Portion', 'Amount to IBME',
            'Start Date', 'End Date', 'Funding Type',
        ])
        for f in Funding.objects.filter(
            researcher=researcher
        ).exclude(status='rejected').order_by('-start_date'):
            ws.append([
                f.title,
                f.organization  or '',
                f.program_name  or '',
                f.get_role_display()   or '',
                f.get_status_display() or '',
                f.currency or 'CAD',
                float(f.grant_total) if f.grant_total else (float(f.amount) if f.amount else ''),
                float(f.amount)         if f.amount         else '',
                float(f.amount_to_ibme) if f.amount_to_ibme else '',
                str(f.start_date) if f.start_date else '',
                str(f.end_date)   if f.end_date   else '',
                f.funding_type or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="pi_report_{researcher.user.last_name}.xlsx"'
        )
        return response

    # ── Date filters ──────────────────────────────────────────
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    date_from     = None
    date_to       = None

    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except ValueError:
        pass
    try:
        if date_to_str:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        pass

    # ── Projects ──────────────────────────────────────────────
    projects_qs = Project.objects.filter(
        researcher=researcher,
        is_deleted=False,
    ).exclude(
        status='rejected',
    ).prefetch_related(
        'team_members',
        Prefetch(
            'tagged_members',
            queryset=CustomUser.objects.select_related('student_profile'),
        ),
        'linked_publications',
        'linked_activities',
        'funding_breakdown',
    )

    if date_from:
        projects_qs = projects_qs.filter(start_date__gte=date_from)
    if date_to:
        projects_qs = projects_qs.filter(start_date__lte=date_to)

    projects = projects_qs.order_by(
        Case(When(status='awarded', then=0), default=1, output_field=IntegerField()),
        '-start_date',
    )

    # ── Funding records ───────────────────────────────────────
    funding_records = Funding.objects.filter(
        researcher=researcher,
    ).exclude(
        status='rejected',
    ).select_related(
        'project',
    ).prefetch_related(
        'project__team_members',
    ).order_by('-start_date')

    if date_from:
        funding_records = funding_records.filter(start_date__gte=date_from)
    if date_to:
        funding_records = funding_records.filter(start_date__lte=date_to)

    # ── Publications ──────────────────────────────────────────
    publications_qs = Publication.objects.filter(
        researcher=researcher, is_active=True
    ).order_by('-publication_date')

    if date_from:
        publications_qs = publications_qs.filter(publication_date__gte=date_from)
    if date_to:
        publications_qs = publications_qs.filter(publication_date__lte=date_to)

    # ── Activities ────────────────────────────────────────────
    activities_qs = Activity.objects.filter(
        researcher=researcher, is_active=True
    ).select_related('conference').prefetch_related(
        'tagged_users', 'objectives'
    ).order_by('-date')

    if date_from:
        activities_qs = activities_qs.filter(date__gte=date_from)
    if date_to:
        activities_qs = activities_qs.filter(date__lte=date_to)

    # ── Supervision ───────────────────────────────────────────
    supervision = SupervisionRecord.objects.filter(researcher=researcher)

    if date_from:
        supervision = supervision.filter(start_date__gte=date_from)
    if date_to:
        supervision = supervision.filter(start_date__lte=date_to)

    supervision = supervision.order_by(
        Case(
            When(status='in_progress', then=0),
            When(status='completed',   then=1),
            default=2,
            output_field=IntegerField(),
        ),
        '-start_date',
    )

    # ── KPI aggregation ───────────────────────────────────────
    # total_funding: PI-role awarded grants only (researcher controls budget)
    # total_funding_all: all roles awarded (full association value)
    proj_stats = projects_qs.aggregate(
        total=Count('id'),
        awarded=Count('id', filter=Q(status='awarded')),
        submitted=Count('id', filter=Q(status='submitted')),
        completed=Count('id', filter=Q(status='completed')),
        pi_funding=Sum(                    # ← renamed
            'total_funding',
            filter=Q(status='awarded', role__in=['pi', 'pa'])
        ),
        all_funding=Sum(                   # ← renamed
            'total_funding',
            filter=Q(status='awarded')
        ),
    )

    return render(request, 'Pages/reports/pi_report.html', {
        'researcher':         researcher,
        'projects':           projects,
        'funding_records':    funding_records,
        'publications':       publications_qs,
        'activities':         activities_qs,
        'supervision':        supervision,
        'date_from':          date_from_str,
        'date_to':            date_to_str,
        'proj_count':         proj_stats['total']      or 0,
        'pub_count':          publications_qs.count(),
        'act_count':          activities_qs.count(),
        'sup_count':          supervision.count(),
        'funding_count':      funding_records.count(),
        'awarded_count':      proj_stats['awarded']    or 0,
        'submitted_count':    proj_stats['submitted']  or 0,
        'completed_count':    proj_stats['completed']  or 0,
        'total_funding':      proj_stats['pi_funding']  or 0,   # ← updated
        'total_funding_all':  proj_stats['all_funding'] or 0,   # ← updated
    })

# ─────────────────────────────────────────────
# Director Reports
# ─────────────────────────────────────────────

@login_required
@admin_required
def reports_list(request):
    

    consenting_students = StudentProfile.objects.filter(
        user__consent_to_share=True,
        degree_level__in=['msc', 'phd'],
    ).count()

    # ── ADD THIS ──────────────────────────────────────────────
    linked_students = SupervisionRecord.objects.filter(
        linked_student__isnull=False,
        status='in_progress',
    ).values('linked_student').distinct().count()

    return render(request, 'Pages/reports/reports_list.html', {
        'consenting_students': consenting_students,
        'linked_students':     linked_students,
    })

@login_required
@admin_required
def enrollment_trends_report(request):
    from datetime import date
    from collections import defaultdict

    current_year  = date.today().year
    n_years       = int(request.GET.get('years', 10))
    filter_pi     = request.GET.get('pi', '')
    filter_degree = request.GET.get('degree', '')
    filter_dept   = request.GET.get('department', '')  # fixed: was 'dept'

    start_year = current_year - n_years + 1

    degree_filter_map = {
        'undergrad': ['bachelors'],
        'msc':       ['masters_thesis', 'masters_non_thesis'],
        'phd':       ['doctorate'],
    }

    def get_level(degree_type):
        if degree_type == 'bachelors':                              return 'undergrad'
        if degree_type in ('masters_thesis', 'masters_non_thesis'): return 'msc'
        if degree_type == 'doctorate':                              return 'phd'
        return 'other'

    records = SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        start_date__isnull=False,
        start_date__year__gte=start_year,   # fixed: respect selected year window
        degree_type__in=['bachelors', 'masters_thesis', 'masters_non_thesis', 'doctorate'],
    ).select_related('researcher__user')

    if filter_pi:
        records = records.filter(researcher__user__id=filter_pi)
    if filter_degree:
        mapped = degree_filter_map.get(filter_degree, [])
        if mapped:
            records = records.filter(degree_type__in=mapped)
    if filter_dept:
        records = records.filter(department__icontains=filter_dept)

    # single DB hit
    record_list = list(records.values(
        'start_date', 'end_date', 'degree_type', 'status', 'department',
        'researcher__user__first_name', 'researcher__user__last_name',
        'researcher__user__id',
    ))

    # precompute per record
    precomputed = []
    for r in record_list:
        s = r['start_date'].year
        e = r['end_date'].year if r['end_date'] else current_year
        precomputed.append({
            'start':      s,
            'end':        e,
            'level':      get_level(r['degree_type']),
            'status':     r['status'],
            'department': r['department'] or '',
            'pi':         f"{r['researcher__user__first_name']} {r['researcher__user__last_name']}".strip(),
            'pi_id':      r['researcher__user__id'],
        })

    # yearly counts
    year_counts = {y: {'undergrad': 0, 'msc': 0, 'phd': 0}
                   for y in range(start_year, current_year + 1)}

    for r in precomputed:
        level = r['level']
        if level == 'other':
            continue
        for y in range(max(r['start'], start_year), min(r['end'], current_year) + 1):
            year_counts[y][level] += 1

    yearly_data = []
    for y in range(start_year, current_year + 1):
        ug  = year_counts[y]['undergrad']
        msc = year_counts[y]['msc']
        phd = year_counts[y]['phd']
        yearly_data.append({
            'year': y, 'undergrad': ug, 'msc': msc, 'phd': phd,
            'total': ug + msc + phd,
        })

    for i, row in enumerate(yearly_data):
        row['yoy'] = 0 if i == 0 else row['total'] - yearly_data[i - 1]['total']

    # summary stats
    total_unique    = len(precomputed)
    total_undergrad = sum(1 for r in precomputed if r['level'] == 'undergrad')
    total_msc       = sum(1 for r in precomputed if r['level'] == 'msc')
    total_phd       = sum(1 for r in precomputed if r['level'] == 'phd')
    total_graduated = sum(1 for r in precomputed if r['status'] == 'completed')
    total_active    = sum(1 for r in precomputed if r['status'] == 'in_progress')

    completed_years = [row for row in yearly_data if row['year'] < current_year]

    first_total  = completed_years[0]['total']  if len(completed_years) >= 2 else 0
    last_total   = completed_years[-1]['total'] if len(completed_years) >= 2 else 0
    growth_delta = last_total - first_total

    if first_total >= 10:
        growth_rate = round(((last_total - first_total) / first_total * 100), 1)
    else:
        growth_rate = None

    growth_end_year = current_year - 1

    # by PI — always set pi_id regardless of level
    by_pi = defaultdict(lambda: {'undergrad': 0, 'msc': 0, 'phd': 0, 'pi_id': None})
    for r in precomputed:
        by_pi[r['pi']]['pi_id'] = r['pi_id']   # fixed: set pi_id for all rows
        if r['level'] in ('undergrad', 'msc', 'phd'):
            by_pi[r['pi']][r['level']] += 1

    by_pi_list = sorted([
        {
            'pi':        k,
            'pi_id':     v['pi_id'],
            'undergrad': v['undergrad'],
            'msc':       v['msc'],
            'phd':       v['phd'],
            'total':     v['undergrad'] + v['msc'] + v['phd'],
        }
        for k, v in by_pi.items()
    ], key=lambda x: x['total'], reverse=True)

    # by department
    by_dept = defaultdict(lambda: {'undergrad': 0, 'msc': 0, 'phd': 0})
    for r in precomputed:
        if r['level'] in ('undergrad', 'msc', 'phd') and r['department']:
            by_dept[r['department']][r['level']] += 1

    by_dept_list = sorted([
        {'dept': k, 'undergrad': v['undergrad'], 'msc': v['msc'], 'phd': v['phd'],
         'total': v['undergrad'] + v['msc'] + v['phd']}
        for k, v in by_dept.items()
    ], key=lambda x: x['total'], reverse=True)

    # department dropdown options
    all_departments = list(
        SupervisionRecord.objects.filter(
            institution__icontains='University of New Brunswick',
            department__isnull=False,
        ).exclude(
            department=''
        ).values_list('department', flat=True).distinct().order_by('department')
    )

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/enrollment_trends_report.html', {
        'yearly_data':     yearly_data,
        'total_unique':    total_unique,
        'total_undergrad': total_undergrad,
        'total_msc':       total_msc,
        'total_phd':       total_phd,
        'total_graduated': total_graduated,
        'total_active':    total_active,
        'growth_rate':     growth_rate,
        'growth_delta': growth_delta,
        'growth_end_year': growth_end_year,
        'start_year':      start_year,
        'current_year':    current_year,
        'n_years':         n_years,
        'by_pi':           by_pi_list,
        'by_dept':         by_dept_list,
        'filter_degree':   filter_degree,
        'filter_pi':       filter_pi,
        'filter_dept':     filter_dept,
        'all_supervisors': all_supervisors,
        'all_departments': all_departments,
        'degree_choices': [
            ('undergrad', 'Undergraduate'),
            ('msc', 'MSc'),
            ('phd', 'PhD'),
        ],
    })


@login_required
@admin_required
def api_enrollment_pi_students(request, pi_id):
    from django.db.models import Case, When, IntegerField, Value
    from datetime import date

    pi = get_object_or_404(CustomUser, id=pi_id, user_type='researcher')

    # respect the same year window as the main report
    n_years    = int(request.GET.get('years', 10))
    start_year = date.today().year - n_years + 1

    records = SupervisionRecord.objects.filter(
        researcher__user=pi,
        start_date__year__gte=start_year,   # fixed: respect selected year window
        degree_type__in=['bachelors', 'masters_thesis', 'masters_non_thesis', 'doctorate'],
    ).select_related(
        'researcher__user',
        'linked_student__user'
    ).annotate(
        degree_order=Case(
            When(degree_type='bachelors',          then=0),
            When(degree_type='masters_thesis',     then=1),
            When(degree_type='masters_non_thesis', then=1),
            When(degree_type='doctorate',          then=2),
            output_field=IntegerField()
        ),
        status_order=Case(
            When(status='in_progress', then=0),
            When(status='completed',   then=1),
            default=Value(2),
            output_field=IntegerField()
        ),
    ).order_by('status_order', 'degree_order', 'start_date')

    students = [{
        'name':       r.student_name,
        'degree':     r.get_degree_type_display() or '',
        'status':     r.get_status_display() or '',
        'start':      r.start_date.strftime('%Y-%m-%d') if r.start_date else '',
        'expected':   r.expected_date.strftime('%Y-%m-%d') if r.expected_date else '',
        'department': r.department or '',
        'linked':     r.linked_student.user.get_full_name()
                      if r.linked_student and r.linked_student.user
                      else None,
    } for r in records]

    return JsonResponse({
        'pi':       pi.get_full_name(),
        'count':    len(students),
        'students': students,
    })


@login_required
@admin_required
def grad_completion_report(request):
    from collections import defaultdict
    from datetime import date
    import json

    cutoff_year  = date.today().year - 10
    current_year = date.today().year

    filter_degree = request.GET.get('degree', '')
    filter_pi     = request.GET.get('pi', '')

    degree_filter_map = {
        'msc': ['masters_thesis', 'masters_non_thesis'],
        'phd': ['doctorate'],
        'pdf': ['postdoc'],
    }

    def get_degree_label(degree_type):
        return {
            'masters_thesis':     'MSc',
            'masters_non_thesis': 'MSc',
            'doctorate':          'PhD',
            'postdoc':            'Post-Doc',
        }.get(degree_type, degree_type)

    def get_degree_level(degree_type):
        return {
            'masters_thesis':     'msc',
            'masters_non_thesis': 'msc',
            'doctorate':          'phd',
            'postdoc':            'pdf',
        }.get(degree_type, '')

    # ── Single DB query for completed records ─────────────────
    sup_records = SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        status='completed',
        degree_start_date__isnull=False,
        degree_type__in=['masters_thesis', 'masters_non_thesis', 'doctorate', 'postdoc'],
        end_date__isnull=False,
        end_date__year__gte=cutoff_year,
        end_date__lte=date.today(),
    ).select_related('researcher__user')

    if filter_degree:
        mapped = degree_filter_map.get(filter_degree, [])
        if mapped:
            sup_records = sup_records.filter(degree_type__in=mapped)
    if filter_pi:
        sup_records = sup_records.filter(researcher__user__id=filter_pi)

    # ── Build records list — single pass ──────────────────────
    records = []
    for r in sup_records:
        grad_date = r.degree_end_date or r.end_date
        start     = r.degree_start_date
        if not start or not grad_date:
            continue
        years = (grad_date - start).days / 365.25
        if years <= 0:
            continue
        records.append({
            'name':            r.student_name,
            'degree':          get_degree_label(r.degree_type),
            'degree_level':    get_degree_level(r.degree_type),
            'department': r.department or '—',
            'start_date':      start,
            'graduation_date': grad_date,
            'years':           round(years, 2),
            'pi':              r.researcher.user.get_full_name(),
            'pi_id':           r.researcher.user.id,
        })

    # ── Summary stats — single pass ───────────────────────────
    total_students = len(records)
    average_time   = round(sum(r['years'] for r in records) / total_students, 2) if total_students else 0

    # ── By degree ─────────────────────────────────────────────
    by_degree = defaultdict(lambda: {'count': 0, 'total': 0.0, 'min': float('inf'), 'max': 0.0})
    for r in records:
        d = r['degree']
        by_degree[d]['count'] += 1
        by_degree[d]['total'] += r['years']
        by_degree[d]['min']    = min(by_degree[d]['min'], r['years'])
        by_degree[d]['max']    = max(by_degree[d]['max'], r['years'])

    by_degree_final = {
        k: {
            'count':   v['count'],
            'average': round(v['total'] / v['count'], 2),
            'min':     round(v['min'], 2),
            'max':     round(v['max'], 2),
        }
        for k, v in sorted(by_degree.items())
    }

    # ── By year ───────────────────────────────────────────────
    by_year = defaultdict(lambda: {'count': 0, 'total': 0.0, 'students': []})
    for r in records:
        y = r['graduation_date'].year
        by_year[y]['count']    += 1
        by_year[y]['total']    += r['years']
        by_year[y]['students'].append(f"{r['name']} ({r['degree']}, {r['years']}y)")

    by_year_final = {
        y: {
            'count':    v['count'],
            'average':  round(v['total'] / v['count'], 2),
            'students': v['students'],
        }
        for y, v in sorted(by_year.items())
    }

    # ── By PI ─────────────────────────────────────────────────
    by_pi = defaultdict(lambda: {'count': 0, 'total': 0.0, 'degrees': set()})
    for r in records:
        by_pi[r['pi']]['count']   += 1
        by_pi[r['pi']]['total']   += r['years']
        by_pi[r['pi']]['degrees'].add(r['degree'])

    by_pi_final = sorted([
        {
            'pi':      k,
            'count':   v['count'],
            'average': round(v['total'] / v['count'], 2),
            'degrees': ', '.join(sorted(v['degrees'])),
        }
        for k, v in by_pi.items()
    ], key=lambda x: x['count'], reverse=True)

    # ── Yearly counts for chart — NO nested loop ──────────────
    # Single query, precompute, then bucket by year
    all_sup = list(SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        degree_type__in=['masters_thesis', 'masters_non_thesis', 'doctorate'],
        start_date__isnull=False,
    ).values('start_date', 'end_date', 'degree_type'))

    chart_years = range(current_year - 9, current_year + 1)
    year_buckets = {y: {'msc': 0, 'phd': 0} for y in chart_years}

    for s in all_sup:
        sy = s['start_date'].year
        ey = s['end_date'].year if s['end_date'] else current_year
        level = 'msc' if s['degree_type'] in ('masters_thesis', 'masters_non_thesis') else 'phd'
        for y in range(max(sy, current_year - 9), min(ey, current_year) + 1):
            year_buckets[y][level] += 1

    yearly_counts = [
        {
            'year':  y,
            'msc':   year_buckets[y]['msc'],
            'phd':   year_buckets[y]['phd'],
            'total': year_buckets[y]['msc'] + year_buckets[y]['phd'],
        }
        for y in sorted(chart_years)
    ]

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grad_completion_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Degree', 'Department', 'PI',
                         'Start Date', 'Graduation Date', 'Years'])
        for r in records:
            writer.writerow([
                r['name'], r['degree'], r['department'], r['pi'],
                r['start_date'], r['graduation_date'], r['years'],
            ])
        return response

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/grad_completion_report.html', {
        'records':        records,
        'total_students': total_students,
        'average_time':   average_time,
        'by_degree':      by_degree_final,
        'by_year':        by_year_final,
        'by_pi':          by_pi_final,
        'yearly_counts':  json.dumps(yearly_counts),
        'cutoff_year':    cutoff_year,
        'current_year':   current_year,
        'filter_degree':  filter_degree,
        'filter_pi':      filter_pi,
        'filter_dept':    '',
        'all_supervisors': all_supervisors,
        'all_departments': [],
        'degree_choices': [('msc', 'MSc'), ('phd', 'PhD'), ('pdf', 'Post-Doc')],
    })

from django.db.models import Prefetch

@login_required
@admin_required
def active_projects_report(request):

    filter_pi = request.GET.get('pi', '')

    active_projects = Project.objects.active().select_related(
        'researcher__user'
    ).prefetch_related(
        'team_members',
        Prefetch(
            'tagged_members',
            queryset=CustomUser.objects.select_related('student_profile'),
        ),
        Prefetch(
            'linked_publications',
            queryset=Publication.objects.filter(is_active=True),
            to_attr='active_publications',
        ),
        Prefetch(
            'linked_activities',
            queryset=Activity.objects.filter(is_active=True),
            to_attr='active_activities',
        ),
    ).order_by('researcher__user__last_name', 'title')

    if filter_pi:
        active_projects = active_projects.filter(researcher__user__id=filter_pi)

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="active_projects_report.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Title', 'PI', 'Status', 'Role', 'Funding Org', 'Funding Type',
            'Start Date', 'End Date', 'Currency', 'Total Funding', 'Awarded to UNB', 'Kept by UNB',
            'HQP (Tagged Students)', 'Academic Collaborators',
            'Industry / Community / Govt Partners',
            'Linked Publications', 'KM Activities', 'IP/Patents',
        ])
        for project in active_projects:
            team       = project.team_members.all()
            tagged     = list(project.tagged_members.all())
            hqp_tagged = [u for u in tagged if u.user_type == 'student']
            pubs       = project.active_publications
            acts       = project.active_activities

            pub_count = len(pubs)
            km_count  = sum(1 for a in acts if a.category == 'knowledge_mobilization')
            ip_count  = sum(1 for p in pubs if p.publication_type == 'patent')

            writer.writerow([
                project.title,
                project.researcher.user.get_full_name(),
                project.get_status_display(),
                project.get_role_display(),
                project.funding_organization or '',
                project.funding_type or '',
                project.start_date or '',
                project.end_date or '',
                project.currency or 'CAD',
                project.total_funding or '',
                project.funding_received or '',
                project.funding_kept_by_unb or '',
                '; '.join(u.get_full_name() for u in hqp_tagged),
                '; '.join(m.name for m in team.filter(
                    Q(is_academic_collaborator=True) | Q(partner_type='academic')
                )),
                '; '.join(
                    f"{m.name} ({m.get_partner_type_display()})"
                    for m in team.filter(
                        partner_type__in=['industry', 'community', 'government', 'other']
                    )
                ),
                pub_count, km_count, ip_count,
            ])
        return response

    # ── Build project_data for HTML render ────────────────────
    project_data = []
    total_hqp = total_pubs = total_km = total_ip = 0

    for project in active_projects:
        team   = project.team_members.all()
        tagged = list(project.tagged_members.all())
        pubs   = project.active_publications
        acts   = project.active_activities

        # ── HQP — compute from prefetched data in Python ─────
        hqp_members    = [u for u in tagged if u.user_type == 'student']
        tagged_collabs = [u for u in tagged if u.user_type == 'researcher']
        t_hqp          = len(hqp_members)

        phd_count = msc_count = pdf_count = 0
        for u in hqp_members:
            try:
                dl = u.student_profile.degree_level
                if dl == 'phd':   phd_count += 1
                elif dl == 'msc': msc_count += 1
                elif dl == 'pdf': pdf_count += 1
            except Exception:
                pass
        other_count = t_hqp - phd_count - msc_count - pdf_count

        # ── Counts — Python, no extra queries ────────────────
        km_acts  = [a for a in acts if a.category == 'knowledge_mobilization']
        ip_pubs  = [p for p in pubs if p.publication_type == 'patent']

        pub_count = len(pubs)
        km_count  = len(km_acts)
        ip_count  = len(ip_pubs)

        total_hqp  += t_hqp
        total_pubs += pub_count
        total_km   += km_count
        total_ip   += ip_count

        project_data.append({
            'project':                project,
            'pi_name':                project.researcher.user.get_full_name(),
            'pi_email':               project.researcher.user.email,
            'conception':             project.conception or '',
            'summary':                generate_project_summary(project),
            'total_hqp':              t_hqp,
            'phd_count':              phd_count,
            'msc_count':              msc_count,
            'pdf_count':              pdf_count,
            'other_count':            other_count,
            'hqp_members':            hqp_members,
            'tagged_collabs':         tagged_collabs,
            'academic_collaborators': team.filter(
                Q(is_academic_collaborator=True) | Q(partner_type='academic')
            ),
            'partners': team.filter(
                partner_type__in=['industry', 'community', 'government', 'other']
            ),
            'linked_pubs':  pubs,
            'pub_count':    pub_count,
            'km_acts':      km_acts,
            'all_acts':     acts,
            'km_count':     km_count,
            'ip_pubs':      ip_pubs,
            'ip_count':     ip_count,
        })

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/active_projects_report.html', {
        'project_data':    project_data,
        'total_projects':  len(project_data),
        'total_hqp':       total_hqp,
        'total_pubs':      total_pubs,
        'total_km':        total_km,
        'total_ip':        total_ip,
        'all_supervisors': all_supervisors,
        'filter_pi':       filter_pi,
    })


@login_required
@admin_required
def funding_analysis_report(request):
    from collections import defaultdict

    projects = Project.objects.filter(
        ccv_active=True,
        is_deleted=False,
    ).select_related('researcher__user')

    funding_records = Funding.objects.filter(
        is_deleted=False,
    ).select_related('researcher__user', 'project')

    EMPTY_CTX = {
        'error': 'No funding data available',
        'total_funding': '$0', 'by_organization': {}, 'by_year': {},
        'top_orgs': [], 'timeline_data': [], 'active_grants_detail': [],
        'min_year': 2024, 'max_year': 2024,
        'last3_by_org': [], 'last3_by_program': [], 'last3_by_pi': [],
        'last3_as_pi': {'amount': 0, 'count': 0},
        'last3_as_coi': {'amount': 0, 'count': 0},
        'depletion_rows': [], 'depletion_max': 0,
        'total_cad': 0, 'total_usd': 0, 'cad_count': 0, 'usd_count': 0,
        'by_pi_list':   [],
        'award_total': 0, 'award_count': 0,
        'schol_total': 0, 'schol_count': 0,
        'current_year': datetime.now().year,
        'last3_start':  datetime.now().year - 3,
        'spark_svg_w':  0,
        'spark_svg_h':  28,
        'year_from':    datetime.now().year,
        'year_to':      datetime.now().year,
    }

    if not projects.exists():
        return render(request, 'Pages/reports/funding_analysis_report.html', EMPTY_CTX)

    try:
        current_year = datetime.now().year
        year_from    = request.GET.get('year_from')
        year_to      = request.GET.get('year_to')

        def parse_year(date_val):
            if not date_val:
                return None
            try:
                s = str(date_val)
                return int(s.split('/')[0] if '/' in s else s[:4])
            except Exception:
                return None

        # ── Deduplicate projects ──────────────────────────────────────
        # Same title + same total_funding = same grant (different amounts = renewal)
        seen = {}
        for p in projects:
            # Use external_id if available, fall back to title-only
            key = p.external_id if p.external_id else p.title.strip().lower()
            if key not in seen:
                seen[key] = p
        projects_deduped = list(seen.values())
        
        all_years    = [y for y in (parse_year(p.start_date) for p in projects_deduped if p.start_date) if y]
        min_year     = min(all_years) if all_years else current_year
        max_year     = max(all_years) if all_years else current_year
        year_from_int = int(year_from) if year_from else min_year
        year_to_int   = int(year_to)   if year_to   else max_year

        # ── KPI totals ────────────────────────────────────────────────
        total_funding_amount = sum(float(p.total_funding or 0) for p in projects_deduped)
        total_grants         = len(projects_deduped)
        average_grant        = total_funding_amount / total_grants if total_grants else 0

        total_cad = sum(float(p.total_funding or 0) for p in projects_deduped if (p.currency or 'CAD') != 'USD')
        total_usd = sum(float(p.total_funding or 0) for p in projects_deduped if p.currency == 'USD')
        cad_count = sum(1 for p in projects_deduped if (p.currency or 'CAD') != 'USD')
        usd_count = sum(1 for p in projects_deduped if p.currency == 'USD')

        # ── By year ───────────────────────────────────────────────────
        by_year = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in projects_deduped:
            y = parse_year(p.start_date)
            if y:
                by_year[y]['amount'] += float(p.total_funding or 0)
                by_year[y]['count']  += 1

        # ── By organization (filtered year range) ─────────────────────
        by_organization = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in projects_deduped:
            sy = parse_year(p.start_date)
            ey = parse_year(p.end_date)
            if not sy or not ey:
                continue
            if sy > year_to_int or ey < year_from_int:
                continue
            org = p.funding_organization or 'Unknown'
            by_organization[org]['amount'] += float(p.total_funding or 0)
            by_organization[org]['count']  += 1

        top_orgs = sorted(
            by_organization.items(), key=lambda x: x[1]['amount'], reverse=True
        )[:10]

        # ── Timeline (secured funding per year) ───────────────────────
        timeline_data = []
        for year in range(year_from_int, year_to_int + 1):
            active_funding = active_count = 0
            for p in projects_deduped:
                sy = parse_year(p.start_date)
                ey = parse_year(p.end_date)
                if sy and ey and sy <= year <= ey:
                    active_funding += float(p.total_funding or 0)
                    active_count   += 1
            timeline_data.append({
                'year':            year,
                'secured_funding': active_funding,
                'active_grants':   active_count,
            })

        # ── Active grants detail table ────────────────────────────────
        active_grants_detail = []
        for p in projects_deduped:
            sy = parse_year(p.start_date)
            ey = parse_year(p.end_date)
            if not sy or not ey:
                continue
            if sy > year_to_int or ey < year_from_int:
                continue
            active_grants_detail.append({
                'title':          p.title,
                'organization':   p.funding_organization or '',
                'pi':             p.researcher.user.get_full_name() if p.researcher and p.researcher.user else '—',
                'role':           p.get_role_display() if p.role else '—',       
                'amount':         int(float(p.total_funding or 0)),
                'amount_to_ibme': int(float(p.funding_received or 0)) if p.funding_received else None,
                'status':         p.get_status_display() if p.status else '',
                'start_year':     sy,
                'end_year':       ey,
                'program':        p.program_name or '',
                'currency':       p.currency or 'CAD',
            })
        active_grants_detail.sort(key=lambda x: x['start_year'])

        # ── Last 3 years ──────────────────────────────────────────────
        last3_start = current_year - 3
        last3_projs = [
            p for p in projects_deduped
            if (parse_year(p.start_date) or 0) >= last3_start
            and p.status not in ('submitted', 'rejected') 
        ]

        PI_ROLES = ('pi', 'pa')

        # Role summary cards
        last3_as_pi  = {'amount': 0, 'count': 0}
        last3_as_coi = {'amount': 0, 'count': 0}
        for p in last3_projs:
            amt = float(p.total_funding or 0)
            if p.role in PI_ROLES:
                last3_as_pi['amount']  += amt
                last3_as_pi['count']   += 1
            else:
                last3_as_coi['amount'] += amt
                last3_as_coi['count']  += 1

        # By funding agency
        last3_org = defaultdict(lambda: {'amount': 0, 'count': 0, 'researchers': set()})
        for p in last3_projs:
            org  = p.funding_organization or 'Unknown'
            name = p.researcher.user.get_full_name() if p.researcher and p.researcher.user else 'Unknown'
            last3_org[org]['amount']      += float(p.total_funding or 0)
            last3_org[org]['count']       += 1
            last3_org[org]['researchers'].add(name)

        last3_by_org = sorted(
            [{'org': k, 'amount': v['amount'], 'count': v['count'],
              'researchers': len(v['researchers'])}
             for k, v in last3_org.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # By program
        last3_prog = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in last3_projs:
            prog = p.program_name or 'Unspecified'
            last3_prog[prog]['amount'] += float(p.total_funding or 0)
            last3_prog[prog]['count']  += 1

        last3_by_program = sorted(
            [{'program': k, 'amount': v['amount'], 'count': v['count']}
             for k, v in last3_prog.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # By researcher (with role split)
        last3_pi_d = defaultdict(lambda: {
            'amount': 0, 'count': 0, 'orgs': set(), 'pi_count': 0, 'coi_count': 0
        })
        for p in last3_projs:
            name = p.researcher.user.get_full_name() if p.researcher and p.researcher.user else 'Unknown'
            last3_pi_d[name]['amount'] += float(p.total_funding or 0)
            last3_pi_d[name]['count']  += 1
            last3_pi_d[name]['orgs'].add(p.funding_organization or 'Unknown')
            if p.role in PI_ROLES:
                last3_pi_d[name]['pi_count']  += 1
            else:
                last3_pi_d[name]['coi_count'] += 1

        last3_by_pi = sorted(
            [{'pi': k, 'amount': v['amount'], 'count': v['count'],
              'orgs': sorted(v['orgs']),
              'pi_count': v['pi_count'], 'coi_count': v['coi_count']}
             for k, v in last3_pi_d.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # ── Grants by PI (Funding records — correct per-researcher amounts) ──
        by_pi = defaultdict(lambda: {
            'amount': 0, 'amount_to_ibme': 0, 'amount_kept': 0, 'count': 0, 'grants': []
        })
        seen_project_ids = set()
        for f in funding_records:
            name = f.researcher.user.get_full_name() if f.researcher and f.researcher.user else 'Unknown'
            by_pi[name]['amount']         += float(f.amount or 0)
            by_pi[name]['amount_to_ibme'] += float(f.amount_to_ibme or 0)
            by_pi[name]['count']          += 1
            if f.project_id and f.project_id not in seen_project_ids:
                seen_project_ids.add(f.project_id)
                by_pi[name]['amount_kept'] += float(
                    f.project.funding_kept_by_unb or 0
                ) if f.project else 0
            by_pi[name]['grants'].append({
                'amount':         float(f.amount or 0),
                'amount_to_ibme': float(f.amount_to_ibme or 0),
                'grant_total':    float(f.grant_total or f.amount or 0),  # ← add
                'currency':       f.currency or 'CAD',
                'start':          parse_year(f.start_date),
                'end':            parse_year(f.end_date),
                'role':           f.role or '',
            })

        by_pi_list = sorted(
            [{'pi': k, 'amount': v['amount'], 'amount_to_ibme': v['amount_to_ibme'],
              'amount_kept': v['amount_kept'], 'count': v['count'], 'grants': v['grants']}
             for k, v in by_pi.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # ── Scholarships / Awards total for PI chart ─────────────────
        from django.db.models import Sum as DjangoSum

        award_qs = Recognition.objects.filter(
            amount__isnull=False, amount__gt=0,
            start_date__year__gte=year_from_int,
            start_date__year__lte=year_to_int,
            researcher__user__user_type='researcher',
        )
        award_total = float(award_qs.aggregate(t=DjangoSum('amount'))['t'] or 0)
        award_count = award_qs.count()

        schol_qs = Recognition.objects.filter(
            amount__isnull=False, amount__gt=0,
            start_date__year__gte=year_from_int,
            start_date__year__lte=year_to_int,
            researcher__user__user_type='student',
        )
        schol_total = float(schol_qs.aggregate(t=DjangoSum('amount'))['t'] or 0)
        schol_count = schol_qs.count()

        # ── Depletion waterfall ───────────────────────────────────────
        # Uses ALL deduped projects (not just currently active) so historical
        # snapshot rows are correct.
        # Linear drawdown model: remaining = amount × (end − t) / (end − start)
        # Shows 6 time-offsets per row: Y, Y+1, Y+2, Y+3, Y+4, Y+5

        DEPL_OFFSETS  = 6
        SPARK_H       = 28   # SVG bar height in px
        SPARK_BAR_W   = 16
        SPARK_GAP     = 4
        BAR_COLORS    = ['#C8102E', '#d63f63', '#e46e91', '#ed9ab5', '#f4c2d3', '#f9e3ea']

        all_proj_parsed = [
            {
                'amount': float(p.total_funding or 0),
                'start':  parse_year(p.start_date),
                'end':    parse_year(p.end_date),
            }
            for p in projects_deduped
            if parse_year(p.start_date) and parse_year(p.end_date)
        ]

        # First pass: compute all amounts so we can normalise sparklines
        raw_rows = []
        for snap_year in range(current_year - 4, current_year + 2):
            active_at_snap = [
                p for p in all_proj_parsed
                if p['start'] >= (current_year - 4)
                and p['start'] <= snap_year
                and p['end'] >= snap_year
            ]
            if not active_at_snap:
                continue

            base_amount = sum(p['amount'] for p in active_at_snap)
            years = []
            for offset in range(DEPL_OFFSETS):
                look_year = snap_year + offset
                amount    = 0.0
                count     = 0
                for p in active_at_snap:
                    if p['end'] < look_year:
                        continue
                    duration = p['end'] - p['start']
                    fraction = (
                        1.0 if duration <= 0
                        else max(0.0, min(1.0, (p['end'] - look_year) / duration))
                    )
                    if fraction > 0:
                        amount += p['amount'] * fraction
                        count  += 1
                years.append({
                    'year':        look_year,
                    'amount':      round(amount),
                    'grant_count': count,
                    'pct':         round(amount / base_amount * 100) if base_amount else 0,
                })
            raw_rows.append({
                'snapshot_year': snap_year,
                'base_amount':   base_amount,
                'years':         years,
            })

        depletion_max = max((r['base_amount'] for r in raw_rows), default=1)

        # Second pass: add SVG sparkline coordinates
        depletion_rows = []
        for row in raw_rows:
            for idx, y in enumerate(row['years']):
                bar_h        = max(2 if y['amount'] > 0 else 0,
                                   round(y['amount'] / depletion_max * SPARK_H))
                y['bar_x']   = idx * (SPARK_BAR_W + SPARK_GAP)
                y['bar_y']   = SPARK_H - bar_h
                y['bar_h']   = bar_h
                y['bar_w']   = SPARK_BAR_W
                y['bar_col'] = BAR_COLORS[min(idx, len(BAR_COLORS) - 1)]
            depletion_rows.append(row)

        spark_svg_w = DEPL_OFFSETS * (SPARK_BAR_W + SPARK_GAP) - SPARK_GAP

        return render(request, 'Pages/reports/funding_analysis_report.html', {
            'total_funding':        f"${total_funding_amount:,.0f}",
            'total_funding_raw':    total_funding_amount,
            'total_grants':         total_grants,
            'average_grant':        f"${average_grant:,.0f}",
            'by_organization':      dict(by_organization),
            'by_year':              dict(sorted(by_year.items())),
            'top_orgs':             top_orgs,
            'timeline_data':        timeline_data,
            'active_grants_detail': active_grants_detail,
            'min_year':             min_year,
            'max_year':             max_year,
            'year_from':            year_from_int,
            'year_to':              year_to_int,
            'current_year':         current_year,
            'last3_start':          last3_start,
            'last3_by_org':         last3_by_org,
            'last3_by_program':     last3_by_program,
            'last3_by_pi':          last3_by_pi,
            'last3_as_pi':          last3_as_pi,
            'last3_as_coi':         last3_as_coi,
            'by_pi_list':           by_pi_list,
            'award_total':          award_total,
            'award_count':          award_count,
            'schol_total':          schol_total,
            'schol_count':          schol_count,
            'depletion_rows':       depletion_rows,
            'depletion_max':        depletion_max,
            'spark_svg_w':          spark_svg_w,
            'spark_svg_h':          SPARK_H,
            'total_cad':            total_cad,
            'total_usd':            total_usd,
            'cad_count':            cad_count,
            'usd_count':            usd_count,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return render(request, 'Pages/reports/funding_analysis_report.html', {
            **EMPTY_CTX,
            'error': f'Error processing data: {e}',
        })

@login_required
@admin_required
def activity_report(request):
    from django.core.paginator import Paginator
    from django.db.models import Count, Q
    from collections import defaultdict

    current_year           = date.today().year
    current_month          = date.today().month
    current_academic_start = current_year if current_month >= 9 else current_year - 1

    year_options = [
        {'value': str(y), 'label': f"{y}/{str(y+1)[-2:]} (Sep {y} – Aug {y+1})"}
        for y in range(current_academic_start, current_academic_start - 6, -1)
    ]

    selected_year      = request.GET.get('year', str(current_academic_start))
    selected_category  = request.GET.get('category', '')
    selected_user_type = request.GET.get('user_type', '')
    selected_pub_type  = request.GET.get('pub_type', '')
    selected_status    = request.GET.get('grant_status', '')
    active_tab         = request.GET.get('tab', 'activities')
    group_by_conf      = request.GET.get('group') == '1'

    try:
        year_int = int(selected_year)
    except Exception:
        year_int = current_academic_start

    start_dt   = date(year_int, 9, 1)
    end_dt     = date(year_int + 1, 8, 31)
    year_label = f"{year_int}/{str(year_int+1)[-2:]} (Sep {year_int} – Aug {year_int+1})"

    # ── Base querysets (filters only, no select/prefetch) ────
    activities_base = Activity.objects.filter(
        date__gte=start_dt, date__lte=end_dt, is_active=True,
    )
    if selected_category:
        activities_base = activities_base.filter(category=selected_category)
    if selected_user_type:
        activities_base = activities_base.filter(researcher__user__user_type=selected_user_type)

    publications_base = Publication.objects.filter(
        publication_date__gte=start_dt, publication_date__lte=end_dt, is_active=True,
    )
    if selected_pub_type:
        publications_base = publications_base.filter(publication_type=selected_pub_type)

    grants_base = Project.objects.filter(
        start_date__gte=start_dt, start_date__lte=end_dt, is_deleted=False,
    )
    if selected_status:
        grants_base = grants_base.filter(status=selected_status)

    # ── Tab badge counts (cheap, single query each) ──────────
    total_count = activities_base.count()
    pub_count   = publications_base.count()
    grant_count = grants_base.count()

    # ── CSV export (early return, only the relevant tab) ────
    if request.GET.get('export') == 'csv':
        if active_tab == 'publications':
            qs = publications_base.select_related('researcher__user').order_by('-publication_date')
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="publications_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Title', 'Type', 'Status', 'Journal', 'Authors', 'Date', 'DOI', 'Researcher'])
            for p in qs:
                writer.writerow([
                    p.title, p.get_publication_type_display(),
                    p.get_status_display() if p.status else '',
                    p.journal or '', p.authors or '',
                    p.publication_date or '', p.doi or '',
                    p.researcher.user.get_full_name(),
                ])
            return response

        elif active_tab == 'grants':
            qs = grants_base.select_related('researcher__user').order_by('-start_date')
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="grants_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow([
                'Title', 'Organization', 'Program', 'Role', 'Status',
                'Total Funding', 'Awarded to UNB', 'Kept by UNB',
                'Start', 'End', 'PI',
            ])
            for g in qs:
                writer.writerow([
                    g.title, g.funding_organization or '', g.program_name or '',
                    g.get_role_display() if g.role else '',
                    g.get_status_display() if g.status else '',
                    g.total_funding or '', g.funding_received or '',
                    g.funding_kept_by_unb or '',
                    g.start_date or '', g.end_date or '',
                    g.researcher.user.get_full_name(),
                ])
            return response

        else:  # activities
            qs = activities_base.select_related('researcher__user').order_by('-date')
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="activities_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Title', 'Category', 'Type', 'Person', 'Email', 'Role', 'Date', 'Description'])
            for a in qs:
                writer.writerow([
                    a.title,
                    a.get_category_display() if hasattr(a, 'get_category_display') else a.category,
                    a.get_activity_type_display(),
                    a.researcher.user.get_full_name(),
                    a.researcher.user.email,
                    a.researcher.user.user_type.title(),
                    a.date.strftime('%Y-%m-%d') if a.date else '',
                    a.description or '',
                ])
            return response

    # ── Defaults so the template never KeyErrors ─────────────
    activities_page = publications_page = grants_page = None
    page_start = page_end = 0
    conference_count = km_count = media_count = 0
    researcher_count = student_count = 0
    awarded_count = submitted_count = rejected_count = 0
    objective_counts = []
    grouped_conferences = []

    # ── Heavy work ONLY for active tab ───────────────────────
    if active_tab == 'activities':
        # 6 counts → 1 aggregate query
        kpis = activities_base.aggregate(
            conf=Count('id', filter=Q(category='conference')),
            km=Count('id', filter=Q(category='knowledge_mobilization')),
            media=Count('id', filter=Q(category='media')),
            researchers=Count('researcher',
                filter=Q(researcher__user__user_type='researcher'), distinct=True),
            students=Count('researcher',
                filter=Q(researcher__user__user_type='student'), distinct=True),
        )
        conference_count = kpis['conf'] or 0
        km_count         = kpis['km'] or 0
        media_count      = kpis['media'] or 0
        researcher_count = kpis['researchers'] or 0
        student_count    = kpis['students'] or 0

        # N+1 loop → 1 annotated query
        objective_counts = list(
            StrategicObjective.objects
                .annotate(count=Count(
                    'activities',
                    filter=Q(
                        activities__date__gte=start_dt,
                        activities__date__lte=end_dt,
                        activities__is_active=True,
                    ),
                ))
                .filter(count__gt=0)
                .values('name', 'count')
        )

        if group_by_conf:
            conf_map = defaultdict(lambda: {
                'conference': None, 'name': '', 'year': None, 'location': '',
                'attendees': [], 'seen_users': set(),
                'student_count': 0, 'researcher_count': 0,
                'msc_count': 0, 'phd_count': 0, 'other_count': 0,
                'earliest_date': None,
            })

            conf_activities = activities_base.filter(
                category='conference'
            ).select_related(
                'researcher__user', 'researcher__user__student_profile', 'conference'
            )

            for a in conf_activities:
                key = f'conf-{a.conference_id}' if a.conference_id else f'no-conf-{a.title}-{a.date}'
                entry = conf_map[key]
                if a.conference and not entry['conference']:
                    entry['conference'] = a.conference
                    entry['name']       = a.conference.name
                    entry['year']       = a.conference.year
                    entry['location']   = a.conference.location or ''
                elif not entry['name']:
                    entry['name'] = a.title

                user_id = a.researcher.user.id
                if user_id not in entry['seen_users']:
                    entry['seen_users'].add(user_id)
                    utype = a.researcher.user.user_type
                    entry['attendees'].append({
                        'name': a.researcher.user.get_full_name(),
                        'user_type': utype,
                        'invited': a.invited, 'keynote': a.keynote,
                    })
                    if utype == 'student':
                        entry['student_count'] += 1
                        try:
                            degree = a.researcher.user.student_profile.degree_level
                            if degree == 'msc':   entry['msc_count'] += 1
                            elif degree == 'phd': entry['phd_count'] += 1
                            else:                 entry['other_count'] += 1
                        except Exception:
                            entry['other_count'] += 1
                    else:
                        entry['researcher_count'] += 1

                if not entry['earliest_date'] or (a.date and a.date < entry['earliest_date']):
                    entry['earliest_date'] = a.date

            grouped_conferences = sorted(
                conf_map.values(),
                key=lambda x: x['earliest_date'] or date.min,
                reverse=True,
            )
        else:
            activities_qs = activities_base.select_related(
                'researcher__user', 'conference'
            ).prefetch_related(
                'objectives', 'tagged_users'
            ).order_by('-date')

            paginator       = Paginator(activities_qs, 25)
            activities_page = paginator.get_page(request.GET.get('page', 1))
            page_start = (activities_page.number - 1) * 25 + 1
            page_end   = min(activities_page.number * 25, total_count)

    elif active_tab == 'publications':
        publications_qs = publications_base.select_related(
            'researcher__user'
        ).order_by('-publication_date')
        pub_paginator     = Paginator(publications_qs, 25)
        publications_page = pub_paginator.get_page(request.GET.get('pub_page', 1))

    elif active_tab == 'grants':
        # 4 counts → 1 aggregate query
        gkpi = grants_base.aggregate(
            awarded=Count('id', filter=Q(status='awarded')),
            submitted=Count('id', filter=Q(status='submitted')),
            rejected=Count('id', filter=Q(status='rejected')),
        )
        awarded_count   = gkpi['awarded'] or 0
        submitted_count = gkpi['submitted'] or 0
        rejected_count  = gkpi['rejected'] or 0

        grants_qs = grants_base.select_related(
            'researcher__user'
        ).prefetch_related('funding_breakdown').order_by('-start_date')
        grant_paginator = Paginator(grants_qs, 25)
        grants_page     = grant_paginator.get_page(request.GET.get('grant_page', 1))

    export_params = (
        f"year={selected_year}&category={selected_category}"
        f"&user_type={selected_user_type}&pub_type={selected_pub_type}"
        f"&grant_status={selected_status}&tab={active_tab}"
    )

    return render(request, 'Pages/reports/activity_report.html', {
        'activities':         activities_page,
        'publications':       publications_page,
        'grants':             grants_page,
        'year_options':       year_options,
        'selected_year':      selected_year,
        'selected_category':  selected_category,
        'selected_user_type': selected_user_type,
        'selected_pub_type':  selected_pub_type,
        'selected_status':    selected_status,
        'active_tab':         active_tab,
        'year_label':         year_label,
        'total_count':        total_count,
        'conference_count':   conference_count,
        'km_count':           km_count,
        'media_count':        media_count,
        'researcher_count':   researcher_count,
        'student_count':      student_count,
        'pub_count':          pub_count,
        'grant_count':        grant_count,
        'awarded_count':      awarded_count,
        'submitted_count':    submitted_count,
        'rejected_count':     rejected_count,
        'page_start':         page_start,
        'page_end':           page_end,
        'export_params':      export_params,
        'group_by_conf':      group_by_conf,
        'grouped_conferences': grouped_conferences,
        'objective_counts':   objective_counts,
    })


@login_required
@admin_required
def export_conference_equity_csv(request):
    students = StudentProfile.objects.select_related('user').filter(
        user__consent_to_share=True,
        degree_level__in=['msc', 'phd'],
    )
    count = students.count()

    if count < 5:
        messages.error(
            request,
            f"Only {count} consenting graduate student{'s' if count != 1 else ''} found. "
            "At least 5 are required to generate this EDI report and protect individual privacy."
        )
        return redirect('reports_list')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="conference_equity_report.csv"'
    writer = csv.writer(response)

    writer.writerow(['IBME Conference Equity Report'])
    writer.writerow([f'Generated: {date.today()}'])
    writer.writerow([f'Total consenting students: {count}'])
    writer.writerow([])

    # ── Section 1: Gender ────────────────────────────────────
    writer.writerow(['Section', 'Category', 'Count'])
    gender_map = {
        'female': 'Female', 'male': 'Male', 'non_binary': 'Non-binary',
        'self_describe': 'Prefer to self-describe', 'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    gender_counts = students.values('gender').annotate(n=Count('id'))
    for row in gender_counts:
        label = gender_map.get(row['gender'], row['gender'] or 'Not provided')
        writer.writerow(['Gender', label, row['n']])
    writer.writerow([])

    # ── Section 2: Indigenous identity ───────────────────────
    indigenous_map = {
        'yes': 'Yes — First Nations, Métis, or Inuit',
        'no': 'No',
        'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    indigenous_counts = students.values('indigenous_identity').annotate(n=Count('id'))
    for row in indigenous_counts:
        label = indigenous_map.get(row['indigenous_identity'], row['indigenous_identity'] or 'Not provided')
        writer.writerow(['Indigenous Identity', label, row['n']])
    writer.writerow([])

    # ── Section 3: Residency ─────────────────────────────────
    residency_map = {
        'citizen': 'Canadian Citizen', 'permanent_resident': 'Permanent Resident',
        'international': 'International Student', 'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    residency_counts = students.values('residency_status').annotate(n=Count('id'))
    for row in residency_counts:
        label = residency_map.get(row['residency_status'], row['residency_status'] or 'Not provided')
        writer.writerow(['Residency Status', label, row['n']])
    writer.writerow([])

    # ── Section 4: Degree level ──────────────────────────────
    degree_map = {'msc': 'MSc', 'phd': 'PhD'}
    degree_counts = students.values('degree_level').annotate(n=Count('id'))
    for row in degree_counts:
        label = degree_map.get(row['degree_level'], row['degree_level'])
        writer.writerow(['Degree Level', label, row['n']])
    writer.writerow([])

    # ── Annotate once for sections 5 & 6 ────────────────────
    current_year = date.today().year
    EQUITY_WINDOW_YEARS = 3
    window_start = current_year - EQUITY_WINDOW_YEARS

    students = students.annotate(
        conf_count=Count(
            'user__researcherprofile__activity',
            filter=Q(
                user__researcherprofile__activity__category='conference',
                user__researcherprofile__activity__is_active=True,
                user__researcherprofile__activity__date__year__gte=window_start,
            ),
            distinct=True,
        ),
        km_count=Count(
            'user__researcherprofile__activity',
            filter=Q(
                user__researcherprofile__activity__category='knowledge_mobilization',
                user__researcherprofile__activity__is_active=True,
                user__researcherprofile__activity__date__year__gte=window_start,
            ),
            distinct=True,
        ),
    )

    # ── Section 5: Conference attendance ────────────────────
    brackets = {'0 conferences': 0, '1–2 conferences': 0, '3–5 conferences': 0, '6+ conferences': 0}
    for s in students:
        c = s.conf_count  # ← fixed: no underscore, from annotation
        if c == 0:        brackets['0 conferences'] += 1
        elif c <= 2:      brackets['1–2 conferences'] += 1
        elif c <= 5:      brackets['3–5 conferences'] += 1
        else:             brackets['6+ conferences'] += 1

    for label, n in brackets.items():
        writer.writerow(['Conference Attendance (last 3 yrs)', label, n])
    writer.writerow([])

    # ── Section 6: KM activities ─────────────────────────────
    km_brackets = {'0 activities': 0, '1–2 activities': 0, '3–5 activities': 0, '6+ activities': 0}
    for s in students:
        k = s.km_count    # ← fixed: from annotation, no per-student query
        if k == 0:        km_brackets['0 activities'] += 1
        elif k <= 2:      km_brackets['1–2 activities'] += 1
        elif k <= 5:      km_brackets['3–5 activities'] += 1
        else:             km_brackets['6+ activities'] += 1

    for label, n in km_brackets.items():
        writer.writerow(['KM Activity Participation (last 3 yrs)', label, n])

    return response

@login_required
@admin_required
def conference_equity_report(request):
     
    conferences = Conference.objects.annotate(
        activity_count=Count('activities')
    ).filter(activity_count__gt=0).order_by('-year', 'name')

    selected_conf = None
    result        = None
    suppressed    = False
    MIN_THRESHOLD = 5

    conf_id = request.GET.get('conference_id')
    if conf_id:
        try:
            selected_conf = Conference.objects.get(id=conf_id)

            EQUITY_WINDOW_YEARS = 3
            window_start = date.today().replace(year=date.today().year - EQUITY_WINDOW_YEARS)

            activities = Activity.objects.filter(
                conference=selected_conf,
                is_active=True,
                date__gte=window_start,
            ).prefetch_related('tagged_users')

            # collect all participants
            participant_ids = set()
            for act in activities:
                participant_ids.add(act.researcher.user.id)
                for u in act.tagged_users.all():
                    participant_ids.add(u.id)

            profiles = StudentProfile.objects.filter(
                user__id__in=participant_ids,
                user__consent_to_share=True,
                degree_level__in=['msc', 'phd'],
            )

            count = profiles.count()

            if count < MIN_THRESHOLD:
                suppressed = True
            else:
                gender_map = {
                    'female': 'Female', 'male': 'Male',
                    'non_binary': 'Non-binary',
                    'self_describe': 'Prefer to self-describe',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                indigenous_map = {
                    'yes': 'Yes', 'no': 'No',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                residency_map = {
                    'citizen': 'Canadian Citizen',
                    'permanent_resident': 'Permanent Resident',
                    'international': 'International Student',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                degree_map = {'msc': 'MSc', 'phd': 'PhD'}

                def get_counts(field, label_map):
                    from collections import Counter
                    counts = Counter(getattr(p, field) for p in profiles)
                    return [
                        {'label': label_map.get(k, k or 'Not provided'), 'count': v}
                        for k, v in sorted(counts.items(), key=lambda x: -x[1])
                    ]

                race_map = {
                    'white':           'White / European descent',
                    'black':           'Black / African descent',
                    'east_asian':      'East Asian',
                    'south_asian':     'South Asian',
                    'southeast_asian': 'Southeast Asian',
                    'latin':           'Latin American / Hispanic',
                    'middle_eastern':  'Middle Eastern / North African',
                    'indigenous':      'Indigenous (other)',
                    'mixed':           'Two or more ethnicities',
                    'other':           'Other',
                    'prefer_not':      'Prefer not to say',
                    None:              'Not provided',
                }

                result = {
                    'total':          count,
                    'gender':         get_counts('gender', gender_map),
                    'indigenous':     get_counts('indigenous_identity', indigenous_map),
                    'residency':      get_counts('residency_status', residency_map),
                    'degree':         get_counts('degree_level', degree_map),
                    'race_ethnicity': get_counts('race_ethnicity', race_map),
                }

        except Conference.DoesNotExist:
            pass

    return render(request, 'Pages/reports/conference_equity_report.html', {
        'conferences':     conferences,
        'selected_conf':   selected_conf,
        'result':          result,
        'suppressed':      suppressed,
        'conf_id':         conf_id,
        'MIN_THRESHOLD':   MIN_THRESHOLD,
    })


@login_required
@admin_required
def conference_equity_summary(request):
    from collections import defaultdict
    from datetime import date, timedelta

    WINDOW_YEARS    = 3
    three_years_ago = date.today() - timedelta(days=WINDOW_YEARS * 365)

    filter_pi   = request.GET.get('pi', '')
    filter_dept = request.GET.get('dept', '')

    # ── Step 1: All conference activities in the 3yr window ──
    conf_activities = Activity.objects.filter(
        category='conference',
        is_active=True,
        date__gte=three_years_ago,
    ).prefetch_related('tagged_users').select_related('researcher__user')

    # ── Step 2: Build user_id → conference count ──────────────
    user_conf_count = defaultdict(int)

    for act in conf_activities:
        if act.researcher.user.user_type == 'student':
            user_conf_count[act.researcher.user_id] += 1
        for u in act.tagged_users.all():
            if u.user_type == 'student':
                user_conf_count[u.id] += 1

    # ── Step 3: Get supervision records with linked students ──
    sup_qs = SupervisionRecord.objects.filter(
        linked_student__isnull=False,
        status='in_progress',
    ).select_related(
        'researcher__user',
        'linked_student__user',
        'linked_student',
    )

    if filter_pi:
        sup_qs = sup_qs.filter(researcher__user__id=filter_pi)
    if filter_dept:
        sup_qs = sup_qs.filter(department__icontains=filter_dept)

    # ── Step 4: Group by PI + department ─────────────────────
    pi_dept_map = defaultdict(lambda: {
        'pi':           '',
        'pi_id':        None,
        'department':   '',
        'msc':          0,
        'phd':          0,
        'other':        0,
        'total':        0,
        'conf_total':   0,
        'students':     [],
        'students_set': set(),
    })

    degree_to_level = {
        'masters_thesis':     'msc',
        'masters_non_thesis': 'msc',
        'doctorate':          'phd',
    }

    for sup in sup_qs:
        dept  = (sup.department or '').strip() or 'Unknown'
        key   = f"{sup.researcher.user_id}|{dept.lower()}"
        entry = pi_dept_map[key]

        entry['pi']         = sup.researcher.user.get_full_name()
        entry['pi_id']      = sup.researcher.user_id
        entry['department'] = dept

        student    = sup.linked_student
        degree_lvl = degree_to_level.get(sup.degree_type, 'other')
        conf_count = user_conf_count.get(student.user_id, 0)

        if student.user_id in entry['students_set']:
            continue
        entry['students_set'].add(student.user_id)

        entry[degree_lvl]   += 1
        entry['total']      += 1
        entry['conf_total'] += conf_count

        entry['students'].append({
            'name':       student.user.get_full_name(),
            'degree':     sup.get_degree_type_display() or '—',
            'conf_count': conf_count,
        })

    # ── Step 5: Compute averages ──────────────────────────────
    rows = []

    for entry in pi_dept_map.values():
        total = entry['total']
        if not total:
            continue

        avg_conf          = round(entry['conf_total'] / total, 1)
        zero_conf         = sum(1 for s in entry['students'] if s['conf_count'] == 0)
        zero_conf_pct     = round(zero_conf / total * 100)
        participation_pct = 100 - zero_conf_pct

        rows.append({
            'pi':                entry['pi'],
            'pi_id':             entry['pi_id'],
            'department':        entry['department'],
            'msc':               entry['msc'],
            'phd':               entry['phd'],
            'other':             entry['other'],
            'total':             total,
            'avg_conf':          avg_conf,
            'conf_total':        entry['conf_total'],
            'zero_conf':         zero_conf,
            'zero_conf_pct':     zero_conf_pct,
            'participation_pct': participation_pct,
        })

    rows.sort(key=lambda x: x['avg_conf'], reverse=True)

    # ── Step 6: Overall summary + highlight below average ─────
    total_students = sum(r['total'] for r in rows)
    total_conf     = sum(r['conf_total'] for r in rows)
    overall_avg    = round(total_conf / total_students, 1) if total_students else 0

    for r in rows:
        r['below_avg'] = r['avg_conf'] < overall_avg

    # ── Dropdown options ──────────────────────────────────────
    all_departments = list(
        SupervisionRecord.objects.filter(
            department__isnull=False,
            linked_student__isnull=False,
        ).exclude(department='').values_list(
            'department', flat=True
        ).distinct().order_by('department')
    )

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="conference_equity_summary.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'PI', 'Department', 'MSc Students', 'PhD Students', 'Other',
            'Total Students', 'Total Conferences Attended',
            'Avg Conferences/Student', 'Students with 0 Conferences', '% with 0 Conferences',
        ])
        for r in rows:
            writer.writerow([
                r['pi'], r['department'],
                r['msc'], r['phd'], r['other'],
                r['total'], r['conf_total'], r['avg_conf'],
                r['zero_conf'], r['zero_conf_pct'],
            ])
        return response

    return render(request, 'Pages/reports/conference_equity_summary.html', {
        'rows':            rows,
        'total_students':  total_students,
        'total_conf':      total_conf,
        'overall_avg':     overall_avg,
        'WINDOW_YEARS':    WINDOW_YEARS,
        'filter_pi':       filter_pi,
        'filter_dept':     filter_dept,
        'all_supervisors': all_supervisors,
        'all_departments': all_departments,
    })
