import base64
import secrets
import string
import uuid

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import get_user_model, authenticate, login, logout, update_session_auth_hash
from django.contrib import messages
from django_ratelimit.decorators import ratelimit
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.contrib.auth.password_validation import validate_password
from django.core.validators import validate_email
import io
from django.db.models import Sum, Count, Q, Value
from django.views.decorators.http import require_http_methods
from django.views.decorators.http import require_POST
from functools import wraps
from defusedxml import ElementTree as ET
import csv
import json
import re
from datetime import datetime, date, timedelta, timezone
import time as time_module

from jsonschema import ValidationError
from .models import (
    CustomUser, Project, ProjectMember,
    ResearcherProfile, Education, Funding,
    Recognition, Activity, Publication,
    StudentProfile, ActivityReview, StudentNotification, SupervisorRequest, SupervisionRecord, StrategicObjective
)
from django.db import models

from django.http import JsonResponse

User = get_user_model()


# ─────────────────────────────────────────────
# Custom Decorators  ← MUST be at the top
# ─────────────────────────────────────────────

def admin_required(view_func):
    """Restrict view to users with user_type='admin'."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not getattr(request.user, 'user_type', None) == 'admin':
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper

def researcher_required(view_func):
    """Restrict view to users with user_type='researcher' or 'admin'."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if getattr(request.user, 'user_type', None) not in ('researcher', 'admin'):
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────
# Basic Pages
# ─────────────────────────────────────────────

def index_view(request):
    return render(request, "Pages/home.html")


@login_required
def log_activity_page(request):
    return render(request, 'Pages/forms/log_activity.html')


# ─────────────────────────────────────────────
# Signup
# ─────────────────────────────────────────────

@ratelimit(key='ip', rate='3/m', block=True, method=['POST'])
@require_http_methods(["GET", "POST"])
def signup_view(request):
    if request.method == "POST":
        name             = request.POST.get("name", "").strip()
        email            = request.POST.get("email", "").strip().lower()
        password         = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        consent          = request.POST.get("consent") == "true"

        try:
            validate_email(email)
        except ValidationError:
            time_module.sleep(1)
            return JsonResponse({"error": "Invalid email format."}, status=400)

        if not name or not email or not password or not confirm_password:
            time_module.sleep(1)
            return JsonResponse({"error": "All required fields must be provided."}, status=400)

        if password != confirm_password:
            time_module.sleep(1)
            return JsonResponse({"error": "Passwords do not match."}, status=400)

        if User.objects.filter(email=email).exists():
            time_module.sleep(1)
            return JsonResponse({"message": "If valid, your account will be created."})

        # Pending account cap
        PENDING_CAP = 20
        if CustomUser.objects.filter(approval_status='pending').count() >= PENDING_CAP:
            time_module.sleep(1)
            return JsonResponse({"message": "If valid, your account will be created."})

        try:
            validate_password(password)
        except ValidationError as e:
            time_module.sleep(1)
            return JsonResponse({"error": list(e.messages)}, status=400)

        username = f"{email.split('@')[0]}_{uuid.uuid4().hex[:6]}"
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            user_type="student",
            consent_to_share=consent,
        )

        parts = name.split()
        raw_first = parts[0] if parts else ''
        raw_last  = ' '.join(parts[1:]) if len(parts) > 1 else ''
        user.first_name = re.sub(r"[^a-zA-Z\s\-\']", '', raw_first).strip()[:100]
        user.last_name  = re.sub(r"[^a-zA-Z\s\-\']", '', raw_last).strip()[:100]
        user.save()

        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=f'{user.get_full_name()} ({user.email}) has registered and is awaiting approval.',
            )

        log_action(request, 'user_registered', target=user,
            summary=f'{user.get_full_name()} registered as {user.user_type}')

        return JsonResponse({
            "success": True,
            "message": "Account created. Awaiting admin approval."
        })

    return render(request, "Pages/User_Auth/signup.html")


# ─────────────────────────────────────────────
# Login
# ─────────────────────────────────────────────

@ratelimit(key='ip', rate='5/m', block=True, method=['POST'])
def login_view(request):
    if request.method == "POST":
        email    = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password")

        user = authenticate(request, email=email, password=password)

        if user is not None:
            if user.approval_status.lower() != "approved":
                time_module.sleep(1)
                log_action(request, 'login_pending', target=user,
                    summary=f'{user.get_full_name()} attempted login (pending approval)')
                return JsonResponse({
                    "error": "Your account is pending approval. Please wait for admin confirmation."
                }, status=403)

            from django.utils import timezone

            if user.temp_password_expires_at and user.temp_password_expires_at < timezone.now():
                time_module.sleep(1)
                log_action(request, 'login_failed',
                    summary=f'{user.get_full_name()} temp password expired')
                return JsonResponse({
                    "error": "Your temporary password has expired. Please contact an administrator."
                }, status=403)

            if user.force_password_change:
                login(request, user)
                return JsonResponse({"redirect": "/set-password/"})

            # ── Only redirect to 2FA if user has a CONFIRMED device ──
            from django_otp.plugins.otp_totp.models import TOTPDevice

            confirmed_devices = TOTPDevice.objects.filter(
                user=user,
                confirmed=True
            )

            if confirmed_devices.exists():
                request.session['pre_2fa_user_id'] = user.id
                request.session['pre_2fa_email']   = email
                return JsonResponse({"redirect": "/login/2fa/"})

            # No confirmed 2FA — force setup
            login(request, user)
            log_action(request, 'user_login', target=user,
                summary=f'{user.get_full_name()} logged in')
            return JsonResponse({"redirect": "/setup/2fa/"})

        time_module.sleep(1)
        log_action(request, 'login_failed',
            summary=f'Failed login attempt for {email}')
        return JsonResponse({"error": "Invalid email or password."}, status=401)

    return render(request, "Pages/User_Auth/login.html")


# ─────────────────────────────────────────────
# 2FA Verification
# ─────────────────────────────────────────────

def handler_403(request, exception=None):
    from django_ratelimit.exceptions import Ratelimited
    if isinstance(exception, Ratelimited):
        return render(request, 'Pages/User_Auth/ratelimited.html', status=429)
    return render(request, 'Pages/errors/403.html', status=403)

@ratelimit(key='ip', rate='10/m', block=True, method=['POST'])
def login_2fa_view(request):
    user_id = request.session.get('pre_2fa_user_id')
    if not user_id:
        return redirect('login')

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return redirect('login')

    error = None

    if request.method == 'POST':
        token = request.POST.get('token', '').strip().replace(' ', '')

        from django_otp import devices_for_user
        from django_otp.plugins.otp_static.models import StaticDevice, StaticToken

        verified = False

        # Check TOTP devices first
        for device in devices_for_user(user):
            if not getattr(device, 'confirmed', True):
                continue  # skip unconfirmed
            if device.verify_token(token):
                verified = True
                break

        # Check backup codes if TOTP failed
        if not verified:
            static_device = StaticDevice.objects.filter(
                user=user, name='backup'
            ).first()
            if static_device:
                static_token = StaticToken.objects.filter(
                    device=static_device,
                    token=token
                ).first()
                if static_token:
                    static_token.delete()  # one-time use
                    verified = True

        if verified:
            request.session.pop('pre_2fa_user_id', None)
            request.session.pop('pre_2fa_email', None)
            login(request, user)
            log_action(request, 'user_login', target=user,
                summary=f'{user.get_full_name()} logged in with 2FA')
            return redirect('dashboard')
        else:
            log_action(request, 'login_failed',
                summary=f'{user.get_full_name()} failed 2FA attempt')
            error = 'Invalid code. Please try again.'

    return render(request, 'Pages/User_Auth/login_2fa.html', {
        'error': error,
        'email': request.session.get('pre_2fa_email', ''),
    })



# ─────────────────────────────────────────────
# 2FA Setup
# ─────────────────────────────────────────────

@login_required
def setup_2fa_view(request):
    from django_otp.plugins.otp_totp.models import TOTPDevice
    from django_otp.plugins.otp_static.models import StaticDevice, StaticToken
    from django.utils import timezone
    import qrcode
    import secrets

    TOTP_SETUP_EXPIRY_MINUTES = 10

    # ── If already confirmed, redirect — don't show QR again ──
    existing_confirmed = TOTPDevice.objects.filter(
        user=request.user, confirmed=True
    ).first()
    if existing_confirmed:
        return redirect('dashboard')

    device, created = TOTPDevice.objects.get_or_create(
        user=request.user,
        name='default',
        defaults={'confirmed': False}
    )

    # ── Delete any extra devices ───────────────────────────────
    TOTPDevice.objects.filter(
        user=request.user
    ).exclude(id=device.id).delete()

    # ── Track when setup started ───────────────────────────────
    if created:
        request.user.totp_setup_started_at = timezone.now()
        request.user.save(update_fields=['totp_setup_started_at'])

    # ── Check if setup window has expired ─────────────────────
    if not device.confirmed and request.user.totp_setup_started_at:
        elapsed = timezone.now() - request.user.totp_setup_started_at
        if elapsed.total_seconds() > TOTP_SETUP_EXPIRY_MINUTES * 60:
            # Delete expired device — force fresh start
            device.delete()
            request.user.totp_setup_started_at = None
            request.user.save(update_fields=['totp_setup_started_at'])
            return render(request, 'Pages/User_Auth/setup_2fa.html', {
                'qr_b64':            None,
                'device':            None,
                'error':             'Your QR code expired after 10 minutes. Refresh the page to generate a new one.',
                'success':           None,
                'backup_codes':      [],
                'show_backup_codes': False,
                'expired':           True,
            })

    error             = None
    success           = None
    qr_b64            = None
    backup_codes      = []
    show_backup_codes = False

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'verify':
            token = request.POST.get('token', '').strip()
            if device.verify_token(token):
                device.confirmed = True
                device.save()

                # Clear setup timestamp
                request.user.totp_setup_started_at = None
                request.user.save(update_fields=['totp_setup_started_at'])

                success = '2FA enabled successfully.'
                log_action(request, 'other', target=request.user,
                    summary=f'{request.user.get_full_name()} enabled 2FA')

                static_device, _ = StaticDevice.objects.get_or_create(
                    user=request.user, name='backup'
                )
                backup_codes      = list(static_device.token_set.values_list('token', flat=True))
                show_backup_codes = True

                return render(request, 'Pages/User_Auth/setup_2fa.html', {
                    'qr_b64':            None,
                    'device':            device,
                    'error':             None,
                    'success':           success,
                    'backup_codes':      backup_codes,
                    'show_backup_codes': show_backup_codes,
                    'expired':           False,
                })
            else:
                error = 'Invalid code — please scan the QR code again and try.'
                log_action(request, 'login_failed', target=request.user,
                    summary=f'{request.user.get_full_name()} failed 2FA setup attempt')

    # ── Generate QR if not yet confirmed ──────────────────────
    if not device.confirmed:
        img = qrcode.make(device.config_url)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        qr_b64 = base64.b64encode(buf.getvalue()).decode()

        static_device, _ = StaticDevice.objects.get_or_create(
            user=request.user, name='backup'
        )
        if static_device.token_set.count() < 8:
            static_device.token_set.all().delete()
            for _ in range(8):
                StaticToken.objects.create(
                    device=static_device,
                    token=secrets.token_hex(4)
                )

    return render(request, 'Pages/User_Auth/setup_2fa.html', {
        'qr_b64':            qr_b64,
        'device':            device,
        'error':             error,
        'success':           success,
        'backup_codes':      backup_codes,
        'show_backup_codes': show_backup_codes,
        'expired':           False,
    })

@login_required
@require_http_methods(["POST"])
def api_reset_user_2fa(request, user_id):
    # ── Superuser only ─────────────────────────────────────
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)

    from django_otp.plugins.otp_totp.models import TOTPDevice
    from django_otp.plugins.otp_static.models import StaticDevice

    try:
        target_user = CustomUser.objects.get(id=user_id)
        TOTPDevice.objects.filter(user=target_user).delete()
        StaticDevice.objects.filter(user=target_user).delete()

        StudentNotification.objects.create(
            user=target_user,
            message='Your 2FA has been reset by an administrator. You will be asked to set it up again on next login.'
        )

        log_action(request, 'other', target=target_user,
            summary=f'{request.user.get_full_name()} reset 2FA for {target_user.get_full_name()}')

        return JsonResponse({'success': True})
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'}, status=404)

# ─────────────────────────────────────────────
# Logout
# ─────────────────────────────────────────────

def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect("login")


# ─────────────────────────────────────────────
# Force Password Change
# ─────────────────────────────────────────────

def generate_temp_password(length=12):
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    pwd = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%"),
    ]
    pwd += [secrets.choice(alphabet) for _ in range(length - 4)]
    secrets.SystemRandom().shuffle(pwd)
    return ''.join(pwd)


@login_required
@require_http_methods(["GET", "POST"])
def set_password(request):
    if not getattr(request.user, 'force_password_change', False):
        return redirect('dashboard')

    if request.method == 'GET':
        return render(request, 'Pages/User_Auth/set_password.html')

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    temp_pw = data.get('temp_password', '').strip()
    new_pw  = data.get('new_password', '').strip()
    conf_pw = data.get('confirm_password', '').strip()

    if not temp_pw:
        return JsonResponse({'success': False, 'error': 'Please enter your temporary password.'}, status=400)

    if not request.user.check_password(temp_pw):
        return JsonResponse({'success': False, 'error': 'Temporary password is incorrect.'}, status=400)

    if new_pw != conf_pw:
        return JsonResponse({'success': False, 'error': 'New passwords do not match.'}, status=400)

    if new_pw == temp_pw:
        return JsonResponse({'success': False, 'error': 'New password cannot be the same as your temporary password.'}, status=400)

    try:
        validate_password(new_pw, user=request.user)
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': list(e.messages)}, status=400)

    request.user.set_password(new_pw)
    request.user.force_password_change = False
    request.user.temp_password_expires_at = None
    request.user.save()

    update_session_auth_hash(request, request.user)

    StudentNotification.objects.create(
        user=request.user,
        message='Your password was changed successfully. If this wasn\'t you, contact your administrator immediately.'
    )

    log_action(request, 'password_changed', target=request.user,
        summary=f'{request.user.get_full_name()} set their password for the first time')

    return JsonResponse({'success': True, 'redirect': '/dashboard/'})


# ─────────────────────────────────────────────
# Change Password
# ─────────────────────────────────────────────

@login_required
@require_http_methods(["POST"])
def api_change_password(request):
    try:
        data       = json.loads(request.body)
        current_pw = data.get('current_password', '').strip()
        new_pw     = data.get('new_password', '').strip()
        conf_pw    = data.get('confirm_password', '').strip()

        if not request.user.check_password(current_pw):
            return JsonResponse({'success': False, 'error': 'Current password is incorrect.'}, status=400)

        if new_pw != conf_pw:
            return JsonResponse({'success': False, 'error': 'New passwords do not match.'}, status=400)

        if new_pw == current_pw:
            return JsonResponse({'success': False, 'error': 'New password cannot be the same as your current password.'}, status=400)

        try:
            validate_password(new_pw, user=request.user)
        except ValidationError as e:
            return JsonResponse({'success': False, 'error': list(e.messages)}, status=400)

        request.user.set_password(new_pw)
        request.user.save()

        StudentNotification.objects.create(
            user=request.user,
            message='Your password was changed. If this wasn\'t you, contact your administrator immediately.'
        )

        update_session_auth_hash(request, request.user)

        log_action(request, 'password_changed', target=request.user,
            summary=f'{request.user.get_full_name()} changed their password')

        return JsonResponse({'success': True})

    except Exception:
        return JsonResponse({'success': False, 'error': 'Internal server error.'}, status=500)


# ─────────────────────────────────────────────
# Admin Create User
# ─────────────────────────────────────────────

@login_required
@admin_required
@require_http_methods(["POST"])
def api_create_user_with_temp_password(request):
    try:
        data       = json.loads(request.body)
        first_name = data.get('first_name', '').strip()
        last_name  = data.get('last_name', '').strip()
        email      = data.get('email', '').strip().lower()
        user_type  = data.get('user_type', '').strip()

        if not all([first_name, last_name, email, user_type]):
            return JsonResponse({'success': False, 'error': 'All fields are required.'}, status=400)

        try:
            validate_email(email)
        except ValidationError:
            return JsonResponse({'success': False, 'error': 'Invalid email format.'}, status=400)

        if user_type not in ('researcher', 'student'):
            return JsonResponse({'success': False, 'error': 'Invalid user type.'}, status=400)

        if CustomUser.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'error': 'A user with this email already exists.'}, status=400)

        temp_password = generate_temp_password()
        username = f"{first_name.lower()}.{last_name.lower()}.{uuid.uuid4().hex[:6]}".replace(' ', '')

        from django.utils import timezone
        from datetime import timedelta

        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            password=temp_password,
            user_type=user_type,
            approval_status='approved',
            force_password_change=True,
        )

        user.temp_password_expires_at = timezone.now() + timedelta(hours=24)
        user.save()

        if user_type == 'researcher':
            ResearcherProfile.objects.get_or_create(user=user, defaults={'research_interests': ''})
        elif user_type == 'student':
            StudentProfile.objects.get_or_create(user=user)

        log_action(request, 'user_created_by_admin', target=user,
            summary=f'{request.user.get_full_name()} created account for {user.get_full_name()} ({email})')

        return JsonResponse({
            'success': True,
            'message': f'Account created for {first_name} {last_name}.',
            'temp_password': temp_password,
            'email': email,
        })

    except Exception:
        return JsonResponse({'success': False, 'error': 'Internal server error.'}, status=500)

# ─────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────

@never_cache
@login_required(login_url="login")
def dashboard_view(request):
    from collections import defaultdict
    from django.db.models.functions import ExtractYear
    from datetime import timedelta
    import json

    context = {"user": request.user}

    if request.user.approval_status != "approved":
        return redirect("login")

    # ─────────────────────────────────────────────
    # Admin
    # ─────────────────────────────────────────────
    if request.user.user_type == 'admin':
        context['total_researchers'] = CustomUser.objects.filter(
            user_type='researcher', approval_status='approved'
        ).count()
        context['total_students'] = CustomUser.objects.filter(
            user_type='student', approval_status='approved'
        ).count()
        context['total_admins'] = CustomUser.objects.filter(
            user_type='admin'
        ).count()

        active_funding_filter = Q(status='awarded') & (
            Q(end_date__gte=date.today()) | Q(end_date__isnull=True)
        )
        funding_agg = Funding.objects.filter(active_funding_filter).aggregate(
            total=Sum('amount'),
            count=Count('id'),
        )
        context['total_funding']           = funding_agg['total'] or 0
        context['total_funding_formatted'] = f"{int(context['total_funding']):,}"
        context['total_grants']            = funding_agg['count'] or 0
        context['total_education']         = Education.objects.count()

        context['researcher_funding'] = ResearcherProfile.objects.annotate(
            total=Sum('funding__amount')
        ).filter(funding__isnull=False).values(
            'user__first_name', 'user__last_name', 'total'
        )[:10]

        context['org_stats'] = (
            Funding.objects.values('organization')
            .annotate(count=Count('id')).order_by('-count')[:10]
        )
        return render(request, "Pages/dashboard.html", context)

    # ─────────────────────────────────────────────
    # Student
    # ─────────────────────────────────────────────
    elif request.user.user_type == 'student':
        MIN_PEER_THRESHOLD = 2
        three_years_ago    = date.today() - timedelta(days=3 * 365)

        student, _ = StudentProfile.objects.get_or_create(user=request.user)
        researcher = ResearcherProfile.objects.filter(user=request.user).first()

        context['student']        = student
        context['student_profile'] = student
        context['researcher']     = researcher
        context['today']          = date.today()
        context['student_education'] = Education.objects.filter(
            researcher=researcher
        ).order_by('-expected_date') if researcher else []

        if researcher:

            activities_list = list(
                Activity.objects.filter(
                    researcher=researcher
                ).order_by('-date')
            )
            context['activity_count']    = len(activities_list)
            context['recent_activities'] = activities_list[:6]
            context['conference_count']  = sum(
                1 for a in activities_list if a.category == 'conference'    
            )
            context['km_count'] = sum(
                1 for a in activities_list if a.category == 'knowledge_mobilization'
            )

            publications_list = list(
                Publication.objects.filter(
                    researcher=researcher
                ).order_by('-publication_date')
            )
            context['publication_count']   = len(publications_list)
            context['recent_publications'] = publications_list[:6]
        else:
            context.update({
                'activity_count': 0, 'recent_activities': [],
                'conference_count': 0, 'km_count': 0,
                'publication_count': 0, 'recent_publications': [],
            })

        # ── Peer stats ────────────────────────────────────────────────
        if student.supervisor:
            peer_students = CustomUser.objects.filter(
                user_type='student',
                student_profile__supervisor=student.supervisor,
            ).exclude(id=request.user.id)
        else:
            peer_students = CustomUser.objects.none()

        peer_student_count = peer_students.count()
        context['peer_student_count'] = peer_student_count
        context['peer_supervisor'] = (
            student.supervisor.user.get_full_name() if student.supervisor else None
        )

        if peer_student_count >= MIN_PEER_THRESHOLD:
            student_researchers = ResearcherProfile.objects.filter(
                user__in=peer_students
            )

            # Single query — compute all peer stats in Python
            peer_activities = list(
                Activity.objects.filter(
                    researcher__in=student_researchers,
                    date__gte=three_years_ago,
                    category__in=['conference', 'knowledge_mobilization'],
                ).values('category', 'date')
            )

            context['peer_conference_total'] = sum(
                1 for a in peer_activities if a['category'] == 'conference'
            )
            context['peer_km_total'] = sum(
                1 for a in peer_activities if a['category'] == 'knowledge_mobilization'
            )

            year_counts = defaultdict(int)
            for a in peer_activities:
                if a['category'] == 'conference' and a['date']:
                    year_counts[a['date'].year] += 1

            context['peer_by_year'] = json.dumps([
                {'year': y, 'count': c}
                for y, c in sorted(year_counts.items())
            ])
            context['peer_stats_available'] = True

        else:
            context.update({
                'peer_conference_total': None,
                'peer_km_total':         None,
                'peer_by_year':          '[]',
                'peer_stats_available':  False,
                'peer_stats_hidden_reason': (
                    'Peer statistics require at least 2 other students under your supervisor.'
                    if student.supervisor
                    else 'Peer statistics are available once you are assigned a supervisor.'
                ),
            })

        return render(request, "Pages/student_dashboard.html", context)

    # ─────────────────────────────────────────────
    # Researcher
    # ─────────────────────────────────────────────
    else:
        context['researcher_funding'] = []
        context['org_stats']          = []

        researcher = ResearcherProfile.objects.filter(user=request.user).first()

        if researcher:
            three_years_ago = date.today() - timedelta(days=3 * 365)

            context['researcher'] = researcher

            context['publication_count'] = Publication.objects.filter(
                researcher=researcher, is_active=True, is_deleted=False
            ).count()

            context['projects_count'] = Project.objects.filter(
                researcher=researcher, is_deleted=False
            ).count()

            context['recent_publications'] = list(
                Publication.objects.filter(
                    researcher=researcher, is_active=True
                ).order_by('-publication_date')[:6]
            )

            # Single aggregation for active funding
            context['total_funding'] = Funding.objects.filter(
                researcher=researcher,
                status='awarded',
            ).filter(
                Q(end_date__gte=date.today()) | Q(end_date__isnull=True)
            ).aggregate(total=Sum('amount'))['total'] or 0
            context['total_funding_formatted'] = f"{int(context['total_funding']):,}"

            context['recent_activities'] = list(
                Activity.objects.filter(
                    researcher=researcher
                ).order_by('-date')[:5]
            )

            # Funding by year
            funding_by_year = list(
                Funding.objects.filter(
                    researcher=researcher,
                ).annotate(year=ExtractYear('start_date'))
                .values('year')
                .annotate(total=Sum('amount'))
                .order_by('year')
            )
            context['funding_by_year'] = [
                {'year': r['year'], 'total': float(r['total'] or 0)}
                for r in funding_by_year
            ]

            # Top funding organizations
            context['top_orgs'] = list(
                Funding.objects.filter(researcher=researcher)
                .values('organization')
                .annotate(total=Sum('amount'))
                .order_by('-total')[:5]
            )

            # Active projects — prefetch to avoid per-project queries
            active_qs = Project.objects.active().filter(
                researcher=researcher
            ).order_by('-start_date')

            active_projects_list = []
            for p in active_qs:
                p.received_percent = (
                    (p.funding_received / p.total_funding * 100)
                    if p.total_funding and p.total_funding > 0
                    and p.funding_received is not None
                    else 0
                )
                active_projects_list.append(p)

            context['active_projects'] = active_projects_list[:3]

            context['supervising_count'] = SupervisionRecord.objects.filter(
                researcher=researcher,
                status='in_progress',
            ).count()

            context['education_records'] = list(
                Education.objects.filter(
                    researcher=researcher
                ).order_by('-expected_date')[:6]
            )

            context['all_award_records'] = list(
                Recognition.objects.filter(
                    researcher=researcher
                ).order_by('-start_date')[:6]
            )

        return render(request, "Pages/dashboard.html", context)


# ─────────────────────────────────────────────
# XML Parsing Helpers
# ─────────────────────────────────────────────

def extract_field_value(section, label_name):

    for field in section.findall('field'):

        if field.get('label') != label_name:
            continue

        val = field.findtext('value')
        if val and val.strip():
            return val.strip()

        lov = field.findtext('lov')
        if lov and lov.strip():
            return lov.strip()

        lov_text = field.findtext('lov/text')
        if lov_text and lov_text.strip():
            return lov_text.strip()

        ref = field.find('refTable')
        if ref is not None:
            for link in ref.findall('linkedWith'):
                value = link.get('value')
                if value:
                    return value

    return None


def extract_organization(field):
    ref = field.find('refTable')
    if ref is None:
        return None

    for link in ref.findall('linkedWith'):
        if link.get('label') == 'Organization':
            return link.get('value')

    return None

def extract_department_for_org(xml_root, target_org=None):
    """
    Finds the department associated with the given organization.
    If target_org is None, returns the first department found.
    """

    fallback = None

    for section in xml_root.findall('.//section'):

        org = None
        dept = None

        for field in section.findall('field'):

            label = field.get('label')

            if label == "Organization":
                org = extract_organization(field)

            elif label == "Department":
                dept = field.findtext('value')

        if dept:
            if not fallback:
                fallback = dept

            if target_org and org and target_org.lower() in org.lower():
                return dept

    return fallback

def parse_date_from_yearmonth(date_str):
    """Convert yyyy/MM format to date object."""
    if not date_str:
        return None
    try:
        parts = date_str.split('/')
        year  = int(parts[0])
        month = int(parts[1]) if len(parts) > 1 else 1
        return datetime(year, month, 1).date()
    except Exception:
        return None


def parse_funding_date(date_str):
    """Parse funding date in various CCV formats → date object."""
    if not date_str:
        return None
    date_str = str(date_str).strip()
    try:
        if '/' in date_str:
            parts = date_str.split('/')
            year  = int(parts[0])
            month = int(parts[1].zfill(2)) if len(parts) > 1 else 1
            return datetime(year, month, 1).date()
        elif len(date_str) == 10 and date_str.count('-') == 2:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        elif len(date_str) == 4 and date_str.isdigit():
            return datetime(int(date_str), 1, 1).date()
    except Exception as e:
        print(f"Error parsing date '{date_str}': {e}")
    return None


def extract_bilingual_value(field_elem):
    if field_elem is None:
        return None
    value_elem = field_elem.find('value')
    if value_elem is not None and value_elem.text:
        return value_elem.text
    bilingual = field_elem.find('bilingual/english')
    if bilingual is not None and bilingual.text:
        return bilingual.text
    return None


def map_funding_role_to_choice(role_str):
    if not role_str:
        return 'other'
    r = role_str.lower()
    if 'principal investigator' in r: return 'pi'
    if 'co-investigator' in r:        return 'co_pi'
    if 'principal applicant' in r:    return 'pa'
    if 'co-applicant' in r:           return 'co_app'
    return 'other'


def determine_activity_category(title, topic, audience, event):
    text = f"{title or ''} {topic or ''} {audience or ''} {event or ''}".lower()
    
    # ── Conference checked FIRST — before workshop/seminar ───
    if any(w in text for w in ['conference', 'symposium', 'congress', 'summit']):
        return 'conference'
    if any(w in text for w in ['outreach', 'community', 'knowledge mobilization',
                                'public engagement', 'workshop', 'seminar',
                                'disabilities', 'partnership', 'gathering']):
        return 'knowledge_mobilization'
    if any(w in text for w in ['broadcast', 'interview', 'media', 'cbc',
                                'global news', 'radio', 'telegraph', 'news']):
        return 'media'
    if any(w in text for w in ['university', 'academic', 'research', 'lecture', 'presentation']):
        return 'academic'
    # ── meeting kept here — avoids "meeting" matching conference ─
    if 'meeting' in text:
        return 'conference'
    return 'conference' if event else 'other'

# ───────────────────────────────────────────────────────────────────────────────
# HELPER — get recordId with fallback
# ───────────────────────────────────────────────────────────────────────────────

def get_record_id(section, fallback_parts):
    """
    Returns section's recordId if present.
    Falls back to a normalised string from fallback_parts if not.
    This handles edge cases where CCV omits recordId on some sections.
    """
    record_id = section.get('recordId')
    if record_id:
        return record_id
    # Fallback: join non-empty parts, normalise
    import re
    combined = '_'.join(str(p) for p in fallback_parts if p)
    return 'fallback:' + re.sub(r'[^a-z0-9_]', '', combined.lower())[:120]


# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_education
# ───────────────────────────────────────────────────────────────────────────────

def parse_xml_education(researcher, xml_root):
    seen_ids = set()
    count_created = count_updated = 0

    for edu_section in xml_root.findall('.//section[@label="Education"]'):
        for degree_section in edu_section.findall('section[@label="Degrees"]'):
            degree_type    = extract_field_value(degree_section, 'Degree Type')
            specialization = extract_field_value(degree_section, 'Specialization')

            # FIX: Extract Degree Name as fallback for specialization
            degree_name = extract_field_value(degree_section, 'Degree Name')

            organization = None
            org_field    = degree_section.find('field[@label="Organization"]')
            if org_field is not None:
                reftable = org_field.find('refTable')
                if reftable is not None:
                    linked = reftable.find('linkedWith[@label="Organization"]')
                    if linked is not None:
                        organization = linked.get('value')
            if not organization:
                organization = extract_field_value(degree_section, 'Other Organization')

            completion_date = extract_field_value(degree_section, 'Degree Received Date')
            if not degree_type:
                continue

            # FIX: Extract Degree Start Date → start_date
            start_date_str = extract_field_value(degree_section, 'Degree Start Date')
            start_date = parse_date_from_yearmonth(start_date_str) if start_date_str else None

            # FIX: Try Degree Received Date first, then Degree Expected Date
            expected_date = None
            if completion_date:
                try:
                    expected_date = (
                        parse_date_from_yearmonth(completion_date)
                        if '/' in completion_date
                        else (datetime.strptime(f"{completion_date}-01-01", "%Y-%m-%d").date()
                              if len(completion_date) == 4 else None)
                    )
                except Exception:
                    pass

            if not expected_date:
                # FIX: Fallback to Degree Expected Date from schema
                expected_date_str = extract_field_value(degree_section, 'Degree Expected Date')
                if expected_date_str:
                    expected_date = parse_date_from_yearmonth(expected_date_str)

            # FIX: Extract Thesis Title
            thesis_title = extract_field_value(degree_section, 'Thesis Title') or ''

            ext_id = get_record_id(degree_section, [degree_type, organization, completion_date])
            seen_ids.add(ext_id)

            

            fields = dict(
                degree_type    = degree_type[:100],
                specialization = specialization or degree_name or '',
                institution    = organization or 'Unknown',
                thesis_title   = thesis_title,                      
                start_date     = start_date,           
                expected_date  = expected_date,
            )

            existing = Education.objects.filter(
                researcher=researcher,
                external_id=ext_id
            ).first()

            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
                existing.save()
                count_updated += 1
            else:
                Education.objects.create(
                    researcher=researcher, external_id=ext_id, **fields
                )
                count_created += 1


    Education.objects.filter(
        researcher=researcher
    ).exclude(
        external_id__in=seen_ids
    ).delete()

    return count_created + count_updated


# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_funding
# ───────────────────────────────────────────────────────────────────────────────


def parse_xml_funding(researcher, xml_root):
    seen_ids = set()
    count_created = count_updated = 0

    for funding_section in xml_root.findall('.//section[@label="Research Funding History"]'):
        status_val = extract_field_value(funding_section, 'Funding Status') or ''
        if status_val == 'Under Review':
            continue

        s = status_val.lower()
        if 'award' in s or 'funded' in s or 'completed' in s:
            inferred_status = 'awarded'
        elif 'rejected' in s or 'unsuccessful' in s or 'declined' in s:
            inferred_status = 'rejected'
        elif 'submitted' in s or 'review' in s or 'pending' in s:
            inferred_status = 'submitted'
        else:
            inferred_status = 'awarded'

        title = extract_field_value(funding_section, 'Funding Title')
        if not title:
            continue

        # ── Resolve linked Project (FK) ──────────────────────────────
        parent_rid = funding_section.get('recordId', '')
        linked_project = Project.all_objects.filter(
            researcher=researcher,
            external_id=parent_rid,
        ).first()

        funding_type = extract_field_value(funding_section, 'Funding Type')
        funding_role = map_funding_role_to_choice(
            extract_field_value(funding_section, 'Funding Role')
        )

        # ── Extract org, program, currency from first Funding Source ─
        org = program_name = None
        currency = 'CAD'

        for source_section in funding_section.findall('section[@label="Funding Sources"]'):
            org = (extract_field_value(source_section, 'Funding Organization') or
                   extract_field_value(source_section, 'Other Funding Organization'))
            program_name = extract_field_value(source_section, 'Program Name')
            currency_str = extract_field_value(source_section, 'Currency of Total Funding') or ''
            currency = 'USD' if 'united states' in currency_str.lower() else 'CAD'
            break

        if not org:
            continue

        # ── Year records branch ───────────────────────────────────────
        # FIX: each year uses its OWN portion only — no dumping full grant
        # total into year-1 IBME. For Co-I/other roles, amount = year portion.
        year_sections = funding_section.findall('section[@label="Funding by Year"]')

        if year_sections:
            for ys in year_sections:
                total = extract_field_value(ys, 'Total Funding')
                try:
                    amount = float(total) if total else 0
                except Exception:
                    amount = 0
                if amount <= 0:
                    continue

                year_portion = extract_field_value(ys, 'Portion of Funding Received')
                try:
                    year_portion_val = float(year_portion) if year_portion else None
                    if year_portion_val == 0:
                        year_portion_val = None
                except Exception:
                    year_portion_val = None

                # FIX: for Co-I/other roles use portion as amount
                # so full project annual budget isn't attributed to this researcher
                co_i_roles = ('co_pi', 'co_app', 'other', 'collaborator')
                if funding_role in co_i_roles:
                    amount = year_portion_val or 0

                # ibme = year's own portion only, never full grant total
                ibme_val = year_portion_val  # None is correct when not recorded

                start = parse_funding_date(extract_field_value(ys, 'Start Date'))
                end   = parse_funding_date(extract_field_value(ys, 'End Date'))

                ext_id = get_record_id(ys, [funding_section.get('recordId', ''), str(start)])
                seen_ids.add(ext_id)

                fields = dict(
                    title=title, organization=org, funding_type=funding_type,
                    program_name=program_name, amount=amount,
                    amount_to_ibme=ibme_val,
                    start_date=start, end_date=end,
                    role=funding_role, status=inferred_status,
                    project=linked_project,
                    currency=currency,
                )

                existing = Funding.all_objects.filter(
                    researcher=researcher, external_id=ext_id
                ).first()

                if existing:
                    if existing.is_deleted:
                        existing.is_deleted = False
                        existing.deleted_at = None
                    for k, v in fields.items():
                        setattr(existing, k, v)
                    existing.save()
                    count_updated += 1
                else:
                    Funding.objects.create(
                        researcher=researcher, external_id=ext_id, **fields
                    )
                    count_created += 1

        else:
            # ── No year records — one record per Funding Source ───────
            # FIX: process ALL sources (removed break), each gets its own
            # Funding row using source recordId for dedup.
            # FIX: amount = Portion of Funding Received, not Total Funding.
            # Explicit 0 (Co-I, no budget) is kept as 0; blank = use total.
            # grant_total = always the XML Total Funding (for blue bar in chart)
            for source_section in funding_section.findall('section[@label="Funding Sources"]'):
                src_org = (extract_field_value(source_section, 'Funding Organization') or
                           extract_field_value(source_section, 'Other Funding Organization'))
                if not src_org:
                    continue

                src_program  = extract_field_value(source_section, 'Program Name') or program_name
                portion_str  = extract_field_value(source_section, 'Portion of Funding Received')
                total_str    = extract_field_value(source_section, 'Total Funding')
                currency_str = extract_field_value(source_section, 'Currency of Total Funding') or ''
                src_currency = 'USD' if 'united states' in currency_str.lower() else 'CAD'

                # ── grant_total: always the XML Total Funding value ───
                # Used for the "Total Grant" blue bar in the PI chart.
                # Stored separately from amount so we never lose the full
                # project value even when researcher's portion is $0.
                try:
                    grant_total = float(total_str) if total_str else 0
                except Exception:
                    grant_total = 0

                # ── amount: researcher's actual allocated portion ──────
                # Distinguish "explicitly 0" (Co-I, no budget) from "not recorded"
                try:
                    portion_val = float(portion_str) if (
                        portion_str is not None and str(portion_str).strip() != ''
                    ) else None
                except Exception:
                    portion_val = None

                if portion_val is not None:
                    # Explicitly recorded — use as-is (0 for Co-I is correct)
                    amount = portion_val
                else:
                    # Not recorded — researcher owns full grant, use total
                    amount = grant_total

                # Skip only when nothing at all is recorded
                if amount <= 0 and portion_val is None and grant_total <= 0:
                    continue  # not break — check remaining sources

                # amount_to_ibme = portion if > 0, otherwise None
                ibme_amount = portion_val if (portion_val and portion_val > 0) else None

                src_start = parse_funding_date(
                    extract_field_value(source_section, 'Funding Start Date') or
                    extract_field_value(funding_section, 'Funding Start Date')
                )
                src_end = parse_funding_date(
                    extract_field_value(source_section, 'Funding End Date') or
                    extract_field_value(funding_section, 'Funding End Date')
                )

                # Use source recordId so each source gets its own Funding row
                src_record_id = source_section.get('recordId', '')
                ext_id = get_record_id(
                    source_section, [title, src_org, str(src_start), src_record_id]
                )
                seen_ids.add(ext_id)

                fields = dict(
                    title=title, organization=src_org, funding_type=funding_type,
                    program_name=src_program,
                    amount=amount,
                    amount_to_ibme=ibme_amount,
                    grant_total=grant_total,
                    start_date=src_start, end_date=src_end,
                    role=funding_role, status=inferred_status,
                    project=linked_project,
                    currency=src_currency,
                )

                existing = Funding.all_objects.filter(
                    researcher=researcher, external_id=ext_id
                ).first()

                if existing:
                    if existing.is_deleted:
                        existing.is_deleted = False
                        existing.deleted_at = None
                    for k, v in fields.items():
                        setattr(existing, k, v)
                    existing.save()
                    count_updated += 1
                else:
                    Funding.objects.create(
                        researcher=researcher, external_id=ext_id, **fields
                    )
                    count_created += 1

    from django.utils import timezone
    Funding.objects.filter(
        researcher=researcher
    ).exclude(
        external_id__in=seen_ids
    ).update(is_deleted=True, deleted_at=timezone.now())

    return count_created + count_updated

# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_recognitions
# ───────────────────────────────────────────────────────────────────────────────

def parse_xml_recognitions(researcher, xml_root):
    seen_ids = set()
    count_created = count_updated = 0

    for recog_section in xml_root.findall('.//section[@label="Recognitions"]'):
        award_name = extract_field_value(recog_section, 'Recognition Name')
        if not award_name:
            continue

        amount_str = extract_field_value(recog_section, 'Amount')
        amount = None
        if amount_str:
            try:
                amount = float(amount_str)
            except Exception:
                pass

        start_date = parse_funding_date(extract_field_value(recog_section, 'Effective Date'))

        # FIX: Extract End Date
        end_date = parse_funding_date(extract_field_value(recog_section, 'End Date'))

        # FIX: Extract Recognition Type
        recognition_type = extract_field_value(recog_section, 'Recognition Type')

        # FIX: Extract primary Organization via refTable, fallback to Other Organization
        organization = None
        org_field = recog_section.find('field[@label="Organization"]')
        if org_field is not None:
            reftable = org_field.find('refTable')
            if reftable is not None:
                linked = reftable.find('linkedWith[@label="Organization"]')
                if linked is not None:
                    organization = linked.get('value')
        if not organization:
            organization = extract_field_value(recog_section, 'Other Organization') or ''

        ext_id = get_record_id(recog_section, [award_name, str(start_date)])
        seen_ids.add(ext_id)

        fields = dict(
            name             = award_name,
            organization     = organization,             
            amount           = amount,
            recognition_type = recognition_type,              
            start_date       = start_date,
            end_date         = end_date,                    
            description      = extract_field_value(recog_section, 'Description') or '',
        )

        existing = Recognition.objects.filter(
            researcher=researcher,
            external_id=ext_id
        ).first()

        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
            existing.save()
            count_updated += 1
        else:
            Recognition.objects.create(
                researcher=researcher, external_id=ext_id, **fields
            )
            count_created += 1

    Recognition.objects.filter(
        researcher=researcher
    ).exclude(
        external_id__in=seen_ids
    ).delete()

    return count_created + count_updated


# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_publications  — switches from generated hash to recordId
# ───────────────────────────────────────────────────────────────────────────────

def extract_description(section):
    """Extract description from CCV bilingual fields."""
    for label in ['Description/Contribution Value/Impact', 'Contribution Value']:
        field = section.find(f'field[@label="{label}"]')
        if field is not None:
            bilingual = field.find('bilingual/english')
            if bilingual is not None and bilingual.text:
                return bilingual.text.strip()
            val = field.find('value')
            if val is not None:
                text = val.get('text')
                if text:
                    return text.strip()
            text = extract_field_value(section, label)
            if text:
                return text.strip()
    return None

def parse_xml_publications(researcher, xml_root):
    seen_external_ids = set()
    count_created = count_updated = 0

    publication_labels = [
        ("Journal Articles",        "Article Title",     "journal"),
        ("Conference Publications",  "Publication Title", "conference"),
        ("Book Chapters",            "Chapter Title",     "chapter"),
        ("Patents",                  "Patent Title",      "patent"),
    ]

    description = None

    for label, title_field, pub_type in publication_labels:
        for pub_section in xml_root.findall(f'.//section[@label="{label}"]'):
            title = extract_field_value(pub_section, title_field)
            if not title:
                continue

            ext_id = get_record_id(pub_section, [title, pub_type])
            seen_external_ids.add(ext_id)

            authors = extract_field_value(pub_section, 'Authors')
            doi = status = None
            status = 'published'
            description = None
            (journal, year_str, volume, issue, pages, publisher,
             open_access, refereed, invited, conference_location,
             book_title, editors, publication_location, contribution_role,
             patent_number, country, filing_date) = [None] * 17

            if label == "Journal Articles":
                journal           = extract_field_value(pub_section, 'Journal')
                year_str          = extract_field_value(pub_section, 'Year')
                volume            = extract_field_value(pub_section, 'Volume')
                issue             = extract_field_value(pub_section, 'Issue')
                pages             = extract_field_value(pub_section, 'Page Range')
                publisher         = extract_field_value(pub_section, 'Publisher')
                description       = (extract_field_value(pub_section, 'Abstract') or
                                     extract_field_value(pub_section, 'Description/Contribution Value/Impact'))
                doi               = extract_field_value(pub_section, 'DOI')

                if isinstance(doi, dict):
                    doi = doi.get('value', {}).get('text')

                if not doi:
                    url_val = extract_field_value(pub_section, 'URL')
                    if isinstance(url_val, list):
                        url_val = url_val[0] if url_val else None
                    if url_val:
                        url_val = str(url_val).strip()
                        if url_val.upper().startswith('DOI:'):
                            doi = url_val[4:].strip()
                        elif 'doi.org/' in url_val.lower():
                            parts = url_val.split('doi.org/')
                            if len(parts) > 1:
                                doi = parts[1].strip()

                if doi:
                    doi = str(doi).replace('https://', '').replace('http://', '').strip()

                contribution_role = extract_field_value(pub_section, 'Contribution Role')
                editors           = extract_field_value(pub_section, 'Editors')
                oa_str            = extract_field_value(pub_section, 'Open Access?')
                open_access       = True if oa_str and oa_str.lower() == 'yes' else (False if oa_str else None)
                ref_str           = extract_field_value(pub_section, 'Refereed?')
                refereed          = True if ref_str and ref_str.lower() == 'yes' else (False if ref_str else None)
                s = (extract_field_value(pub_section, 'Publishing Status') or '').lower()
                if 'revision' in s:                   status = 'revision_requested'
                elif 'rejected' in s:                 status = 'rejected'
                elif 'accepted' in s or 'press' in s: status = 'accepted'
                elif 'under review' in s:             status = 'under_review'

            elif label == "Conference Publications":
                journal             = extract_field_value(pub_section, 'Conference Name')
                year_str            = extract_field_value(pub_section, 'Conference Date')
                pages               = extract_field_value(pub_section, 'Page Range')
                conference_location = extract_field_value(pub_section, 'Conference Location')
                city                = extract_field_value(pub_section, 'City')
                if conference_location and city:
                    conference_location = f"{city}, {conference_location}"
                elif city:
                    conference_location = city
                doi         = extract_field_value(pub_section, 'DOI')
                description = (extract_field_value(pub_section, 'Abstract') or
                               extract_field_value(pub_section, 'Description/Contribution Value/Impact'))

                if isinstance(doi, dict):
                    doi = doi.get('value', {}).get('text')

                if not doi:
                    url_val = extract_field_value(pub_section, 'URL')
                    if isinstance(url_val, list):
                        url_val = url_val[0] if url_val else None
                    if url_val:
                        url_val = str(url_val).strip()
                        if url_val.upper().startswith('DOI:'):
                            doi = url_val[4:].strip()
                        elif 'doi.org/' in url_val.lower():
                            parts = url_val.split('doi.org/')
                            if len(parts) > 1:
                                doi = parts[1].strip()

                if doi:
                    doi = str(doi).replace('https://', '').replace('http://', '').strip()

                contribution_role    = extract_field_value(pub_section, 'Contribution Role')
                editors              = extract_field_value(pub_section, 'Editors')
                publisher            = extract_field_value(pub_section, 'Publisher')
                publication_location = extract_field_value(pub_section, 'Publication Location')
                inv_str              = extract_field_value(pub_section, 'Invited?')
                invited              = True if inv_str and inv_str.lower() == 'yes' else (False if inv_str else None)
                ref_str              = extract_field_value(pub_section, 'Refereed?')
                refereed             = True if ref_str and ref_str.lower() == 'yes' else (False if ref_str else None)

            elif label == "Book Chapters":
                book_title        = extract_field_value(pub_section, 'Book Title')
                journal           = book_title
                year_str          = extract_field_value(pub_section, 'Year')
                pages             = extract_field_value(pub_section, 'Page Range')
                publisher         = extract_field_value(pub_section, 'Publisher')
                editors           = extract_field_value(pub_section, 'Editors')
                contribution_role = extract_field_value(pub_section, 'Contribution Role')
                doi               = extract_field_value(pub_section, 'DOI')
                description       = (extract_field_value(pub_section, 'Abstract') or
                                     extract_field_value(pub_section, 'Description/Contribution Value/Impact'))

                if isinstance(doi, dict):
                    doi = doi.get('value', {}).get('text')

                if not doi:
                    url_val = extract_field_value(pub_section, 'URL')
                    if isinstance(url_val, list):
                        url_val = url_val[0] if url_val else None
                    if url_val:
                        url_val = str(url_val).strip()
                        if url_val.upper().startswith('DOI:'):
                            doi = url_val[4:].strip()
                        elif 'doi.org/' in url_val.lower():
                            parts = url_val.split('doi.org/')
                            if len(parts) > 1:
                                doi = parts[1].strip()

                if doi:
                    doi = str(doi).replace('https://', '').replace('http://', '').strip()

                publication_location = extract_field_value(pub_section, 'Publication Location')
                volume               = extract_field_value(pub_section, 'Volume')
                ref_str              = extract_field_value(pub_section, 'Refereed?')
                refereed             = True if ref_str and ref_str.lower() == 'yes' else (False if ref_str else None)

            elif label == "Patents":
                patent_number   = extract_field_value(pub_section, 'Patent Number')
                journal         = patent_number
                country         = extract_field_value(pub_section, 'Patent Location')
                authors         = (extract_field_value(pub_section, 'Inventors') or
                                   extract_field_value(pub_section, 'Authors') or
                                   researcher.user.get_full_name())
                filing_date_str = extract_field_value(pub_section, 'Filing Date')
                year_str        = extract_field_value(pub_section, 'Year Issued') or filing_date_str
                if filing_date_str and len(filing_date_str.strip()) == 10:
                    try:
                        filing_date = datetime.strptime(filing_date_str.strip(), '%Y-%m-%d').date()
                    except Exception:
                        filing_date = None
                patent_status = extract_field_value(pub_section, 'Patent Status') or ''
                status = ('granted'
                          if 'grant' in patent_status.lower() or 'issued' in patent_status.lower()
                          else 'pending')
                description = extract_description(pub_section)

            # ── Date parsing ──────────────────────────────────────────
            pub_date = None
            if year_str:
                try:
                    if '/' in year_str and len(year_str) in (6, 7):
                        pub_date = parse_date_from_yearmonth(year_str)
                    elif len(year_str.strip()) == 10 and '-' in year_str:
                        pub_date = datetime.strptime(year_str.strip(), "%Y-%m-%d").date()
                    elif len(year_str.strip()) == 4:
                        pub_date = date(int(year_str.strip()), 1, 1)
                except Exception:
                    pass

            if not pub_date:
                fallback = (extract_field_value(pub_section, 'Year') or
                            extract_field_value(pub_section, 'Year Issued'))
                if fallback and len(fallback.strip()) == 4:
                    try:
                        pub_date = date(int(fallback.strip()), 1, 1)
                    except Exception:
                        pass

            if not pub_date:
                continue

            language   = extract_field_value(pub_section, 'Language')
            pub_fields = dict(
                title=title[:500], authors=authors or 'Unknown', journal=journal or '',
                publication_date=pub_date, doi=doi, publication_type=pub_type, status=status,
                language=language, pages=pages, refereed=refereed,
                contribution_role=contribution_role,
                publication_location=publication_location, volume=volume, issue=issue,
                publisher=publisher, open_access=open_access,
                conference_location=conference_location, invited=invited,
                book_title=book_title, editors=editors, patent_number=patent_number,
                country=country, filing_date=filing_date, is_active=True,
                abstract=description,
            )

            # ── Step 1: Exact match by external_id ────────────────────
            existing = Publication.all_objects.filter(
                researcher=researcher, external_id=ext_id
            ).first()

            # ── Step 2: Fallback → match manual publication by title + date
            if not existing:
                existing = Publication.all_objects.filter(
                    researcher=researcher,
                    title__iexact=title,
                    publication_date=pub_date,
                    source='manual',
                ).first()

                if existing:
                    existing.source      = 'ccv'
                    existing.external_id = ext_id
                    logger.info(
                        f"[CCV] Manual publication '{title}' claimed by CCV import "
                        f"(researcher={researcher.user.email})"
                    )

            if existing:
                if existing.is_deleted:
                    seen_external_ids.add(ext_id)
                    continue

                # ── Protect manually set statuses ─────────────────────
                MANUAL_STATUSES = ['accepted', 'under_review', 'revision_requested', 'draft']
                if existing.status in MANUAL_STATUSES:
                    pub_fields['status'] = existing.status

                for field, value in pub_fields.items():
                    setattr(existing, field, value)

                existing.is_active = True
                existing.save()
                auto_link_publication_authors(existing)
                count_updated += 1

            else:
                new_pub = Publication.objects.create(
                    researcher=researcher,
                    external_id=ext_id,
                    **pub_fields,
                )
                auto_link_publication_authors(new_pub)
                count_created += 1

    Publication.objects.filter(
        researcher=researcher, is_active=True
    ).exclude(external_id__in=seen_external_ids).update(is_active=False)

    return count_created + count_updated

# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_activities  — switches from generated hash to recordId
# ───────────────────────────────────────────────────────────────────────────────

def parse_xml_activities(researcher, xml_root):
    seen_external_ids = set()
    count_created = count_updated = 0

    for label in ["Presentations", "Broadcast Interviews", "Text Interviews"]:
        for activity_section in xml_root.findall(f'.//section[@label="{label}"]'):
            title = event = date_str = topic = co_presenters = audience = None
            invited = keynote = network = forum = city = country = None
            description = None

            if label == "Presentations":
                title         = extract_field_value(activity_section, 'Presentation Title')
                event         = extract_field_value(activity_section, 'Conference / Event Name')
                country       = extract_field_value(activity_section, 'Location')
                city          = extract_field_value(activity_section, 'City')
                date_str      = extract_field_value(activity_section, 'Presentation Year')
                activity_type = 'presentation'
                co_presenters = extract_field_value(activity_section, 'Co-Presenters')
                audience      = extract_field_value(activity_section, 'Main Audience')
                invited       = extract_field_value(activity_section, 'Invited?')
                keynote       = extract_field_value(activity_section, 'Keynote?')
                description   = extract_field_value(activity_section, 'Description / Contribution Value')

            elif label == "Broadcast Interviews":
                topic         = extract_field_value(activity_section, 'Topic')
                title         = topic or extract_field_value(activity_section, 'Program')
                network       = extract_field_value(activity_section, 'Network')
                event         = network
                date_str      = extract_field_value(activity_section, 'First Broadcast Date')
                activity_type = 'broadcast'
                description   = extract_field_value(activity_section, 'Description / Contribution Value')

            elif label == "Text Interviews":
                topic         = extract_field_value(activity_section, 'Topic')
                title         = topic
                forum         = extract_field_value(activity_section, 'Forum')
                event         = forum
                date_str      = extract_field_value(activity_section, 'Publication Date')
                activity_type = 'text_interview'
                description   = extract_field_value(activity_section, 'Description / Contribution Value')

            if not title:
                continue

            activity_date = None
            if date_str:
                try:
                    ds = date_str.strip()
                    if len(ds) == 10 and ds.count('-') == 2:
                        activity_date = datetime.strptime(ds, "%Y-%m-%d").date()
                    elif len(ds) == 4:
                        activity_date = date(int(ds), 1, 1)
                    else:
                        activity_date = parse_date_from_yearmonth(date_str)
                except Exception:
                    pass

            if not activity_date:
                continue

            if city and country:
                location = f"{city}, {country}"
            elif city:
                location = city
            elif country:
                location = country
            else:
                location = None

            def to_bool(val):
                if val is None:
                    return None
                return val.strip().lower() in ('yes', 'true', '1')

            category = determine_activity_category(title, topic, audience, event)

            ext_id = get_record_id(activity_section, [title, str(activity_date), activity_type])
            seen_external_ids.add(ext_id)

            final_description = description or event or ''

            act_fields = dict(
                activity_type = activity_type,
                title         = title[:255],
                description   = final_description,
                date          = activity_date,
                category      = category,
                is_active     = True,
                source        = 'ccv',
                location      = location,
                invited       = to_bool(invited),
                keynote       = to_bool(keynote),
                co_presenters = co_presenters or '',
                audience      = audience or '',
            )

            # ── Step 1: Exact CCV match ───────────────────────────────
            existing = Activity.all_objects.filter(
                researcher  = researcher,
                external_id = ext_id,
                source      = 'ccv',
            ).first()
            
            # ── Step 2: Fallback → match manual activity by title + date
            if not existing:
                existing = Activity.all_objects.filter(
                    researcher    = researcher,
                    title__iexact = title,
                    date          = activity_date,
                    source        = 'manual',
                ).first()

                if existing:
                    existing.source      = 'ccv'
                    existing.external_id = ext_id
                    logger.info(
                        f"[CCV] Manual activity '{title}' on {activity_date} claimed by CCV import "
                        f"(researcher={researcher.user.email})"
                    )

            if existing:
                if existing.is_deleted:
                    seen_external_ids.add(ext_id)
                    continue

                # ── Preserve manually set category on claimed activities ──
                manual_category = existing.category

                for k, v in act_fields.items():
                    setattr(existing, k, v)

                # If this was a manual activity, keep the user's category
                if existing.source == 'ccv' and manual_category:
                    existing.category = manual_category

                existing.is_active = True
                existing.save()
                count_updated += 1

            else:
                Activity.objects.create(
                    researcher  = researcher,
                    external_id = ext_id,
                    **act_fields,
                )
                count_created += 1

    Activity.objects.filter(
        researcher=researcher, source='ccv', is_active=True
    ).exclude(external_id__in=seen_external_ids).update(is_active=False)

    return count_created + count_updated


# ───────────────────────────────────────────────────────────────────────────────
# parse_xml_projects
# ───────────────────────────────────────────────────────────────────────────────


import logging
logger = logging.getLogger(__name__)

def parse_xml_projects(researcher, xml_root):
    seen_external_ids = set()
    count_created = count_updated = 0

    for proj_section in xml_root.findall('.//section[@label="Research Funding History"]'):
        title = extract_field_value(proj_section, 'Funding Title')
        if not title:
            continue

        ext_id = get_record_id(proj_section, [title])
        seen_external_ids.add(ext_id)

        start_date = parse_funding_date(extract_field_value(proj_section, 'Funding Start Date'))
        end_date   = parse_funding_date(extract_field_value(proj_section, 'Funding End Date'))

        status_str = extract_field_value(proj_section, 'Funding Status') or ''
        s = status_str.lower()
        if 'award' in s or 'funded' in s:
            status = 'awarded'
        elif 'completed' in s or 'closed' in s:
            status = 'completed'
        elif 'rejected' in s or 'unsuccessful' in s or 'declined' in s:
            status = 'rejected'
        elif 'review' in s or 'submitted' in s or 'pending' in s or 'under' in s:
            status = 'submitted'
        else:
            status = 'pending'

        role         = map_funding_role_to_choice(extract_field_value(proj_section, 'Funding Role'))
        funding_type = extract_field_value(proj_section, 'Funding Type')

        description = None
        desc_field  = proj_section.find('field[@label="Project Description"]')
        if desc_field is not None:
            description = extract_bilingual_value(desc_field)

        # ── Funding aggregation ───────────────────────────────────────
        funding_org          = None
        program_name         = None
        currency             = 'CAD'
        total_funding_sum    = 0.0
        funding_received_sum = 0.0

        for funding_src in proj_section.findall('section[@label="Funding Sources"]'):
            org_main  = extract_field_value(funding_src, 'Funding Organization')
            org_other = extract_field_value(funding_src, 'Other Funding Organization')
            org = (
                f"{org_main} / {org_other}" if (org_main and org_other)
                else (org_main or org_other)
            )

            if not funding_org and org:
                funding_org  = org
                program_name = extract_field_value(funding_src, 'Program Name')
                currency_str = (extract_field_value(funding_src, 'Currency of Total Funding') or '').lower()
                currency     = 'USD' if 'united states' in currency_str else 'CAD'

            total_str    = extract_field_value(funding_src, 'Total Funding')
            received_str = extract_field_value(funding_src, 'Portion of Funding Received')

            try:
                total_funding_sum += float(total_str) if total_str else 0
            except Exception as e:
                logger.warning(
                    f"[CCV] Failed to parse total funding for '{title}' "
                    f"(researcher={researcher.user.email}): {total_str!r} | {e}"
                )

            try:
                funding_received_sum += float(received_str) if received_str else 0
            except Exception as e:
                logger.warning(
                    f"[CCV] Failed to parse funding received for '{title}' "
                    f"(researcher={researcher.user.email}): {received_str!r} | {e}"
                )

        total_funding    = total_funding_sum    or None
        funding_received = funding_received_sum or None

        # ── CCV-controlled fields (always overwrite) ──────────────────
        ccv_fields = dict(
            description = description,
            ccv_active  = True,
            source      = 'ccv',
            currency    = currency,
        )

        # ── Protected fields (respect manual edits) ───────────────────
        first_create_only = dict(
            title                = title[:500],
            status               = status,
            role                 = role,
            funding_type         = funding_type,
            funding_organization = funding_org or '',
            program_name         = program_name or '',
            total_funding        = total_funding,
            funding_received     = funding_received,
            start_date           = start_date,
            end_date             = end_date,
        )

        # ── Step 1: Try exact CCV match ───────────────────────────────
        existing = Project.all_objects.filter(
            researcher  = researcher,
            external_id = ext_id,
            source      = 'ccv',
        ).first()

        # ── Step 2: Fallback → match manual project by title ──────────
        if not existing:
            existing = Project.all_objects.filter(
                researcher   = researcher,
                title__iexact = title,
                source       = 'manual',
            ).first()

            if existing:
                # Claim this manual project as CCV-controlled
                existing.source      = 'ccv'
                existing.external_id = ext_id
                logger.info(
                    f"[CCV] Manual project '{title}' claimed by CCV import "
                    f"(researcher={researcher.user.email})"
                )

        # ── UPDATE ────────────────────────────────────────────────────
        if existing:
            if existing.is_deleted:
                continue

            for k, v in ccv_fields.items():
                setattr(existing, k, v)

            if not existing.manually_overridden:
                for k, v in first_create_only.items():
                    setattr(existing, k, v)

            existing.ccv_active = True
            existing.save()
            project = existing
            count_updated += 1

        # ── CREATE ────────────────────────────────────────────────────
        else:
            project = Project.objects.create(
                researcher  = researcher,
                external_id = ext_id,
                **ccv_fields,
                **first_create_only,
            )
            count_created += 1

        # ── Refresh CCV investigators only ────────────────────────────
        # is_academic_collaborator=True  → came from CCV → safe to refresh
        # is_academic_collaborator=False → manually added partner → preserve
        project.team_members.filter(is_academic_collaborator=True).delete()

        for inv_section in proj_section.findall('section[@label="Other Investigators"]'):
            inv_name = extract_field_value(inv_section, 'Investigator Name')
            if inv_name:
                ProjectMember.objects.create(
                    project                  = project,
                    name                     = inv_name,
                    role                     = map_funding_role_to_choice(
                        extract_field_value(inv_section, 'Role')
                    ),
                    is_academic_collaborator = True,
                    partner_type             = 'academic',
                )

    # ── Deactivate CCV projects no longer in XML ──────────────────────
    Project.objects.filter(
        researcher = researcher,
        source     = 'ccv',
        ccv_active = True,
    ).exclude(external_id__in=seen_external_ids).update(ccv_active=False)

    return count_created + count_updated

@login_required
@researcher_required
@require_http_methods(["POST"])
def api_update_project_status(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project = Project.objects.get(id=project_id, researcher=researcher)
        data = json.loads(request.body)
        new_status = data.get('status', '')
        valid_statuses = ['awarded', 'completed', 'pending', 'submitted', 'rejected']
        if new_status in valid_statuses:
            project.status = new_status
            project.manually_overridden = True
            project.save()
            log_action(request, 'project_updated', target=project,
                       summary=f'{request.user.get_full_name()} updated status of "{project.title}" to {new_status}')
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_http_methods(["POST"])
def api_update_publication_status(request, publication_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        pub = Publication.objects.get(id=publication_id, researcher=researcher)
        data = json.loads(request.body)
        new_status = data.get('status', '')
        valid = ['published', 'accepted', 'under_review', 'revision_requested', 'rejected', 'pending', 'draft', 'granted']
        if new_status in valid:
            pub.status = new_status
            pub.save()
            log_action(request, 'publication_updated', target=pub,
                       summary=f'{request.user.get_full_name()} updated status of "{pub.title}" to {new_status}')
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Invalid status'}, status=400)
    except Publication.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


def get_student_researcher_proxy(user):
    """
    Returns the ResearcherProfile used to store a student's
    publications and activities. Creates one if absent.

    NOTE:
    - This is NOT a real researcher profile.
    - The associated user has user_type='student'.
    - Do NOT use this for researcher-only queries.

    Raises:
        ValueError: if called with a non-student user.
    """

    # Hard safety check (better than assert in production)
    if user.user_type != 'student':
        raise ValueError(
            f"get_student_researcher_proxy() called with non-student user: "
            f"{user.email} (type={user.user_type})"
        )

    profile, created = ResearcherProfile.objects.get_or_create(
        user=user,
        defaults={'research_interests': ''}
    )

    if created:
        try:
            log_action(
                None,
                'other',
                target=profile,
                summary=f'ResearcherProfile proxy created for student {user.email}'
            )
        except Exception:
            # Logging should never break the flow
            pass

    return profile

import re
from datetime import date
def parse_xml_student_profile(user, xml_root):

    profile, _ = StudentProfile.objects.get_or_create(user=user, defaults={})

    if not profile.manually_overridden:

        for edu_section in xml_root.findall('.//section[@label="Education"]'):

            latest_degree_section = None
            latest_date = None

            for degree_section in edu_section.findall('.//section[@label="Degrees"]'):

                received = degree_section.find('field[@label="Degree Received Date"]')
                expected = degree_section.find('field[@label="Degree Expected Date"]')
                start    = degree_section.find('field[@label="Degree Start Date"]')

                # explicit priority: completed > ongoing > started
                if received is not None:
                    date_field = received
                elif expected is not None:
                    date_field = expected
                else:
                    date_field = start

                date_val = ''
                if date_field is not None:
                    date_val = (date_field.findtext('value') or '').strip()

                try:
                    parts = date_val.replace('/', '-').split('-')
                    year  = int(parts[0])
                    month = int(parts[1]) if len(parts) > 1 else 1
                    d     = date(year, month, 1)
                except (ValueError, IndexError):
                    d = None

                if d is not None:
                    if latest_date is None or d > latest_date:
                        latest_date           = d
                        latest_degree_section = degree_section
                else:
                    latest_degree_section = degree_section  # fallback — last wins

            # ── process only the latest degree ───────────────────────
            if latest_degree_section is None:
                continue

            degree_type_field = latest_degree_section.find('field[@label="Degree Type"]')
            if degree_type_field is not None:
                val = (
                    degree_type_field.findtext('lov') or
                    degree_type_field.findtext('value') or
                    degree_type_field.findtext('lov/text') or
                    ''
                ).strip().lower()

                if 'bachelor' in val or 'bsc' in val:
                    profile.degree_level = 'undergrad'
                elif 'master' in val or 'msc' in val:
                    profile.degree_level = 'msc'
                elif 'phd' in val or 'doctor' in val:
                    profile.degree_level = 'phd'
                elif 'postdoc' in val:
                    profile.degree_level = 'pdf'

            spec_field = latest_degree_section.find('field[@label="Specialization"]')
            dept_field = latest_degree_section.find('field[@label="Department"]')

            # Priority 1: Specialization field
            if spec_field is not None:
                val = spec_field.findtext('value', '').strip()
                if val:
                    profile.department = val

            # Priority 2: Department field
            if not profile.department and dept_field is not None:
                val = dept_field.findtext('value', '').strip()
                if val:
                    profile.department = val

            # Priority 3: Any department field anywhere in the XML
            if not profile.department:
                for field in xml_root.findall('.//field[@label="Department"]'):
                    val = field.findtext('value')
                    if val and val.strip():
                        profile.department = val.strip()
                        break

            # Priority 4: Infer from Degree Name
            if not profile.department:
                degree_name_field = latest_degree_section.find('field[@label="Degree Name"]')
                if degree_name_field is not None:
                    val = degree_name_field.findtext('value', '').strip()
                    if val and ' of ' in val.lower():
                        profile.department = val.split(' of ', 1)[-1].strip()

            # ── thesis title ──────────────────────────────────────────
            thesis_field = latest_degree_section.find('field[@label="Thesis Title"]')
            if thesis_field is not None:
                val = thesis_field.findtext('value', '').strip()
                if val:
                    profile.thesis_title = val

            # ── dates ─────────────────────────────────────────────────
            for field_label, attr in [
                ('Degree Start Date',    'start_date'),
                ('Degree Expected Date', 'expected_end_date'),
                ('Degree Received Date', 'graduation_date'),
            ]:
                f = latest_degree_section.find(f'field[@label="{field_label}"]')
                if f is not None:
                    val = (f.findtext('value') or '').strip()
                    if val:
                        try:
                            parts = val.replace('/', '-').split('-')
                            year  = int(parts[0])
                            month = int(parts[1]) if len(parts) > 1 else 1
                            setattr(profile, attr, date(year, month, 1))
                        except (ValueError, IndexError):
                            pass

            break  # one edu section is enough

    # ── supervisor parsing — runs regardless of manually_overridden ───
    if not profile.supervisor:
        for edu_section in xml_root.findall('.//section[@label="Education"]'):
            for degree_section in edu_section.findall('.//section[@label="Degrees"]'):
                for sup_section in degree_section.findall('.//section[@label="Supervisors"]'):

                    sup_name = (
                        extract_field_value(sup_section, 'Supervisor Name') or ''
                    ).strip()

                    if not sup_name:
                        continue

                    clean = re.sub(r'[^a-zA-Z\s]', '', sup_name.lower()).strip()
                    parts = clean.split()
                    if len(parts) < 2:
                        continue

                    first = parts[0]
                    last  = parts[-1]

                    # try normal order
                    matches = list(
                        ResearcherProfile.objects.filter(
                            user__first_name__iexact=first,
                            user__last_name__iexact=last,
                            user__user_type='researcher',
                        )[:2]
                    )

                    if len(matches) == 1:
                        profile.supervisor = matches[0]
                        break

                    # try reversed
                    reversed_matches = list(
                        ResearcherProfile.objects.filter(
                            user__first_name__iexact=last,
                            user__last_name__iexact=first,
                            user__user_type='researcher',
                        )[:2]
                    )

                    if len(reversed_matches) == 1:
                        profile.supervisor = reversed_matches[0]
                        break

                if profile.supervisor:
                    break
            if profile.supervisor:
                break

    # ── student researcher proxy for pubs/activities ──────────────────
    researcher = get_student_researcher_proxy(user)

    try:
        with transaction.atomic():
            profile.save() 
            pub_count   = parse_xml_publications(researcher, xml_root)
            act_count   = parse_xml_activities(researcher, xml_root)
            recog_count = parse_xml_recognitions(researcher, xml_root)
            edu_count   = parse_xml_education(researcher, xml_root)
    except Exception as e:
        raise

    return {
        'degree_level': profile.degree_level,
        'department':   profile.department,
        'publications': pub_count,
        'activities':   act_count,
        'recognitions':   recog_count,
        'education':     edu_count,
    }

def parse_xml_supervision(researcher, xml_root):
    from .models import SupervisionRecord

    seen_ids = set()
    count_created = count_updated = 0

    # ── preload existing records — one DB hit instead of N ──
    existing_records = {
        r.external_id: r
        for r in SupervisionRecord.objects.filter(researcher=researcher)
    }

    for sup_section in xml_root.findall('.//section[@label="Student/Postdoctoral Supervision"]'):

        # ── strip all extracted values ──────────────────
        student_name = (extract_field_value(sup_section, 'Student Name') or '').strip()
        if not student_name:
            continue

        ext_id = get_record_id(sup_section, [student_name, str(researcher.id)])
        seen_ids.add(ext_id)

        # ── fuzzy degree matching ────────────────────────
        degree_raw   = extract_field_value(sup_section, 'Degree Type or Postdoctoral Status') or ''
        degree_clean = degree_raw.lower().replace("\u2019", "'").replace("'", "'").strip()

        if 'bachelor' in degree_clean:
            degree_type = 'bachelors'
        elif 'non-thesis' in degree_clean or 'non thesis' in degree_clean:
            degree_type = 'masters_non_thesis'
        elif 'master' in degree_clean:
            degree_type = 'masters_thesis'
        elif 'doctor' in degree_clean or 'phd' in degree_clean:
            degree_type = 'doctorate'
        elif 'postdoc' in degree_clean or 'post-doc' in degree_clean:
            degree_type = 'postdoc'
        elif 'research associate' in degree_clean:
            degree_type = 'research_associate'
        else:
            degree_type = None

        # ── improved status detection ───────────────────
        status_raw   = extract_field_value(sup_section, 'Student Degree Status') or ''
        status_clean = status_raw.lower()

        if 'completed' in status_clean:
            status = 'completed'
        elif any(x in status_clean for x in ['progress', 'ongoing', 'current', 'enrolled']):
            status = 'in_progress'
        else:
            status = None

        role        = (extract_field_value(sup_section, 'Supervision Role') or '').strip()
        institution = (extract_field_value(sup_section, 'Student Institution') or '').strip()
        thesis      = (extract_field_value(sup_section, 'Thesis/Project Title') or '').strip()
        position    = (extract_field_value(sup_section, 'Present Position') or '').strip()
        org         = (extract_field_value(sup_section, 'Present Organization') or '').strip()
        residency   = (extract_field_value(sup_section, 'Student Canadian Residency Status') or '').strip()

        start_date   = parse_funding_date(extract_field_value(sup_section, 'Supervision Start Date'))
        end_date     = parse_funding_date(extract_field_value(sup_section, 'Supervision End Date'))
        degree_start = parse_funding_date(extract_field_value(sup_section, 'Student Degree Start Date'))
        degree_end   = parse_funding_date(extract_field_value(sup_section, 'Student Degree Received Date'))
        expected     = parse_funding_date(extract_field_value(sup_section, 'Student Degree Expected Date'))

        # CCV-owned fields — always safe to overwrite
        ccv_fields = dict(
            student_name      = student_name,
            institution       = institution,
            degree_type       = degree_type,
            supervision_role  = role,
            status            = status,
            start_date        = start_date,
            end_date          = end_date,
            degree_start_date = degree_start,
            degree_end_date   = degree_end,
            thesis_title      = thesis,
            present_position  = position,
            present_org       = org,
            residency_status  = residency,
        )

        # Manually editable fields — protected if overridden
        manual_fields = dict(
            expected_date = expected,
        )

        existing = existing_records.get(ext_id)

        if existing:
            for k, v in ccv_fields.items():
                setattr(existing, k, v)
            if not existing.manually_overridden:
                existing.expected_date = manual_fields['expected_date']
            existing.save()
            try_link_student(existing)  # ← inside loop, not after
            count_updated += 1
        else:
            new_record = SupervisionRecord.objects.create(
                researcher  = researcher,
                external_id = ext_id,
                **ccv_fields,
                expected_date = expected,
            )
            try_link_student(new_record)  # ← inside loop
            count_created += 1

    # ── soft delete stale records ───────────────────
    # SupervisionRecord has no is_active field so we hard delete
    # only records that were NOT manually overridden
    SupervisionRecord.objects.filter(
        researcher=researcher,
    ).exclude(
        external_id__in=seen_ids
    ).filter(
        manually_overridden=False  # protect manually edited records
    ).delete()

    return count_created + count_updated


@login_required
@researcher_required
@require_http_methods(["POST"])
def api_update_supervision_record(request, record_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        record     = SupervisionRecord.objects.get(id=record_id, researcher=researcher)
        data       = json.loads(request.body)

        # ── reject empty payload ─────────────────────────
        if not data:
            return JsonResponse({'success': False, 'error': 'No data provided'}, status=400)

        updated = False

        # ── linked_student (manual override) ─────────────────────
        if 'linked_student_id' in data:
            sid = data.get('linked_student_id')
            if sid:
                # Only allow linking actual students, not researchers
                student = StudentProfile.objects.filter(
                    id=sid,
                    user__user_type='student'
                ).first()

                if not student:
                    return JsonResponse({
                        'success': False,
                        'error': 'Student not found or invalid account type.'
                    }, status=404)

                # Name similarity check
                record_name   = re.sub(r'[^a-zA-Z\s]', '', record.student_name.lower()).strip()
                student_name  = re.sub(r'[^a-zA-Z\s]', '', student.user.get_full_name().lower()).strip()
                record_parts  = set(record_name.split())
                student_parts = set(student_name.split())

                if not record_parts & student_parts:
                    return JsonResponse({
                        'success': False,
                        'error': f'Name mismatch — this record is for "{record.student_name}" '
                                f'but selected account belongs to "{student.user.get_full_name()}".'
                    }, status=400)

                if record.linked_student != student:
                    record.linked_student = student
                    updated = True
            else:
                if record.linked_student is not None:
                    record.linked_student = None
                    updated = True

        # ── department ────────────────────────────────────────
        if 'department' in data:
            new_val = (data.get('department') or '').strip() or None
            if record.department != new_val:
                record.department = new_val
                updated = True

        # ── expected_date ─────────────────────────────────────
        if 'expected_date' in data:
            from django.utils.dateparse import parse_date
            raw = data.get('expected_date') or ''
            if raw:
                parsed = parse_date(raw)
                if not parsed:
                    return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)
            else:
                parsed = None
            if record.expected_date != parsed:
                record.expected_date = parsed
                updated = True

        # ── status ────────────────────────────────────────────
        if 'status' in data:
            new_val = data.get('status')
            if new_val in ('in_progress', 'completed') and record.status != new_val:
                record.status = new_val
                updated = True

        # ── degree_type ───────────────────────────────────────
        if 'degree_type' in data:
            valid   = {c[0] for c in SupervisionRecord.DEGREE_CHOICES}
            new_val = data.get('degree_type')
            if new_val in valid and record.degree_type != new_val:
                record.degree_type = new_val
                updated = True

        # ── only save and flag if something actually changed ──
        if updated:
            record.manually_overridden = True
            record.save()
            log_action(request, 'profile_updated', target=record,
                       summary=f'{request.user.get_full_name()} updated supervision record for "{record.student_name}"')

        return JsonResponse({
            'success': True,
            'updated': updated,
            'record': {
                'department':    record.department,
                'expected_date': str(record.expected_date) if record.expected_date else None,
                'status':        record.status,
                'degree_type':   record.degree_type,
            }
        })

    except SupervisionRecord.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Record not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': 'Internal server error'}, status=500)


import re

def try_link_student(record):
    if record.linked_student:
        return

    clean = re.sub(r'[^a-zA-Z\s]', '', record.student_name.lower()).strip()
    name_parts = clean.split()
    if len(name_parts) < 2:
        return

    first = name_parts[0]
    last  = name_parts[-1]

    matches = StudentProfile.objects.filter(
        user__first_name__iexact=first,
        user__last_name__iexact=last,
    )

    if matches.count() == 1:
        student = matches.first()
        record.linked_student = student
        record.save(update_fields=['linked_student'])

        # ── Notify student only on NEW link, no duplicates ────
        already_notified = StudentNotification.objects.filter(
            user=student.user,
            message__icontains="supervision record for you"
        ).exists()

        if not already_notified:
            StudentNotification.objects.create(
                user=student.user,
                message=(
                    f'{record.researcher.user.get_full_name()} appears as your supervisor in a record. '
                    f'If they are your current supervisor, you can send them a request from your profile.'
                )
            )
        return

    if matches.count() == 0:
        reversed_matches = StudentProfile.objects.filter(
            user__first_name__iexact=last,
            user__last_name__iexact=first,
        )
        if reversed_matches.count() == 1:
            student = reversed_matches.first()
            record.linked_student = student
            record.save(update_fields=['linked_student'])

            already_notified = StudentNotification.objects.filter(
                user=student.user,
                message__icontains="supervision record for you"
            ).exists()

            if not already_notified:
                StudentNotification.objects.create(
                    user=student.user,
                    message=(
                        f'{record.researcher.user.get_full_name()} appears as your supervisor in a record. '
                        f'If they are your current supervisor, you can send them a request from your profile.'
                    )
                )


# at the end of approve_and_create_profiles in admin.py
# and at the end of api_create_user_with_temp_password in views.py

def reverse_link_supervision(user):
    """When a student registers, check if any supervision records match them."""
    try:
        profile = user.student_profile
    except Exception:
        return

    clean = re.sub(r'[^a-zA-Z\s]', '', user.get_full_name().lower()).strip()
    parts = clean.split()
    if len(parts) < 2:
        return

    first, last = parts[0], parts[-1]

    unlinked = SupervisionRecord.objects.filter(
        linked_student__isnull=True,
    ).filter(
        Q(student_name__iexact=f"{first} {last}") |
        Q(student_name__iexact=f"{last} {first}")
    )

    for record in unlinked:
        try_link_student(record)


# ─────────────────────────────────────────────
# Bulk Upload
# ─────────────────────────────────────────────

def validate_ccv_structure(root, user_type='researcher'):
    # Common sections both researchers and students have
    common_sections = ['Personal Information']

    # Researcher-only required sections
    researcher_sections = [
        'Activities',
        'Publications',
    ]

    # Optional — new researchers may not have any grants yet
    # researcher_optional = ['Research Funding History']

    required = common_sections
    if user_type == 'researcher':
        required += researcher_sections

    for section in required:
        if root.find(f'.//section[@label="{section}"]') is None:
            return False, f'Missing required section: {section}'
    return True, None


def process_xml_file(file_obj):
    try:
        tree = ET.parse(file_obj)
        xml_root = tree.getroot()
    except ET.ParseError as e:
        return {
            'filename': file_obj.name,
            'success': False,
            'error': f'Invalid XML: {e}'
        }

    try:
        if "generic-cv" not in xml_root.tag:
            return {
                'filename': file_obj.name,
                'success': False,
                'error': 'Invalid CCV XML format'
            }

        is_valid, error = validate_ccv_structure(xml_root)
        if not is_valid:
            return {
                'filename': file_obj.name,
                'success': False,
                'error': error
            }

        ident_data = {}
        ident_section = xml_root.find('.//section[@label="Identification"]')
        if ident_section is not None:
            ident_data = {
                'first_name':       extract_field_value(ident_section, 'First Name'),
                'last_name':        extract_field_value(ident_section, 'Family Name'),
                'title':            extract_field_value(ident_section, 'Title'),
                'sex':              extract_field_value(ident_section, 'Sex'),
                'language':         extract_field_value(ident_section, 'Correspondence Language'),
                'residency_status': extract_field_value(ident_section, 'Canadian Residency Status'),
            }

        first_name = ident_data.get('first_name')
        last_name  = ident_data.get('last_name')

        if not first_name or not last_name:
            return {
                'filename': file_obj.name,
                'success': False,
                'error': 'Could not find researcher name in XML'
            }

        email = None
        email_section = xml_root.find('.//section[@label="Email"]')
        if email_section is not None:
            email = extract_field_value(email_section, 'Email Address')

        if not email:
            return {
                'filename': file_obj.name,
                'success': False,
                'error': f'No email found for {first_name} {last_name}.'
            }

        org_section  = xml_root.find('.//section[@label="Organization"]')
        organization = extract_field_value(org_section, "Organization") if org_section is not None else ""

        try:
            email = email.strip().lower()
            user  = CustomUser.objects.get(email=email)

            if user.user_type != 'researcher':
                return {
                    'filename': file_obj.name,
                    'success': False,
                    'error': f'{email} exists but is not a researcher. Admin approval required.'
                }

            if user.approval_status == 'rejected':
                return {
                    'filename': file_obj.name,
                    'success': False,
                    'error': f'{first_name} {last_name} ({email}) has been rejected. Cannot import data.',
                }

            if user.approval_status == 'pending':
                user.first_name = first_name
                user.last_name  = last_name
                user.save()
                ResearcherProfile.objects.get_or_create(
                    user=user, defaults={'research_interests': ''}
                )
                return {
                    'filename':       file_obj.name,
                    'success':        True,
                    'researcher':     f"{first_name} {last_name}",
                    'email':          email,
                    'education':      0, 'funding': 0, 'recognitions': 0,
                    'publications':   0, 'activities': 0, 'projects': 0,
                    'supervisor_count': 0, 'total_records': 0,
                    'note': 'Pending approval — data will import after admin approves.',
                }

        except CustomUser.DoesNotExist:
            PENDING_CAP = 20
            if CustomUser.objects.filter(approval_status='pending').count() >= PENDING_CAP:
                return {
                    'filename': file_obj.name,
                    'success': False,
                    'error': 'Too many pending accounts. Approve or reject existing users before uploading more.',
                }

            username = f"{first_name.lower()}.{last_name.lower()}.{uuid.uuid4().hex[:6]}".replace(' ', '')

            user = CustomUser.objects.create_user(
                email=email,
                username=username,
                first_name=first_name,
                last_name=last_name,
                user_type='researcher',
                approval_status='pending',
                organization=organization,
                password=secrets.token_urlsafe(16),
            )
            user.save()

            ResearcherProfile.objects.get_or_create(
                user=user, defaults={'research_interests': ''}
            )

            for admin_user in CustomUser.objects.filter(user_type='admin'):
                StudentNotification.objects.create(
                    user=admin_user,
                    message=f'{first_name} {last_name} ({email}) has been uploaded via CCV and is awaiting approval.',
                )

            return {
                'filename':       file_obj.name,
                'success':        True,
                'researcher':     f"{first_name} {last_name}",
                'email':          email,
                'education':      0, 'funding': 0, 'recognitions': 0,
                'publications':   0, 'activities': 0, 'projects': 0,
                'supervisor_count': 0, 'total_records': 0,
                'note': 'New user created as pending — data will import after admin approves.',
            }

        user.first_name = first_name
        user.last_name  = last_name
        user.save()

        researcher, _ = ResearcherProfile.objects.get_or_create(
            user=user, defaults={'research_interests': ''}
        )
        researcher.title             = ident_data.get('title') or ''
        researcher.sex               = ident_data.get('sex') or ''
        researcher.language          = ident_data.get('language') or ''
        researcher.residency_status  = ident_data.get('residency_status') or ''

        submission = xml_root.find('.//submission')
        if submission is not None:
            researcher.ccv_identifier = submission.get('ccvIdentifier') or ''
        researcher.save()

        try:
            with transaction.atomic():
                education_count    = parse_xml_education(researcher, xml_root)
                funding_count      = parse_xml_funding(researcher, xml_root)
                recognition_count  = parse_xml_recognitions(researcher, xml_root)
                publication_count  = parse_xml_publications(researcher, xml_root)
                activity_count     = parse_xml_activities(researcher, xml_root)
                project_count      = parse_xml_projects(researcher, xml_root)
                supervision_count  = parse_xml_supervision(researcher, xml_root)
        except Exception as e:
            return {
                'filename': file_obj.name,
                'success': False,
                'error': f'Data import failed and was rolled back: {e}'
            }

        return {
            'filename':         file_obj.name,
            'success':          True,
            'researcher':       f"{first_name} {last_name}",
            'email':            email,
            'education':        education_count,
            'funding':          funding_count,
            'recognitions':     recognition_count,
            'publications':     publication_count,
            'activities':       activity_count,
            'projects':         project_count,
            'supervisor_count': supervision_count,
            'total_records': (
                education_count + funding_count + recognition_count +
                publication_count + activity_count + project_count
            ),
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'filename': file_obj.name,
            'success': False,
            'error': f'Processing error: {e}'
        }


@ratelimit(key='ip', rate='2/m', block=True)
@login_required(login_url='login')
@require_http_methods(["POST"])
def bulk_upload(request):
    if getattr(request.user, 'user_type', None) != 'admin':
        return JsonResponse({'success': False, 'error': 'Admins only'}, status=403)

    files = request.FILES.getlist('files')
    if not files:
        return JsonResponse({'success': False, 'error': 'No files provided'}, status=400)

    # File count limit
    MAX_FILES = 10
    if len(files) > MAX_FILES:
        return JsonResponse({'success': False, 'error': f'Maximum {MAX_FILES} files per upload.'}, status=400)

    results = []
    successful = failed = 0

    for file_obj in files:
        filename = file_obj.name

        # Reject non-XML
        if not filename.lower().endswith('.xml'):
            results.append({
                'filename': filename,
                'success': False,
                'error': 'Only XML files are allowed'
            })
            failed += 1
            continue

        # Reject large files
        if file_obj.size > 5 * 1024 * 1024:
            results.append({
                'filename': filename,
                'success': False,
                'error': 'File too large (max 5MB)'
            })
            failed += 1
            continue

        try:
            tree = ET.parse(file_obj)
            xml_root = tree.getroot()

            if "generic-cv" not in xml_root.tag:
                raise ValueError("Invalid CCV XML format")

            is_valid, error = validate_ccv_structure(xml_root)
            if not is_valid:
                results.append({
                    'filename': filename,
                    'success': False,
                    'error': error
                })
                failed += 1
                continue

            file_obj.seek(0)

            # Process as researcher (pending approval handled inside) in atomic transactions

            try:
                with transaction.atomic():
                    result = process_xml_file(file_obj)
                    if not result['success']:
                        raise ValueError(result.get('error', 'Processing failed'))
            except ValueError as e:
                result = {
                    'filename': filename,
                    'success': False,
                    'error': str(e)
                }

        except Exception as e:
            result = {
                'filename': filename,
                'success': False,
                'error': f'Invalid XML: {str(e)}'
            }

        results.append(result)

        if result['success']:
            log_action(
                request,
                'ccv_uploaded',
                summary=f"CCV uploaded: {result.get('researcher', 'Unknown')}",
                details={
                    'filename': filename,
                    'records': result.get('total_records', 0)
                }
            )
            successful += 1
        else:
            failed += 1

    summary = f"Processed {len(files)} file(s): {successful} successful, {failed} failed"

    return JsonResponse({
        'success': successful > 0,
        'message': summary,
        'results': results,
    })

@ratelimit(key='ip', rate='3/m', block=True)
@login_required
@require_http_methods(["POST"])
def student_upload_ccv(request):

    # Role check
    if request.user.user_type != "student":
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)

    files = request.FILES.getlist('files')
    if not files:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

    file_obj = files[0]

    # File type check
    if file_obj.content_type not in ['text/xml', 'application/xml']:
        return JsonResponse({'error': 'Invalid file type'}, status=400)

    # File size check
    if file_obj.size > 5 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'File too large (max 5MB)'}, status=400)

    # ── Parse XML ───────────────────────────────
    try:
        tree = ET.parse(file_obj)
        xml_root = tree.getroot()

        # remove namespaces
        for elem in xml_root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]

        # Validate CCV format
        if "generic-cv" not in xml_root.tag:
            return JsonResponse({'success': False, 'error': 'Invalid CCV format'}, status=400)
        
        is_valid, error = validate_ccv_structure(xml_root, user_type='student')
        if not is_valid:
            return JsonResponse({'success': False, 'error': error}, status=400)

    except ET.ParseError as e:
        return JsonResponse({'success': False, 'error': f'Invalid XML: {e}'}, status=400)

    # ── Validate ownership ───────────────────────
    xml_email_field = xml_root.find('.//field[@label="Email Address"]/value')

    if xml_email_field is None or not xml_email_field.text:
        return JsonResponse({
            'success': False,
            'error': 'Could not verify CCV ownership'
        }, status=403)

    xml_email = xml_email_field.text.strip().lower()
    user_email = request.user.email.strip().lower()

    if xml_email != user_email:
        return JsonResponse({
            'success': False,
            'error': 'This CCV belongs to a different user'
        }, status=403)

    # ── Process XML ──────────────────────────────
    try:
        result = parse_xml_student_profile(request.user, xml_root)

        log_action(
            request,
            'ccv_student_uploaded',
            summary=f'{request.user.get_full_name()} uploaded CCV',
            details={
                'activities':   result['activities'],
                'publications': result['publications'],
                'recognitions': result['recognitions'], 
            }
        )

        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=(
                    f'{request.user.get_full_name()} uploaded their CCV | '
                    f"{result['activities']} activities, "
                    f"{result['publications']} publications, "
                    f"{result['recognitions']} recognitions imported."
                )
            )

        return JsonResponse({
            'success': True,
            'message': (
                f"Profile updated — {result['activities']} activities, "
                f"{result['publications']} publications, "
                f"{result['recognitions']} recognitions imported."
            ),
            'degree_level': result['degree_level'] or '',
            'department':   result['department'] or '',
        })

    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@ratelimit(key='ip', rate='3/m', block=True)
@login_required
@researcher_required
@require_http_methods(["POST"])
def researcher_upload_ccv(request):

    files = request.FILES.getlist('files')
    if not files:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

    file_obj = files[0]

    # File type check
    if file_obj.content_type not in ['text/xml', 'application/xml']:
        return JsonResponse({'error': 'Invalid file type'}, status=400)

    # File size check
    if file_obj.size > 5 * 1024 * 1024:
        return JsonResponse({'success': False, 'error': 'File too large'}, status=400)

    # ── Parse XML ───────────────────────────────
    try:
        tree = ET.parse(file_obj)
        xml_root = tree.getroot()

        # remove namespaces
        for elem in xml_root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]

        # Validate CCV structure (AFTER parsing)
        if "generic-cv" not in xml_root.tag:
            return JsonResponse({'success': False, 'error': 'Invalid CCV format'}, status=400)

        is_valid, error = validate_ccv_structure(xml_root, user_type='researcher')
        if not is_valid:
            return JsonResponse({'success': False, 'error': error}, status=400)

    except ET.ParseError as e:
        return JsonResponse({'success': False, 'error': f'Invalid XML: {e}'}, status=400)

    # ── Validate email ownership ─────────────────
    xml_email_field = xml_root.find('.//field[@label="Email Address"]/value')

    if xml_email_field is None or not xml_email_field.text:
        return JsonResponse({
            'success': False,
            'error': 'Could not verify CCV ownership.'
        }, status=403)

    if xml_email_field.text.strip().lower() != request.user.email.strip().lower():
        return JsonResponse({
            'success': False,
            'error': 'This CCV belongs to a different user.'
        }, status=403)

    # ── Process XML ──────────────────────────────
    try:
        researcher, _ = ResearcherProfile.objects.get_or_create(user=request.user)

        total = 0
        
        try:
            with transaction.atomic():
                total += parse_xml_education(researcher, xml_root)
                total += parse_xml_recognitions(researcher, xml_root)
                total += parse_xml_projects(researcher, xml_root)
                total += parse_xml_funding(researcher, xml_root)
                total += parse_xml_publications(researcher, xml_root)
                total += parse_xml_activities(researcher, xml_root)
                total += parse_xml_supervision(researcher, xml_root)
        except Exception as e:
            return JsonResponse({'error': 'Import failed and was rolled back. Please try again.'}, status=500)

        log_action(
            request,
            'ccv_uploaded',
            summary=f'CCV uploaded: {request.user.get_full_name()}',
            details={'records': total, 'filename': file_obj.name}
        )

        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=f'{request.user.get_full_name()} uploaded their CCV | {total} records imported.'
            )

        return JsonResponse({
            'success': True,
            'message': f'Profile updated — {total} records imported.'
        })

    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)



# ─────────────────────────────────────────────
# Director Reports
# ─────────────────────────────────────────────

@login_required
@admin_required
def reports_list(request):
    from .models import SupervisionRecord

    consenting_students = StudentProfile.objects.filter(
        user__consent_to_share=True,
        degree_level__in=['msc', 'phd'],
    ).count()

    # ── ADD THIS ──────────────────────────────────────────────
    linked_students = SupervisionRecord.objects.filter(
        linked_student__isnull=False,
        status='in_progress',
    ).values('linked_student').distinct().count()

    return render(request, 'Pages/reports/reports_list.html', {
        'consenting_students': consenting_students,
        'linked_students':     linked_students,
    })

@login_required
@admin_required
def enrollment_trends_report(request):
    from datetime import date
    from collections import defaultdict

    current_year  = date.today().year
    n_years       = int(request.GET.get('years', 10))
    filter_pi     = request.GET.get('pi', '')
    filter_degree = request.GET.get('degree', '')
    filter_dept   = request.GET.get('department', '')  # fixed: was 'dept'

    start_year = current_year - n_years + 1

    degree_filter_map = {
        'undergrad': ['bachelors'],
        'msc':       ['masters_thesis', 'masters_non_thesis'],
        'phd':       ['doctorate'],
    }

    def get_level(degree_type):
        if degree_type == 'bachelors':                              return 'undergrad'
        if degree_type in ('masters_thesis', 'masters_non_thesis'): return 'msc'
        if degree_type == 'doctorate':                              return 'phd'
        return 'other'

    records = SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        start_date__isnull=False,
        start_date__year__gte=start_year,   # fixed: respect selected year window
        degree_type__in=['bachelors', 'masters_thesis', 'masters_non_thesis', 'doctorate'],
    ).select_related('researcher__user')

    if filter_pi:
        records = records.filter(researcher__user__id=filter_pi)
    if filter_degree:
        mapped = degree_filter_map.get(filter_degree, [])
        if mapped:
            records = records.filter(degree_type__in=mapped)
    if filter_dept:
        records = records.filter(department__icontains=filter_dept)

    # single DB hit
    record_list = list(records.values(
        'start_date', 'end_date', 'degree_type', 'status', 'department',
        'researcher__user__first_name', 'researcher__user__last_name',
        'researcher__user__id',
    ))

    # precompute per record
    precomputed = []
    for r in record_list:
        s = r['start_date'].year
        e = r['end_date'].year if r['end_date'] else current_year
        precomputed.append({
            'start':      s,
            'end':        e,
            'level':      get_level(r['degree_type']),
            'status':     r['status'],
            'department': r['department'] or '',
            'pi':         f"{r['researcher__user__first_name']} {r['researcher__user__last_name']}".strip(),
            'pi_id':      r['researcher__user__id'],
        })

    # yearly counts
    year_counts = {y: {'undergrad': 0, 'msc': 0, 'phd': 0}
                   for y in range(start_year, current_year + 1)}

    for r in precomputed:
        level = r['level']
        if level == 'other':
            continue
        for y in range(max(r['start'], start_year), min(r['end'], current_year) + 1):
            year_counts[y][level] += 1

    yearly_data = []
    for y in range(start_year, current_year + 1):
        ug  = year_counts[y]['undergrad']
        msc = year_counts[y]['msc']
        phd = year_counts[y]['phd']
        yearly_data.append({
            'year': y, 'undergrad': ug, 'msc': msc, 'phd': phd,
            'total': ug + msc + phd,
        })

    for i, row in enumerate(yearly_data):
        row['yoy'] = 0 if i == 0 else row['total'] - yearly_data[i - 1]['total']

    # summary stats
    total_unique    = len(precomputed)
    total_undergrad = sum(1 for r in precomputed if r['level'] == 'undergrad')
    total_msc       = sum(1 for r in precomputed if r['level'] == 'msc')
    total_phd       = sum(1 for r in precomputed if r['level'] == 'phd')
    total_graduated = sum(1 for r in precomputed if r['status'] == 'completed')
    total_active    = sum(1 for r in precomputed if r['status'] == 'in_progress')

    completed_years = [row for row in yearly_data if row['year'] < current_year]

    first_total  = completed_years[0]['total']  if len(completed_years) >= 2 else 0
    last_total   = completed_years[-1]['total'] if len(completed_years) >= 2 else 0
    growth_delta = last_total - first_total

    if first_total >= 10:
        growth_rate = round(((last_total - first_total) / first_total * 100), 1)
    else:
        growth_rate = None

    growth_end_year = current_year - 1

    # by PI — always set pi_id regardless of level
    by_pi = defaultdict(lambda: {'undergrad': 0, 'msc': 0, 'phd': 0, 'pi_id': None})
    for r in precomputed:
        by_pi[r['pi']]['pi_id'] = r['pi_id']   # fixed: set pi_id for all rows
        if r['level'] in ('undergrad', 'msc', 'phd'):
            by_pi[r['pi']][r['level']] += 1

    by_pi_list = sorted([
        {
            'pi':        k,
            'pi_id':     v['pi_id'],
            'undergrad': v['undergrad'],
            'msc':       v['msc'],
            'phd':       v['phd'],
            'total':     v['undergrad'] + v['msc'] + v['phd'],
        }
        for k, v in by_pi.items()
    ], key=lambda x: x['total'], reverse=True)

    # by department
    by_dept = defaultdict(lambda: {'undergrad': 0, 'msc': 0, 'phd': 0})
    for r in precomputed:
        if r['level'] in ('undergrad', 'msc', 'phd') and r['department']:
            by_dept[r['department']][r['level']] += 1

    by_dept_list = sorted([
        {'dept': k, 'undergrad': v['undergrad'], 'msc': v['msc'], 'phd': v['phd'],
         'total': v['undergrad'] + v['msc'] + v['phd']}
        for k, v in by_dept.items()
    ], key=lambda x: x['total'], reverse=True)

    # department dropdown options
    all_departments = list(
        SupervisionRecord.objects.filter(
            institution__icontains='University of New Brunswick',
            department__isnull=False,
        ).exclude(
            department=''
        ).values_list('department', flat=True).distinct().order_by('department')
    )

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/enrollment_trends_report.html', {
        'yearly_data':     yearly_data,
        'total_unique':    total_unique,
        'total_undergrad': total_undergrad,
        'total_msc':       total_msc,
        'total_phd':       total_phd,
        'total_graduated': total_graduated,
        'total_active':    total_active,
        'growth_rate':     growth_rate,
        'growth_delta': growth_delta,
        'growth_end_year': growth_end_year,
        'start_year':      start_year,
        'current_year':    current_year,
        'n_years':         n_years,
        'by_pi':           by_pi_list,
        'by_dept':         by_dept_list,
        'filter_degree':   filter_degree,
        'filter_pi':       filter_pi,
        'filter_dept':     filter_dept,
        'all_supervisors': all_supervisors,
        'all_departments': all_departments,
        'degree_choices': [
            ('undergrad', 'Undergraduate'),
            ('msc', 'MSc'),
            ('phd', 'PhD'),
        ],
    })


@login_required
@admin_required
def api_enrollment_pi_students(request, pi_id):
    from django.db.models import Case, When, IntegerField, Value
    from datetime import date

    pi = get_object_or_404(CustomUser, id=pi_id, user_type='researcher')

    # respect the same year window as the main report
    n_years    = int(request.GET.get('years', 10))
    start_year = date.today().year - n_years + 1

    records = SupervisionRecord.objects.filter(
        researcher__user=pi,
        start_date__year__gte=start_year,   # fixed: respect selected year window
        degree_type__in=['bachelors', 'masters_thesis', 'masters_non_thesis', 'doctorate'],
    ).select_related(
        'researcher__user',
        'linked_student__user'
    ).annotate(
        degree_order=Case(
            When(degree_type='bachelors',          then=0),
            When(degree_type='masters_thesis',     then=1),
            When(degree_type='masters_non_thesis', then=1),
            When(degree_type='doctorate',          then=2),
            output_field=IntegerField()
        ),
        status_order=Case(
            When(status='in_progress', then=0),
            When(status='completed',   then=1),
            default=Value(2),
            output_field=IntegerField()
        ),
    ).order_by('status_order', 'degree_order', 'start_date')

    students = [{
        'name':       r.student_name,
        'degree':     r.get_degree_type_display() or '',
        'status':     r.get_status_display() or '',
        'start':      r.start_date.strftime('%Y-%m-%d') if r.start_date else '',
        'expected':   r.expected_date.strftime('%Y-%m-%d') if r.expected_date else '',
        'department': r.department or '',
        'linked':     r.linked_student.user.get_full_name()
                      if r.linked_student and r.linked_student.user
                      else None,
    } for r in records]

    return JsonResponse({
        'pi':       pi.get_full_name(),
        'count':    len(students),
        'students': students,
    })


@login_required
@admin_required
def grad_completion_report(request):
    from collections import defaultdict
    from datetime import date
    import json

    cutoff_year  = date.today().year - 10
    current_year = date.today().year

    filter_degree = request.GET.get('degree', '')
    filter_pi     = request.GET.get('pi', '')

    degree_filter_map = {
        'msc': ['masters_thesis', 'masters_non_thesis'],
        'phd': ['doctorate'],
        'pdf': ['postdoc'],
    }

    def get_degree_label(degree_type):
        return {
            'masters_thesis':     'MSc',
            'masters_non_thesis': 'MSc',
            'doctorate':          'PhD',
            'postdoc':            'Post-Doc',
        }.get(degree_type, degree_type)

    def get_degree_level(degree_type):
        return {
            'masters_thesis':     'msc',
            'masters_non_thesis': 'msc',
            'doctorate':          'phd',
            'postdoc':            'pdf',
        }.get(degree_type, '')

    # ── Single DB query for completed records ─────────────────
    sup_records = SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        status='completed',
        degree_start_date__isnull=False,
        degree_type__in=['masters_thesis', 'masters_non_thesis', 'doctorate', 'postdoc'],
        end_date__isnull=False,
        end_date__year__gte=cutoff_year,
        end_date__lte=date.today(),
    ).select_related('researcher__user')

    if filter_degree:
        mapped = degree_filter_map.get(filter_degree, [])
        if mapped:
            sup_records = sup_records.filter(degree_type__in=mapped)
    if filter_pi:
        sup_records = sup_records.filter(researcher__user__id=filter_pi)

    # ── Build records list — single pass ──────────────────────
    records = []
    for r in sup_records:
        grad_date = r.degree_end_date or r.end_date
        start     = r.degree_start_date
        if not start or not grad_date:
            continue
        years = (grad_date - start).days / 365.25
        if years <= 0:
            continue
        records.append({
            'name':            r.student_name,
            'degree':          get_degree_label(r.degree_type),
            'degree_level':    get_degree_level(r.degree_type),
            'department': r.department or '—',
            'start_date':      start,
            'graduation_date': grad_date,
            'years':           round(years, 2),
            'pi':              r.researcher.user.get_full_name(),
            'pi_id':           r.researcher.user.id,
        })

    # ── Summary stats — single pass ───────────────────────────
    total_students = len(records)
    average_time   = round(sum(r['years'] for r in records) / total_students, 2) if total_students else 0

    # ── By degree ─────────────────────────────────────────────
    by_degree = defaultdict(lambda: {'count': 0, 'total': 0.0, 'min': float('inf'), 'max': 0.0})
    for r in records:
        d = r['degree']
        by_degree[d]['count'] += 1
        by_degree[d]['total'] += r['years']
        by_degree[d]['min']    = min(by_degree[d]['min'], r['years'])
        by_degree[d]['max']    = max(by_degree[d]['max'], r['years'])

    by_degree_final = {
        k: {
            'count':   v['count'],
            'average': round(v['total'] / v['count'], 2),
            'min':     round(v['min'], 2),
            'max':     round(v['max'], 2),
        }
        for k, v in sorted(by_degree.items())
    }

    # ── By year ───────────────────────────────────────────────
    by_year = defaultdict(lambda: {'count': 0, 'total': 0.0, 'students': []})
    for r in records:
        y = r['graduation_date'].year
        by_year[y]['count']    += 1
        by_year[y]['total']    += r['years']
        by_year[y]['students'].append(f"{r['name']} ({r['degree']}, {r['years']}y)")

    by_year_final = {
        y: {
            'count':    v['count'],
            'average':  round(v['total'] / v['count'], 2),
            'students': v['students'],
        }
        for y, v in sorted(by_year.items())
    }

    # ── By PI ─────────────────────────────────────────────────
    by_pi = defaultdict(lambda: {'count': 0, 'total': 0.0, 'degrees': set()})
    for r in records:
        by_pi[r['pi']]['count']   += 1
        by_pi[r['pi']]['total']   += r['years']
        by_pi[r['pi']]['degrees'].add(r['degree'])

    by_pi_final = sorted([
        {
            'pi':      k,
            'count':   v['count'],
            'average': round(v['total'] / v['count'], 2),
            'degrees': ', '.join(sorted(v['degrees'])),
        }
        for k, v in by_pi.items()
    ], key=lambda x: x['count'], reverse=True)

    # ── Yearly counts for chart — NO nested loop ──────────────
    # Single query, precompute, then bucket by year
    all_sup = list(SupervisionRecord.objects.filter(
        institution__icontains='University of New Brunswick',
        degree_type__in=['masters_thesis', 'masters_non_thesis', 'doctorate'],
        start_date__isnull=False,
    ).values('start_date', 'end_date', 'degree_type'))

    chart_years = range(current_year - 9, current_year + 1)
    year_buckets = {y: {'msc': 0, 'phd': 0} for y in chart_years}

    for s in all_sup:
        sy = s['start_date'].year
        ey = s['end_date'].year if s['end_date'] else current_year
        level = 'msc' if s['degree_type'] in ('masters_thesis', 'masters_non_thesis') else 'phd'
        for y in range(max(sy, current_year - 9), min(ey, current_year) + 1):
            year_buckets[y][level] += 1

    yearly_counts = [
        {
            'year':  y,
            'msc':   year_buckets[y]['msc'],
            'phd':   year_buckets[y]['phd'],
            'total': year_buckets[y]['msc'] + year_buckets[y]['phd'],
        }
        for y in sorted(chart_years)
    ]

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grad_completion_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Degree', 'Department', 'PI',
                         'Start Date', 'Graduation Date', 'Years'])
        for r in records:
            writer.writerow([
                r['name'], r['degree'], r['department'], r['pi'],
                r['start_date'], r['graduation_date'], r['years'],
            ])
        return response

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/grad_completion_report.html', {
        'records':        records,
        'total_students': total_students,
        'average_time':   average_time,
        'by_degree':      by_degree_final,
        'by_year':        by_year_final,
        'by_pi':          by_pi_final,
        'yearly_counts':  json.dumps(yearly_counts),
        'cutoff_year':    cutoff_year,
        'current_year':   current_year,
        'filter_degree':  filter_degree,
        'filter_pi':      filter_pi,
        'filter_dept':    '',
        'all_supervisors': all_supervisors,
        'all_departments': [],
        'degree_choices': [('msc', 'MSc'), ('phd', 'PhD'), ('pdf', 'Post-Doc')],
    })

from django.db.models import Prefetch

@login_required
@admin_required
def active_projects_report(request):

    filter_pi = request.GET.get('pi', '')

    active_projects = Project.objects.active().select_related(
        'researcher__user'
    ).prefetch_related(
        'team_members',
        Prefetch(
            'tagged_members',
            queryset=CustomUser.objects.select_related('student_profile'),
        ),
        Prefetch(
            'linked_publications',
            queryset=Publication.objects.filter(is_active=True),
            to_attr='active_publications',
        ),
        Prefetch(
            'linked_activities',
            queryset=Activity.objects.filter(is_active=True),
            to_attr='active_activities',
        ),
    ).order_by('researcher__user__last_name', 'title')

    if filter_pi:
        active_projects = active_projects.filter(researcher__user__id=filter_pi)

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="active_projects_report.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Title', 'PI', 'Status', 'Role', 'Funding Org', 'Funding Type',
            'Start Date', 'End Date', 'Currency', 'Total Funding', 'Awarded to UNB', 'Kept by UNB',
            'HQP (Tagged Students)', 'Academic Collaborators',
            'Industry / Community / Govt Partners',
            'Linked Publications', 'KM Activities', 'IP/Patents',
        ])
        for project in active_projects:
            team       = project.team_members.all()
            tagged     = list(project.tagged_members.all())
            hqp_tagged = [u for u in tagged if u.user_type == 'student']
            pubs       = project.active_publications
            acts       = project.active_activities

            pub_count = len(pubs)
            km_count  = sum(1 for a in acts if a.category == 'knowledge_mobilization')
            ip_count  = sum(1 for p in pubs if p.publication_type == 'patent')

            writer.writerow([
                project.title,
                project.researcher.user.get_full_name(),
                project.get_status_display(),
                project.get_role_display(),
                project.funding_organization or '',
                project.funding_type or '',
                project.start_date or '',
                project.end_date or '',
                project.currency or 'CAD',
                project.total_funding or '',
                project.funding_received or '',
                project.funding_kept_by_unb or '',
                '; '.join(u.get_full_name() for u in hqp_tagged),
                '; '.join(m.name for m in team.filter(
                    Q(is_academic_collaborator=True) | Q(partner_type='academic')
                )),
                '; '.join(
                    f"{m.name} ({m.get_partner_type_display()})"
                    for m in team.filter(
                        partner_type__in=['industry', 'community', 'government', 'other']
                    )
                ),
                pub_count, km_count, ip_count,
            ])
        return response

    # ── Build project_data for HTML render ────────────────────
    project_data = []
    total_hqp = total_pubs = total_km = total_ip = 0

    for project in active_projects:
        team   = project.team_members.all()
        tagged = list(project.tagged_members.all())
        pubs   = project.active_publications
        acts   = project.active_activities

        # ── HQP — compute from prefetched data in Python ─────
        hqp_members    = [u for u in tagged if u.user_type == 'student']
        tagged_collabs = [u for u in tagged if u.user_type == 'researcher']
        t_hqp          = len(hqp_members)

        phd_count = msc_count = pdf_count = 0
        for u in hqp_members:
            try:
                dl = u.student_profile.degree_level
                if dl == 'phd':   phd_count += 1
                elif dl == 'msc': msc_count += 1
                elif dl == 'pdf': pdf_count += 1
            except Exception:
                pass
        other_count = t_hqp - phd_count - msc_count - pdf_count

        # ── Counts — Python, no extra queries ────────────────
        km_acts  = [a for a in acts if a.category == 'knowledge_mobilization']
        ip_pubs  = [p for p in pubs if p.publication_type == 'patent']

        pub_count = len(pubs)
        km_count  = len(km_acts)
        ip_count  = len(ip_pubs)

        total_hqp  += t_hqp
        total_pubs += pub_count
        total_km   += km_count
        total_ip   += ip_count

        project_data.append({
            'project':                project,
            'pi_name':                project.researcher.user.get_full_name(),
            'pi_email':               project.researcher.user.email,
            'conception':             project.conception or '',
            'summary':                generate_project_summary(project),
            'total_hqp':              t_hqp,
            'phd_count':              phd_count,
            'msc_count':              msc_count,
            'pdf_count':              pdf_count,
            'other_count':            other_count,
            'hqp_members':            hqp_members,
            'tagged_collabs':         tagged_collabs,
            'academic_collaborators': team.filter(
                Q(is_academic_collaborator=True) | Q(partner_type='academic')
            ),
            'partners': team.filter(
                partner_type__in=['industry', 'community', 'government', 'other']
            ),
            'linked_pubs':  pubs,
            'pub_count':    pub_count,
            'km_acts':      km_acts,
            'all_acts':     acts,
            'km_count':     km_count,
            'ip_pubs':      ip_pubs,
            'ip_count':     ip_count,
        })

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    return render(request, 'Pages/reports/active_projects_report.html', {
        'project_data':    project_data,
        'total_projects':  len(project_data),
        'total_hqp':       total_hqp,
        'total_pubs':      total_pubs,
        'total_km':        total_km,
        'total_ip':        total_ip,
        'all_supervisors': all_supervisors,
        'filter_pi':       filter_pi,
    })


@login_required
@admin_required
def funding_analysis_report(request):
    from collections import defaultdict

    projects = Project.objects.filter(
        ccv_active=True,
        is_deleted=False,
    ).select_related('researcher__user')

    funding_records = Funding.objects.filter(
        is_deleted=False,
    ).select_related('researcher__user', 'project')

    EMPTY_CTX = {
        'error': 'No funding data available',
        'total_funding': '$0', 'by_organization': {}, 'by_year': {},
        'top_orgs': [], 'timeline_data': [], 'active_grants_detail': [],
        'min_year': 2024, 'max_year': 2024,
        'last3_by_org': [], 'last3_by_program': [], 'last3_by_pi': [],
        'last3_as_pi': {'amount': 0, 'count': 0},
        'last3_as_coi': {'amount': 0, 'count': 0},
        'depletion_rows': [], 'depletion_max': 0,
        'total_cad': 0, 'total_usd': 0, 'cad_count': 0, 'usd_count': 0,
        'by_pi_list':   [],
        'current_year': datetime.now().year,
        'last3_start':  datetime.now().year - 3,
        'spark_svg_w':  0,
        'spark_svg_h':  28,
        'year_from':    datetime.now().year,
        'year_to':      datetime.now().year,
    }

    if not projects.exists():
        return render(request, 'Pages/reports/funding_analysis_report.html', EMPTY_CTX)

    try:
        current_year = datetime.now().year
        year_from    = request.GET.get('year_from')
        year_to      = request.GET.get('year_to')

        def parse_year(date_val):
            if not date_val:
                return None
            try:
                s = str(date_val)
                return int(s.split('/')[0] if '/' in s else s[:4])
            except Exception:
                return None

        # ── Deduplicate projects ──────────────────────────────────────
        # Same title + same total_funding = same grant (different amounts = renewal)
        seen = {}
        for p in projects:
            # Use external_id if available, fall back to title-only
            key = p.external_id if p.external_id else p.title.strip().lower()
            if key not in seen:
                seen[key] = p
        projects_deduped = list(seen.values())
        
        all_years    = [y for y in (parse_year(p.start_date) for p in projects_deduped if p.start_date) if y]
        min_year     = min(all_years) if all_years else current_year
        max_year     = max(all_years) if all_years else current_year
        year_from_int = int(year_from) if year_from else min_year
        year_to_int   = int(year_to)   if year_to   else max_year

        # ── KPI totals ────────────────────────────────────────────────
        total_funding_amount = sum(float(p.total_funding or 0) for p in projects_deduped)
        total_grants         = len(projects_deduped)
        average_grant        = total_funding_amount / total_grants if total_grants else 0

        total_cad = sum(float(p.total_funding or 0) for p in projects_deduped if (p.currency or 'CAD') != 'USD')
        total_usd = sum(float(p.total_funding or 0) for p in projects_deduped if p.currency == 'USD')
        cad_count = sum(1 for p in projects_deduped if (p.currency or 'CAD') != 'USD')
        usd_count = sum(1 for p in projects_deduped if p.currency == 'USD')

        # ── By year ───────────────────────────────────────────────────
        by_year = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in projects_deduped:
            y = parse_year(p.start_date)
            if y:
                by_year[y]['amount'] += float(p.total_funding or 0)
                by_year[y]['count']  += 1

        # ── By organization (filtered year range) ─────────────────────
        by_organization = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in projects_deduped:
            sy = parse_year(p.start_date)
            ey = parse_year(p.end_date)
            if not sy or not ey:
                continue
            if sy > year_to_int or ey < year_from_int:
                continue
            org = p.funding_organization or 'Unknown'
            by_organization[org]['amount'] += float(p.total_funding or 0)
            by_organization[org]['count']  += 1

        top_orgs = sorted(
            by_organization.items(), key=lambda x: x[1]['amount'], reverse=True
        )[:10]

        # ── Timeline (secured funding per year) ───────────────────────
        timeline_data = []
        for year in range(year_from_int, year_to_int + 1):
            active_funding = active_count = 0
            for p in projects_deduped:
                sy = parse_year(p.start_date)
                ey = parse_year(p.end_date)
                if sy and ey and sy <= year <= ey:
                    active_funding += float(p.total_funding or 0)
                    active_count   += 1
            timeline_data.append({
                'year':            year,
                'secured_funding': active_funding,
                'active_grants':   active_count,
            })

        # ── Active grants detail table ────────────────────────────────
        active_grants_detail = []
        for p in projects_deduped:
            sy = parse_year(p.start_date)
            ey = parse_year(p.end_date)
            if not sy or not ey:
                continue
            if sy > year_to_int or ey < year_from_int:
                continue
            active_grants_detail.append({
                'title':          p.title,
                'organization':   p.funding_organization or '',
                'pi':             p.researcher.user.get_full_name() if p.researcher and p.researcher.user else '—',
                'role':           p.get_role_display() if p.role else '—',       
                'amount':         int(float(p.total_funding or 0)),
                'amount_to_ibme': int(float(p.funding_received or 0)) if p.funding_received else None,
                'status':         p.get_status_display() if p.status else '',
                'start_year':     sy,
                'end_year':       ey,
                'program':        p.program_name or '',
                'currency':       p.currency or 'CAD',
            })
        active_grants_detail.sort(key=lambda x: x['start_year'])

        # ── Last 3 years ──────────────────────────────────────────────
        last3_start = current_year - 3
        last3_projs = [
            p for p in projects_deduped
            if (parse_year(p.start_date) or 0) >= last3_start
            and p.status not in ('submitted', 'rejected') 
        ]

        PI_ROLES = ('pi', 'pa')

        # Role summary cards
        last3_as_pi  = {'amount': 0, 'count': 0}
        last3_as_coi = {'amount': 0, 'count': 0}
        for p in last3_projs:
            amt = float(p.total_funding or 0)
            if p.role in PI_ROLES:
                last3_as_pi['amount']  += amt
                last3_as_pi['count']   += 1
            else:
                last3_as_coi['amount'] += amt
                last3_as_coi['count']  += 1

        # By funding agency
        last3_org = defaultdict(lambda: {'amount': 0, 'count': 0, 'researchers': set()})
        for p in last3_projs:
            org  = p.funding_organization or 'Unknown'
            name = p.researcher.user.get_full_name() if p.researcher and p.researcher.user else 'Unknown'
            last3_org[org]['amount']      += float(p.total_funding or 0)
            last3_org[org]['count']       += 1
            last3_org[org]['researchers'].add(name)

        last3_by_org = sorted(
            [{'org': k, 'amount': v['amount'], 'count': v['count'],
              'researchers': len(v['researchers'])}
             for k, v in last3_org.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # By program
        last3_prog = defaultdict(lambda: {'amount': 0, 'count': 0})
        for p in last3_projs:
            prog = p.program_name or 'Unspecified'
            last3_prog[prog]['amount'] += float(p.total_funding or 0)
            last3_prog[prog]['count']  += 1

        last3_by_program = sorted(
            [{'program': k, 'amount': v['amount'], 'count': v['count']}
             for k, v in last3_prog.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # By researcher (with role split)
        last3_pi_d = defaultdict(lambda: {
            'amount': 0, 'count': 0, 'orgs': set(), 'pi_count': 0, 'coi_count': 0
        })
        for p in last3_projs:
            name = p.researcher.user.get_full_name() if p.researcher and p.researcher.user else 'Unknown'
            last3_pi_d[name]['amount'] += float(p.total_funding or 0)
            last3_pi_d[name]['count']  += 1
            last3_pi_d[name]['orgs'].add(p.funding_organization or 'Unknown')
            if p.role in PI_ROLES:
                last3_pi_d[name]['pi_count']  += 1
            else:
                last3_pi_d[name]['coi_count'] += 1

        last3_by_pi = sorted(
            [{'pi': k, 'amount': v['amount'], 'count': v['count'],
              'orgs': sorted(v['orgs']),
              'pi_count': v['pi_count'], 'coi_count': v['coi_count']}
             for k, v in last3_pi_d.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # ── Grants by PI (Funding records — correct per-researcher amounts) ──
        by_pi = defaultdict(lambda: {
            'amount': 0, 'amount_to_ibme': 0, 'amount_kept': 0, 'count': 0, 'grants': []
        })
        seen_project_ids = set()
        for f in funding_records:
            name = f.researcher.user.get_full_name() if f.researcher and f.researcher.user else 'Unknown'
            by_pi[name]['amount']         += float(f.amount or 0)
            by_pi[name]['amount_to_ibme'] += float(f.amount_to_ibme or 0)
            by_pi[name]['count']          += 1
            if f.project_id and f.project_id not in seen_project_ids:
                seen_project_ids.add(f.project_id)
                by_pi[name]['amount_kept'] += float(
                    f.project.funding_kept_by_unb or 0
                ) if f.project else 0
            by_pi[name]['grants'].append({
                'amount':         float(f.amount or 0),
                'amount_to_ibme': float(f.amount_to_ibme or 0),
                'grant_total':    float(f.grant_total or f.amount or 0),  # ← add
                'currency':       f.currency or 'CAD',
                'start':          parse_year(f.start_date),
                'end':            parse_year(f.end_date),
                'role':           f.role or '',
            })

        by_pi_list = sorted(
            [{'pi': k, 'amount': v['amount'], 'amount_to_ibme': v['amount_to_ibme'],
              'amount_kept': v['amount_kept'], 'count': v['count'], 'grants': v['grants']}
             for k, v in by_pi.items()],
            key=lambda x: x['amount'], reverse=True
        )

        # ── Depletion waterfall ───────────────────────────────────────
        # Uses ALL deduped projects (not just currently active) so historical
        # snapshot rows are correct.
        # Linear drawdown model: remaining = amount × (end − t) / (end − start)
        # Shows 6 time-offsets per row: Y, Y+1, Y+2, Y+3, Y+4, Y+5

        DEPL_OFFSETS  = 6
        SPARK_H       = 28   # SVG bar height in px
        SPARK_BAR_W   = 16
        SPARK_GAP     = 4
        BAR_COLORS    = ['#C8102E', '#d63f63', '#e46e91', '#ed9ab5', '#f4c2d3', '#f9e3ea']

        all_proj_parsed = [
            {
                'amount': float(p.total_funding or 0),
                'start':  parse_year(p.start_date),
                'end':    parse_year(p.end_date),
            }
            for p in projects_deduped
            if parse_year(p.start_date) and parse_year(p.end_date)
        ]

        # First pass: compute all amounts so we can normalise sparklines
        raw_rows = []
        for snap_year in range(current_year - 4, current_year + 2):
            active_at_snap = [
                p for p in all_proj_parsed
                if p['start'] >= (current_year - 4)
                and p['start'] <= snap_year
                and p['end'] >= snap_year
            ]
            if not active_at_snap:
                continue

            base_amount = sum(p['amount'] for p in active_at_snap)
            years = []
            for offset in range(DEPL_OFFSETS):
                look_year = snap_year + offset
                amount    = 0.0
                count     = 0
                for p in active_at_snap:
                    if p['end'] < look_year:
                        continue
                    duration = p['end'] - p['start']
                    fraction = (
                        1.0 if duration <= 0
                        else max(0.0, min(1.0, (p['end'] - look_year) / duration))
                    )
                    if fraction > 0:
                        amount += p['amount'] * fraction
                        count  += 1
                years.append({
                    'year':        look_year,
                    'amount':      round(amount),
                    'grant_count': count,
                    'pct':         round(amount / base_amount * 100) if base_amount else 0,
                })
            raw_rows.append({
                'snapshot_year': snap_year,
                'base_amount':   base_amount,
                'years':         years,
            })

        depletion_max = max((r['base_amount'] for r in raw_rows), default=1)

        # Second pass: add SVG sparkline coordinates
        depletion_rows = []
        for row in raw_rows:
            for idx, y in enumerate(row['years']):
                bar_h        = max(2 if y['amount'] > 0 else 0,
                                   round(y['amount'] / depletion_max * SPARK_H))
                y['bar_x']   = idx * (SPARK_BAR_W + SPARK_GAP)
                y['bar_y']   = SPARK_H - bar_h
                y['bar_h']   = bar_h
                y['bar_w']   = SPARK_BAR_W
                y['bar_col'] = BAR_COLORS[min(idx, len(BAR_COLORS) - 1)]
            depletion_rows.append(row)

        spark_svg_w = DEPL_OFFSETS * (SPARK_BAR_W + SPARK_GAP) - SPARK_GAP

        return render(request, 'Pages/reports/funding_analysis_report.html', {
            'total_funding':        f"${total_funding_amount:,.0f}",
            'total_funding_raw':    total_funding_amount,
            'total_grants':         total_grants,
            'average_grant':        f"${average_grant:,.0f}",
            'by_organization':      dict(by_organization),
            'by_year':              dict(sorted(by_year.items())),
            'top_orgs':             top_orgs,
            'timeline_data':        timeline_data,
            'active_grants_detail': active_grants_detail,
            'min_year':             min_year,
            'max_year':             max_year,
            'year_from':            year_from_int,
            'year_to':              year_to_int,
            'current_year':         current_year,
            'last3_start':          last3_start,
            'last3_by_org':         last3_by_org,
            'last3_by_program':     last3_by_program,
            'last3_by_pi':          last3_by_pi,
            'last3_as_pi':          last3_as_pi,
            'last3_as_coi':         last3_as_coi,
            'by_pi_list':           by_pi_list,
            'depletion_rows':       depletion_rows,
            'depletion_max':        depletion_max,
            'spark_svg_w':          spark_svg_w,
            'spark_svg_h':          SPARK_H,
            'total_cad':            total_cad,
            'total_usd':            total_usd,
            'cad_count':            cad_count,
            'usd_count':            usd_count,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return render(request, 'Pages/reports/funding_analysis_report.html', {
            **EMPTY_CTX,
            'error': f'Error processing data: {e}',
        })

@login_required
@admin_required
def activity_report(request):
    from django.core.paginator import Paginator

    current_year           = date.today().year
    current_month          = date.today().month
    current_academic_start = current_year if current_month >= 9 else current_year - 1

    year_options = [
        {'value': str(y), 'label': f"{y}/{str(y+1)[-2:]} (Sep {y} – Aug {y+1})"}
        for y in range(current_academic_start, current_academic_start - 6, -1)
    ]

    selected_year      = request.GET.get('year', str(current_academic_start))
    selected_category  = request.GET.get('category', '')
    selected_user_type = request.GET.get('user_type', '')
    selected_pub_type  = request.GET.get('pub_type', '')
    selected_status    = request.GET.get('grant_status', '')
    active_tab         = request.GET.get('tab', 'activities')

    try:
        year_int = int(selected_year)
    except Exception:
        year_int = current_academic_start

    start_dt   = date(year_int, 9, 1)
    end_dt     = date(year_int + 1, 8, 31)
    year_label = f"{year_int}/{str(year_int+1)[-2:]} (Sep {year_int} – Aug {year_int+1})"

    # ── Activities ────────────────────────────────────────────
    activities = Activity.objects.filter(
        date__gte=start_dt, date__lte=end_dt,
        is_active=True,
    ).select_related(
        'researcher__user', 'conference'
    ).prefetch_related(
        'objectives', 'tagged_users'  # ← add tagged_users here
    ).order_by('-date')
    
    if selected_category:
        activities = activities.filter(category=selected_category)
    if selected_user_type:
        activities = activities.filter(researcher__user__user_type=selected_user_type)

    # ── Conference grouping ───────────────────────────────────
    group_by_conf     = request.GET.get('group') == '1'
    grouped_conferences = []

    if group_by_conf:
        from collections import defaultdict

        conf_map = defaultdict(lambda: {
            'conference':       None,
            'name':             '',
            'year':             None,
            'location':         '',
            'attendees':        [],
            'seen_users':       set(),
            'student_count':    0,
            'researcher_count': 0,
            'msc_count':        0,
            'phd_count':        0,
            'other_count':      0,
            'earliest_date':    None,
        })

        # FIX 3: use start_dt from selected year, not hardcoded 3yr window
        conf_activities = activities.filter(
            category='conference',
            date__gte=start_dt,
        ).select_related('researcher__user', 'conference')

        for a in conf_activities:
            if a.conference_id:
                key = f'conf-{a.conference_id}'
            else:
                key = f'no-conf-{a.title}-{a.date}'

            entry = conf_map[key]

            if a.conference and not entry['conference']:
                entry['conference'] = a.conference
                entry['name']       = a.conference.name
                entry['year']       = a.conference.year
                entry['location']   = a.conference.location or ''
            elif not entry['name']:
                entry['name'] = a.title

            user_id = a.researcher.user.id

            if user_id not in entry['seen_users']:
                entry['seen_users'].add(user_id)
                utype = a.researcher.user.user_type

                entry['attendees'].append({
                    'name':      a.researcher.user.get_full_name(),
                    'user_type': utype,
                    'invited':   a.invited,
                    'keynote':   a.keynote,
                })

                if utype == 'student':
                    entry['student_count'] += 1

                    # FIX 2: removed broken select_related, access directly with try/except
                    try:
                        degree = a.researcher.user.student_profile.degree_level
                        if degree == 'msc':
                            entry['msc_count'] += 1
                        elif degree == 'phd':
                            entry['phd_count'] += 1
                        else:
                            entry['other_count'] += 1
                    except Exception:
                        entry['other_count'] += 1
                else:
                    entry['researcher_count'] += 1

            if not entry['earliest_date'] or (a.date and a.date < entry['earliest_date']):
                entry['earliest_date'] = a.date

        grouped_conferences = sorted(
            conf_map.values(),
            key=lambda x: x['earliest_date'] or date.min,
            reverse=True,
        )

    # ── Publications ──────────────────────────────────────────
    publications = Publication.objects.filter(
        publication_date__gte=start_dt,
        publication_date__lte=end_dt,
        is_active=True,
    ).select_related('researcher__user').order_by('-publication_date')

    if selected_pub_type:
        publications = publications.filter(publication_type=selected_pub_type)

    # ── Grants ────────────────────────────────────────────────
    grants = Project.objects.filter(
        start_date__gte=start_dt,
        start_date__lte=end_dt,
        is_deleted=False,
    ).select_related('researcher__user').prefetch_related(
        'funding_breakdown'
    ).order_by('-start_date')

    if selected_status:
        grants = grants.filter(status=selected_status)

    # ── CSV exports ───────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        if active_tab == 'publications':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="publications_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Title', 'Type', 'Status', 'Journal', 'Authors', 'Date', 'DOI', 'Researcher'])
            for p in publications:
                writer.writerow([
                    p.title, p.get_publication_type_display(),
                    p.get_status_display() if p.status else '',
                    p.journal or '', p.authors or '',
                    p.publication_date or '', p.doi or '',
                    p.researcher.user.get_full_name(),
                ])
            return response

        elif active_tab == 'grants':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="grants_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow([
                'Title', 'Organization', 'Program', 'Role', 'Status',
                'Total Funding', 'Awarded to UNB', 'Kept by UNB',
                'Start', 'End', 'PI',
            ])
            # FIX 1: removed dead grant_program_lookup — use g.program_name directly
            for g in grants:
                writer.writerow([
                    g.title,
                    g.funding_organization or '',
                    g.program_name or '',
                    g.get_role_display() if g.role else '',
                    g.get_status_display() if g.status else '',
                    g.total_funding or '',
                    g.funding_received or '',
                    g.funding_kept_by_unb or '',
                    g.start_date or '',
                    g.end_date or '',
                    g.researcher.user.get_full_name(),
                ])
            return response

        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="activities_{selected_year}.csv"'
            writer = csv.writer(response)
            writer.writerow(['Title', 'Category', 'Type', 'Person', 'Email', 'Role', 'Date', 'Description'])
            for a in activities:
                writer.writerow([
                    a.title,
                    a.get_category_display() if hasattr(a, 'get_category_display') else a.category,
                    a.get_activity_type_display(),
                    a.researcher.user.get_full_name(),
                    a.researcher.user.email,
                    a.researcher.user.user_type.title(),
                    a.date.strftime('%Y-%m-%d') if a.date else '',
                    a.description or '',
                ])
            return response

    # ── Stats ─────────────────────────────────────────────────
    total_count      = activities.count()
    conference_count = activities.filter(category='conference').count()
    km_count         = activities.filter(category='knowledge_mobilization').count()
    media_count      = activities.filter(category='media').count()
    researcher_count = activities.filter(
        researcher__user__user_type='researcher').values('researcher').distinct().count()
    student_count    = activities.filter(
        researcher__user__user_type='student').values('researcher').distinct().count()

    
    # ── Strategic objectives breakdown ───────────────────────
    objective_counts = []
    for obj in StrategicObjective.objects.all():
        count = obj.activities.filter(
            date__gte=start_dt,
            date__lte=end_dt,
            is_active=True,
        ).count()
        if count > 0:
            objective_counts.append({'name': obj.name, 'count': count})

    paginator  = Paginator(activities, 25)
    page_obj   = paginator.get_page(request.GET.get('page', 1))
    page_start = (page_obj.number - 1) * 25 + 1
    page_end   = min(page_obj.number * 25, total_count)

    pub_paginator = Paginator(publications, 25)
    pub_page_obj  = pub_paginator.get_page(request.GET.get('pub_page', 1))

    grant_paginator = Paginator(grants, 25)
    grant_page_obj  = grant_paginator.get_page(request.GET.get('grant_page', 1))

    export_params = (
        f"year={selected_year}&category={selected_category}"
        f"&user_type={selected_user_type}&pub_type={selected_pub_type}"
        f"&grant_status={selected_status}&tab={active_tab}"
    )

    return render(request, 'Pages/reports/activity_report.html', {
        'activities':         page_obj,
        'publications':       pub_page_obj,
        'grants':             grant_page_obj,
        'year_options':       year_options,
        'selected_year':      selected_year,
        'selected_category':  selected_category,
        'selected_user_type': selected_user_type,
        'selected_pub_type':  selected_pub_type,
        'selected_status':    selected_status,
        'active_tab':         active_tab,
        'year_label':         year_label,
        'total_count':        total_count,
        'conference_count':   conference_count,
        'km_count':           km_count,
        'media_count':        media_count,
        'researcher_count':   researcher_count,
        'student_count':      student_count,
        'pub_count':          publications.count(),
        'grant_count':        grants.count(),
        'awarded_count':      grants.filter(status='awarded').count(),
        'submitted_count':    grants.filter(status='submitted').count(),
        'rejected_count':     grants.filter(status='rejected').count(),
        'page_start':         page_start,
        'page_end':           page_end,
        'export_params':      export_params,
        'group_by_conf':      group_by_conf,
        'grouped_conferences': grouped_conferences,
        'objective_counts': objective_counts,
    })

@login_required
@admin_required
def export_report_csv(request, report_type):
    return JsonResponse({'error': 'Export feature coming soon'})


# ─────────────────────────────────────────────
# Publications
# ─────────────────────────────────────────────

def auto_link_publication_authors(publication):
    if not publication.authors:
        return

    from .models import PublicationAuthor, StudentProfile, ResearcherProfile

    raw = publication.authors.replace('*', '').strip()
    if ';' in raw:
        authors_list = [a.strip().lower() for a in raw.split(';') if a.strip()]
    else:
        authors_list = [a.strip().lower() for a in raw.split(',') if a.strip()]

    # ── Extract last names from author strings to pre-filter ──
    candidate_last_names = set()
    for author in authors_list:
        if ',' in author:
            candidate_last_names.add(author.split(',')[0].strip())
        else:
            parts = author.split()
            if parts:
                candidate_last_names.add(parts[-1].strip())

    if not candidate_last_names:
        return

    def is_match(user):
        full_name = user.get_full_name().lower().strip()
        last_name = user.last_name.lower().strip()
        first_name = user.first_name.lower().strip()

        if not full_name or not last_name or len(last_name) < 3:
            return False

        for author in authors_list:
            author = author.strip()

            # Exact full name match
            if full_name == author:
                return True

            # Handle "LastName, FirstName" or "LastName, F" format
            if ',' in author:
                parts = [p.strip() for p in author.split(',', 1)]
                a_last = parts[0]
                a_first = parts[1] if len(parts) > 1 else ''

                if a_last == last_name:
                    if not a_first:
                        return True
                    if first_name and (first_name == a_first or first_name[0] == a_first[0]):
                        return True
                continue

            # Handle "FirstName LastName" format
            author_parts = author.split()
            if len(author_parts) >= 2:
                if last_name == author_parts[-1]:
                    if first_name and (first_name == author_parts[0] or first_name[0] == author_parts[0][0]):
                        return True

            # Single last name only
            if len(author.split()) == 1 and author == last_name:
                return True

        return False

    # ── Only load users whose last name appears in the authors string ──
    for student in StudentProfile.objects.select_related('user').filter(
        user__last_name__in=candidate_last_names
    ):
        if student.user.id == publication.researcher.user.id:
            continue
        if is_match(student.user):
            PublicationAuthor.objects.get_or_create(
                publication=publication,
                student=student,
            )

    for researcher in ResearcherProfile.objects.select_related('user').filter(
        user__last_name__in=candidate_last_names
    ):
        if researcher.id == publication.researcher.id:
            continue
        if is_match(researcher.user):
            StudentNotification.objects.get_or_create(
                user=researcher.user,
                message=f'You are listed as a co-author on "{publication.title}" by {publication.researcher.user.get_full_name()}.',
            )

def log_action(request, action, target=None, summary=None, details=None):

    from .models import AuditLog
    ip = None
    if request:
        ip = (
            request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
            or request.META.get('REMOTE_ADDR')
        )
    AuditLog.objects.create(
        user        = request.user if request and request.user.is_authenticated else None,
        action      = action,
        target_type = target.__class__.__name__ if target else None,
        target_id   = target.pk if target else None,
        summary     = summary or '',
        details     = details or {},
        ip_address  = ip,
    )

@login_required
def add_publication(request):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
    except ResearcherProfile.DoesNotExist:
        researcher = ResearcherProfile.objects.create(user=request.user)

    if request.method == 'POST':
        try:
            title            = request.POST.get('title', '').strip()
            authors          = request.POST.get('authors', '').strip()
            journal          = request.POST.get('journal', '').strip()
            publication_date = request.POST.get('publication_date', '').strip()

            if not all([title, authors, journal, publication_date]):
                return JsonResponse({'success': False,
                                     'error': 'Title, Authors, Journal, and Date are required'}, status=400)

            # ── Duplicate check ───────────────────────────────
            existing = Publication.objects.filter(
                researcher=researcher,
                title__iexact=title,
                publication_date=publication_date,
                is_active=True,
            ).first()

            if existing:
                return JsonResponse({
                    'success': False,
                    'error': 'You already have a publication with this title and date.'
                }, status=400)

            pub = Publication.objects.create(
                researcher=researcher,
                title=title,
                authors=authors,
                journal=journal,
                publication_date=publication_date,
                doi=request.POST.get('doi'),
                url=request.POST.get('url'),
                abstract=request.POST.get('abstract'),
                publication_type=request.POST.get('publication_type', 'journal'),
                status=request.POST.get('status', 'published'),
                source='manual',
            )

            log_action(request, 'publication_added', target=pub,
                        summary=f'{request.user.get_full_name()} added "{title}"')
            
            for admin_user in CustomUser.objects.filter(user_type='admin'):
                StudentNotification.objects.create(
                    user=admin_user,
                    message=f'{request.user.get_full_name()} added a publication "{title}".',
                )

            auto_link_publication_authors(pub)
            
            return JsonResponse({'success': True, 'message': f'Publication "{title}" added!'})
        
        except Exception as e:
            return JsonResponse({'error': 'Internal server error'}, status=500)

    return render(request, 'Pages/forms/add_publication.html', {'researcher': researcher})


from django.shortcuts import render

@login_required
def view_linked_publications(request):
    if request.user.user_type != 'student':
        return redirect('dashboard')

    try:
        student = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        return render(request, 'Pages/forms/linked_publications.html', {
            'publications': [], 'count': 0, 'peer_count': 0,
        })

    VISIBLE_STATUSES = ['published', 'accepted', 'under_review']

    if student.supervisor:
        peer_students = StudentProfile.objects.filter(
            supervisor=student.supervisor
        ).exclude(user=request.user)

        peer_researcher_ids = list(
            peer_students.values_list('user__researcherprofile__id', flat=True)
        )

        co_supervisor_ids = list(
            student.co_supervisors.values_list('id', flat=True)
        )

        group_ids = [student.supervisor.id] + co_supervisor_ids + [r for r in peer_researcher_ids if r]

        publications = Publication.objects.filter(
            is_active=True,
            researcher__id__in=group_ids,
            status__in=VISIBLE_STATUSES,
        ).distinct().select_related(
            'researcher__user'
        ).order_by('-publication_date')

        peer_count = peer_students.count()

    elif student.department:
        publications = Publication.objects.filter(
            is_active=True,
            researcher__user__student_profile__department=student.department,
            status__in=VISIBLE_STATUSES,
        ).distinct().select_related(
            'researcher__user'
        ).order_by('-publication_date')

        peer_count = StudentProfile.objects.filter(
            department=student.department
        ).exclude(user=request.user).count()

    else:
        publications = Publication.objects.none()
        peer_count = 0

    return render(request, 'Pages/forms/linked_publications.html', {
        'publications': publications,
        'count':        publications.count(),
        'peer_count':   peer_count,
        'supervisor':   student.supervisor,
        'no_context':   not student.supervisor and not student.department,
    })

@login_required
def view_publications(request):
    try:
        researcher   = ResearcherProfile.objects.get(user=request.user)
        publications = Publication.objects.filter(researcher=researcher).distinct()
    except ResearcherProfile.DoesNotExist:
        publications = Publication.objects.none()

    sort_by = request.GET.get("sort", "recent")
    order   = {"oldest": "publication_date", "title-asc": "title",
               "title-desc": "-title"}.get(sort_by, "-publication_date")
    publications = publications.order_by(order)

    search_query = request.GET.get("search")
    if search_query:
        q = Q()
        if " AND " in search_query:
            for part in search_query.split(" AND "):
                q &= Q(title__icontains=part.strip()) | Q(authors__icontains=part.strip()) | Q(journal__icontains=part.strip())
        elif " OR " in search_query:
            for part in search_query.split(" OR "):
                q |= Q(title__icontains=part.strip()) | Q(authors__icontains=part.strip()) | Q(journal__icontains=part.strip())
        else:
            q = Q(title__icontains=search_query) | Q(authors__icontains=search_query) | Q(journal__icontains=search_query)
        publications = publications.filter(q)

    return render(request, 'Pages/forms/view_publications.html', {
        'publications': publications, 'publication_count': publications.count()
    })


@login_required
def delete_publication(request, publication_id):
    try:
        researcher  = ResearcherProfile.objects.get(user=request.user)
        publication = Publication.objects.get(id=publication_id, researcher=researcher)
        title       = publication.title
        publication.soft_delete(user=request.user)
        log_action(request, 'publication_deleted', target=publication,
            summary=f'{request.user.get_full_name()} deleted "{title}"')
        return JsonResponse({'success': True, 'message': f'Publication "{title}" deleted.'})
    except Publication.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Publication not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_activity(request, activity_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        activity = Activity.objects.get(id=activity_id, researcher=researcher)
        title = activity.title

        # Notify tagged users before deleting
        tagged_users = list(activity.tagged_users.all())
        for u in tagged_users:
            if u.user_type == 'researcher':
                msg = f'{request.user.get_full_name()} deleted "{title}" which you were tagged in.'
            else:
                msg = f'{request.user.get_full_name()} deleted "{title}" which you were tagged in. You may want to log it yourself.'
    
            StudentNotification.objects.create(user=u, message=msg)

        activity.soft_delete(user=request.user)
        log_action(request, 'activity_deleted', target=activity,
                   summary=f'{request.user.get_full_name()} deleted activity "{title}"')
        return JsonResponse({'success': True, 'message': f'Activity "{title}" deleted.'})
    except Activity.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Activity not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@researcher_required
@require_http_methods(["POST"])
def delete_project(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project = Project.objects.get(id=project_id, researcher=researcher)
        title = project.title
        project.soft_delete(user=request.user)
        log_action(request, 'project_deleted', target=project,
                   summary=f'{request.user.get_full_name()} deleted project "{title}"')
        return JsonResponse({'success': True, 'message': f'Project "{title}" deleted.'})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
def export_publications_csv(request):
    try:
        researcher = request.user.researcherprofile
    except ResearcherProfile.DoesNotExist:
        return HttpResponse("No researcher profile found.", status=404)

    publications = Publication.objects.filter(
        researcher=researcher, is_active=True
    ).order_by('-publication_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="publications.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Title', 'Authors', 'Type', 'Status', 'Journal / Conference / Publisher',
        'Publication Date', 'Year', 'Volume', 'Issue', 'Pages', 'DOI', 'Language',
        'Refereed', 'Open Access', 'Contribution Role', 'Conference Location',
        'Book Title', 'Editors', 'Patent Number', 'Country', 'Filing Date',
    ])
    for p in publications:
        writer.writerow([
            p.title, p.authors or '', p.get_publication_type_display(), p.get_status_display(),
            p.journal or '', p.publication_date or '',
            str(p.publication_date)[:4] if p.publication_date else '',
            p.volume or '', p.issue or '', p.pages or '', p.doi or '', p.language or '',
            'Yes' if p.refereed else ('No' if p.refereed is False else ''),
            'Yes' if p.open_access else ('No' if p.open_access is False else ''),
            p.contribution_role or '', p.conference_location or '', p.book_title or '',
            p.editors or '', p.patent_number or '', p.country or '', p.filing_date or '',
        ])
    return response

@login_required
def export_activities_csv(request):
    try:
        researcher = request.user.researcherprofile
    except ResearcherProfile.DoesNotExist:
        return HttpResponse("No researcher profile found.", status=404)

    activities = Activity.objects.filter(
        researcher=researcher,
        is_active=True
    ).order_by('-date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="activities.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Title',
        'Type',
        'Category',
        'Date',
        'Year',
        'Location',
        'Invited',
        'Keynote',
        'Co-presenters',
        'Audience',
        'Description',
    ])

    for a in activities:
        writer.writerow([
            a.title,
            a.get_activity_type_display(),
            a.get_category_display() if a.category else '',
            a.date or '',
            str(a.date)[:4] if a.date else '',
            a.location or '',
            'Yes' if a.invited else ('No' if a.invited is False else ''),
            'Yes' if a.keynote else ('No' if a.keynote is False else ''),
            a.co_presenters or '',
            a.audience or '',
            a.description or '',
        ])

    return response

# ─────────────────────────────────────────────
# Projects
# ─────────────────────────────────────────────

from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
@researcher_required
def add_project(request):

    def parse_date(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date() if d else None
        except:
            return None

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            title                = data.get('title', '').strip()
            description          = data.get('description', '').strip()
            funding_organization = data.get('funding_organization', '').strip()
            funding_type         = data.get('funding_type', '').strip()
            total_funding        = data.get('total_funding')
            funding_received     = data.get('funding_received')
            start_date           = data.get('start_date')
            end_date             = data.get('end_date')
            role                 = data.get('role', 'pi')
            status               = data.get('status', 'awarded')
            conception           = data.get('conception', '').strip()
            next_steps           = data.get('next_steps', '').strip()
            members              = data.get('members', [])
            hqp_ids              = data.get('hqp_ids', [])
            currency             = data.get('currency', 'CAD')

            if not title:
                return JsonResponse({'success': False, 'error': 'Title is required'}, status=400)

            researcher, _ = ResearcherProfile.objects.get_or_create(user=request.user)

            VALID_PARTNER_TYPES = ['academic', 'industry', 'community', 'government', 'other']
            VALID_ROLES = ['pi', 'co_pi', 'pa', 'co_app', 'other']

            with transaction.atomic():

                # ── Create Project ─────────────────────────────
                project = Project.objects.create(
                    researcher           = researcher,
                    title                = title,
                    description          = description or None,
                    funding_organization = funding_organization or None,
                    funding_type         = funding_type or None,
                    total_funding        = Decimal(total_funding) if total_funding else None,
                    funding_received     = Decimal(funding_received) if funding_received else None,
                    start_date           = parse_date(start_date),
                    end_date             = parse_date(end_date),
                    role                 = role,
                    status               = status,
                    conception           = conception or None,
                    next_steps           = next_steps or None,
                    source               = 'manual',
                    ccv_active           = True,
                    currency = currency,
                )

                # ── External collaborators / partners ─────────
                seen = set()

                for m in members:
                    name = (m.get('name') or '').strip()
                    key = name.lower()

                    if not name or key in seen:
                        continue
                    seen.add(key)

                    member_role  = m.get('role', 'other')
                    partner_type = m.get('partner_type', 'academic')

                    if member_role not in VALID_ROLES:
                        member_role = 'other'

                    if partner_type not in VALID_PARTNER_TYPES:
                        partner_type = 'other'

                    ProjectMember.objects.create(
                        project      = project,
                        name         = name,
                        role         = member_role,
                        partner_type = partner_type,
                        is_academic_collaborator = (partner_type == 'academic'),
                    )

                # ── HQP tagging ───────────────────────────────
                if hqp_ids:
                    tagged = CustomUser.objects.filter(
                        id__in=hqp_ids,
                        user_type='student',
                        approval_status='approved'
                    )

                    project.tagged_members.set(tagged)

                    for u in tagged:
                        StudentNotification.objects.create(
                            user=u,
                            message=f'{request.user.get_full_name()} has added you as HQP on the project "{title}".'
                        )

                # ── Logging ───────────────────────────────────
                log_action(
                    request,
                    'project_created',
                    target=project,
                    summary=f'{request.user.get_full_name()} created project "{title}"'
                )

                # ── Notify admins ─────────────────────────────────
                for admin_user in CustomUser.objects.filter(user_type='admin'):
                    StudentNotification.objects.create(
                        user=admin_user,
                        message=f'{request.user.get_full_name()} created a new project "{title}".'
                    )

            return JsonResponse({
                'success': True,
                'message': f'Project "{title}" created successfully.',
                'project_id': project.id,
            })

        except Exception as e:
            return JsonResponse({'error': 'Internal server error'}, status=500)

    # ── GET ──────────────────────────────────────────────────
    students = CustomUser.objects.filter(
        user_type='student',
        approval_status='approved',
    ).order_by('last_name', 'first_name')

    return render(request, 'Pages/projects/add_project.html', {
        'students': students,
    })



# ─────────────────────────────────────────────
# Activities
# ─────────────────────────────────────────────

from django.db.models import Q

@login_required
def view_activities(request):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        activities = Activity.objects.filter(
            Q(researcher=researcher) | Q(tagged_users=request.user),
            is_active=True
        ).prefetch_related('objectives').distinct()
    except ResearcherProfile.DoesNotExist:
        activities = Activity.objects.none()

    sort_by = request.GET.get("sort", "recent")
    order = {
        "oldest": "date",
        "title-asc": "title",
        "title-desc": "-title"
    }.get(sort_by, "-date")
    activities = activities.order_by(order)

    search_query = request.GET.get("search")
    if search_query:
        q = Q()
        if " AND " in search_query:
            for part in search_query.split(" AND "):
                q &= Q(title__icontains=part.strip())
        elif " OR " in search_query:
            for part in search_query.split(" OR "):
                q |= Q(title__icontains=part.strip())
        else:
            q = Q(title__icontains=search_query)
        activities = activities.filter(q)

    return render(request, 'Pages/forms/view_activities.html', {
        'activities': activities
    })


@login_required
def api_get_objectives(request):
    from .models import StrategicObjective
    return JsonResponse({
        'objectives': [
            {'id': o.id, 'name': o.name}
            for o in StrategicObjective.objects.all()
        ]
    })

@login_required
@require_http_methods(["POST"])
def api_log_activity(request):
    if request.user.user_type not in ('student', 'researcher'):
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)
    try:
        data          = json.loads(request.body)
        title         = data.get('title', '').strip()
        activity_type = data.get('activity_type', 'presentation')
        date_str      = data.get('date', '')
        location      = data.get('location', '')
        description   = data.get('description', '')
        invited       = data.get('invited')
        keynote       = data.get('keynote')
        conference_id = data.get('conference_id') or None
        co_presenters = data.get('co_presenters', '')
        force         = data.get('force', False)

        

        if not title or not date_str:
            return JsonResponse({'success': False, 'error': 'Title and date are required'}, status=400)

        try:
            activity_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format'}, status=400)

        category = data.get('category', '').strip()
        category_was_inferred = False

        if not category:
            category = determine_activity_category(title, None, None, location)
            category_was_inferred = True

        # Only enforce conference requirement if user explicitly chose it
        if category == 'conference' and not conference_id and not category_was_inferred:
            return JsonResponse({
                'success': False,
                'error': 'Please select or create a conference for conference activities.'
            }, status=400)
        
        researcher = ResearcherProfile.objects.filter(user=request.user).first()
        if not researcher:
            researcher = ResearcherProfile.objects.create(
                user=request.user, research_interests=''
            )

        # ── Merged check: already owner OR tagged ─────────────
        already_associated = Activity.objects.filter(
            date=activity_date,
            title__iexact=title,
            is_active=True,
        ).filter(
            Q(researcher=researcher) | Q(tagged_users=request.user)
        ).first()

        if already_associated:
            return JsonResponse({
                'success': False,
                'error': 'This activity is already on your profile.',
            }, status=400)

        # ── Similar activity by someone else ──────────────────
        if not force:
            similar_qs = Activity.objects.filter(
                date=activity_date,
                title__iexact=title,
                is_active=True,
            ).exclude(researcher=researcher)

            # ── FIX: include conference in match if provided ──
            if conference_id:
                similar_qs = similar_qs.filter(conference_id=conference_id)

            similar = similar_qs.first()

            if similar:
                return JsonResponse({
                    'success': False,
                    'similar_found': True,
                    'similar': {
                        'id':        similar.id,
                        'title':     similar.title,
                        'date':      str(similar.date),
                        'location':  similar.location or '',
                        'logged_by': similar.researcher.user.get_full_name(),
                    },
                    'message': 'A similar activity was found. Were you at the same event?'
                })

        # ── Create activity ───────────────────────────────────
        activity = Activity.objects.create(
            researcher=researcher,
            activity_type=activity_type,
            title=title,
            description=description,
            date=activity_date,
            location=location or None,
            invited=invited,
            keynote=keynote,
            category=category,
            source='manual',
            is_active=True,
            co_presenters=co_presenters or '',
            conference_id=conference_id or None,
        )

        objective_ids = data.get('objectives', [])

        if objective_ids:

            objs = StrategicObjective.objects.filter(id__in=objective_ids)
            activity.objectives.set(objs)

        log_action(request, 'activity_submitted', target=activity,
                   summary=f'{request.user.get_full_name()} logged "{activity.title}"')

        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=f'{request.user.get_full_name()} logged a new activity "{activity.title}".',
            )

        tagged_ids = data.get('tagged_users', [])
        if tagged_ids:
            User = get_user_model()
            tagged = User.objects.filter(id__in=tagged_ids)
            activity.tagged_users.set(tagged)
            for u in tagged:
                StudentNotification.objects.create(
                    user=u,
                    message=f'{request.user.get_full_name()} tagged you in activity "{activity.title}".'
                )

        return JsonResponse({
            'success': True,
            'message': 'Activity logged successfully.',
            'activity_id': activity.id
        })

    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_tag_me_on_activity(request, activity_id):
    """Student confirms they attended the same event — tags themselves on existing activity."""
    try:
        activity = Activity.objects.get(id=activity_id, is_active=True)

        if activity.researcher.user == request.user:
            return JsonResponse({'success': False, 'error': 'This is your own activity.'}, status=400)

        # ── FIX: explicit duplicate tag check ─────────────────
        if activity.tagged_users.filter(id=request.user.id).exists():
            return JsonResponse({
                'success': False,
                'error': 'You are already linked to this activity.'
            }, status=400)

        activity.tagged_users.add(request.user)

        StudentNotification.objects.create(
            user=activity.researcher.user,
            message=f'{request.user.get_full_name()} linked themselves to your activity "{activity.title}".'
        )

        return JsonResponse({
            'success': True,
            'message': f'You\'ve been linked to "{activity.title}".'
        })
    except Activity.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Activity not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def api_search_conferences(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})
    from .models import Conference
    conferences = Conference.objects.filter(
        Q(name__icontains=query) | Q(acronym__icontains=query)
    ).order_by('-year')[:10]
    return JsonResponse({'results': [{
        'id':       c.id,
        'name':     c.name,
        'acronym':  c.acronym or '',
        'year':     c.year or '',
        'location': c.location or '',
        'display':  str(c),
    } for c in conferences]})

@login_required
@require_http_methods(["POST"])
def api_create_conference(request):
    from .models import Conference
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'Name required'}, status=400)
        # parse year from name if present e.g. "EMBC 2024"
        import re
        year_match = re.search(r'\b(20\d{2})\b', name)
        year = int(year_match.group(1)) if year_match else None
        conf, created = Conference.objects.get_or_create(
            name=name,
            defaults={'year': year}
        )
        return JsonResponse({'success': True, 'id': conf.id, 'display': str(conf)})
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_http_methods(["POST"])
def api_review_activity(request, activity_id):
    from django.utils import timezone
    if request.user.user_type != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin only'}, status=403)
    try:
        review       = ActivityReview.objects.get(activity_id=activity_id)
        data         = json.loads(request.body)
        action       = data.get('action')
        reason       = data.get('reason', '').strip()
        student_user = review.activity.researcher.user

        if action == 'approve':
            review.status = 'approved'
            review.reviewed_by = request.user
            review.reviewed_at = timezone.now()
            review.save()

            log_action(request, 'activity_approved', target=review.activity,
                        summary=f'Approved "{review.activity.title}" by {student_user.get_full_name()}')

            StudentNotification.objects.create(
                user=student_user,
                message=f"Your activity \"{review.activity.title}\" was approved.",
            )

            # ── Copy to tagged researchers ─────────────────────────
            tagged_ids = review.activity.tagged_users.values_list('id', flat=True)
            for tagged_user in CustomUser.objects.filter(id__in=tagged_ids, user_type='researcher'):
                try:
                    researcher_profile = ResearcherProfile.objects.get(user=tagged_user)
                    obj, created = Activity.objects.get_or_create(
                        researcher=researcher_profile,
                        title=review.activity.title,
                        date=review.activity.date,
                        defaults={
                            'activity_type': review.activity.activity_type,
                            'category':      review.activity.category,
                            'description':   review.activity.description,
                            'location':      review.activity.location,
                            'invited':       review.activity.invited,
                            'keynote':       review.activity.keynote,
                            'co_presenters': review.activity.co_presenters,
                            'audience':      review.activity.audience,
                            'source':        'manual',
                            'is_active':     True,
                        }
                    )
                    if not created and not obj.is_active:
                        obj.is_active = True
                        obj.save(update_fields=['is_active'])
                except ResearcherProfile.DoesNotExist:
                    pass

            return JsonResponse({'success': True, 'message': 'Activity approved.'})

        elif action == 'reject':
            title = review.activity.title
            review.activity.delete()
            StudentNotification.objects.create(
                user=student_user,
                message=f"Your activity \"{title}\" was rejected. Reason: {reason or 'No reason provided'}",
            )
            return JsonResponse({'success': True, 'message': 'Activity rejected and deleted.'})

        return JsonResponse({'success': False, 'error': 'Invalid action'}, status=400)

    except ActivityReview.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Review not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


from django.db.models import Q

@login_required
def api_get_activity(request, activity_id):
    try:
        user = request.user

        # Get research profile safely
        researcher_profile = getattr(user, "researcherprofile", None)

        activity = Activity.objects.filter(
            Q(id=activity_id) &
            (
                Q(researcher=researcher_profile) | 
                Q(tagged_users=user)
            )
        ).distinct().first()

        if not activity:
            return JsonResponse({'error': 'Activity not found'}, status=404)

        tagged = [
            {'name': u.get_full_name() or u.username, 'user_type': u.user_type}
            for u in activity.tagged_users.all()
        ]

        return JsonResponse({
            'id': activity.id,
            'title': activity.title,
            'description': activity.description or '',
            'activity_type': activity.activity_type,
            'activity_type_display': activity.get_activity_type_display(),
            'category': activity.category or '',
            'category_display': activity.get_category_display() if activity.category else '',
            'date': activity.date.strftime('%b %d, %Y') if activity.date else None,
            'location': activity.location or '',
            'invited': activity.invited,
            'keynote': activity.keynote,
            'co_presenters': activity.co_presenters or '',
            'audience': activity.audience or '',
            'tagged_users': tagged,
            'source': activity.source, 
            'objectives': [o.name for o in activity.objectives.all()],
        })

    except Exception as e:
        print("API ERROR:", e)  # shows the real error in terminal
        return JsonResponse({'error': 'Internal server error'}, status=500)




# ─────────────────────────────────────────────
# Projects
# ─────────────────────────────────────────────

@login_required
@researcher_required
def view_projects(request):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        projects   = Project.objects.filter(researcher=researcher, is_deleted=False)
    except ResearcherProfile.DoesNotExist:
        projects = Project.objects.none()

    sort_by = request.GET.get("sort", "recent")
    order   = {"oldest": "start_date", "title-asc": "title",
               "title-desc": "-title"}.get(sort_by, "-start_date")
    projects = projects.order_by(order)

    search_query = request.GET.get("search")
    if search_query:
        q = Q()
        if " AND " in search_query:
            for part in search_query.split(" AND "):
                q &= Q(title__icontains=part.strip())
        elif " OR " in search_query:
            for part in search_query.split(" OR "):
                q |= Q(title__icontains=part.strip())
        else:
            q = Q(title__icontains=search_query)
        projects = projects.filter(q)

    projects_list = []
    for project in projects:
        project.total_funding_received = project.funding_received or 0
        projects_list.append(project)

    return render(request, 'Pages/projects/view_projects.html', {'projects': projects_list})


@login_required
@researcher_required
def api_get_project(request, id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=id, researcher=researcher)

        funding_type = project.funding_type

        # ── CCV team members ──────────────────────────────────
        team_members = [{
            'id':                       m.id,
            'name':                     m.name,
            'role':                     m.role,
            'role_display':             m.get_role_display(),
            'is_academic_collaborator': m.is_academic_collaborator,
            'partner_type':             m.partner_type or 'academic',
            'partner_type_display':     m.get_partner_type_display() if m.partner_type else 'Academic',
            'manually_added':           not m.is_academic_collaborator,  # ← CCV ones have is_academic_collaborator=True
        } for m in project.team_members.all()]

        # ── Tagged RIMS members ───────────────────────────────
        tagged_members = []
        for u in project.tagged_members.all():
            try:
                sp             = u.student_profile
                degree_level   = sp.degree_level or ''
                degree_display = sp.degree_label or ''
                department     = sp.department or ''
            except Exception:
                degree_level = degree_display = department = ''

            tagged_members.append({
                'id':             u.id,
                'name':           u.get_full_name(),
                'user_type':      u.user_type,
                'is_hqp':         u.user_type == 'student',
                'degree_level':   degree_level,
                'degree_display': degree_display,
                'department':     department,
            })

        from .utils import mask_financial_fields

        project_data = {
            'id':                   project.id,
            'title':                project.title,
            'summary':              generate_project_summary(project),
            'description':          project.description or '',
            'status':               project.status,
            'is_active':            project.is_active,
            'funding_type':         funding_type,
            'role':                 project.role,
            'role_display':         project.get_role_display(),
            'start_date':           project.start_date.isoformat() if project.start_date else None,
            'end_date':             project.end_date.isoformat()   if project.end_date   else None,
            'funding_organization': project.funding_organization,
            'total_funding':        int(project.total_funding)       if project.total_funding       else None,
            'funding_received':     int(project.funding_received)    if project.funding_received    else None,
            'funding_kept_by_unb':  int(project.funding_kept_by_unb) if project.funding_kept_by_unb else None,
            'next_steps':           project.next_steps or '',
            'conception':           project.conception or '',
            'ip_activities':        project.ip_activities or '',
            'team_members':         team_members,
            'tagged_members':       tagged_members,
            'currency':             project.currency or 'CAD', 
        }

        return JsonResponse(mask_financial_fields(project_data, request.user))
    
    except Exception as e:
        return JsonResponse({'error': str(e), 'type': type(e).__name__}, status=404)


def generate_project_summary(project):
    parts    = []
    role_map = {
        'pi':     'Principal Investigator-led',
        'co_pi':  'Co-Investigator',
        'pa':     'Principal Applicant-led',
        'co_app': 'Co-applicant',
        'other':  'Team member',
    }
    role_str     = role_map.get(project.role, '') if project.role else ''
    funding_type = (project.funding_type or 'project').lower()
    opening      = f"{role_str} {funding_type}".strip() if role_str else funding_type.capitalize()
    parts.append(opening)

    if project.funding_organization:
        parts[-1] += f" funded by {project.funding_organization}"

    if project.program_name:
        parts[-1] += f" under the {project.program_name} program"

    if project.start_date and project.end_date:
        parts.append(f"running from {project.start_date} to {project.end_date}")
    elif project.start_date:
        parts.append(f"starting {project.start_date}")

    if project.total_funding:
        parts.append(f"with total funding of ${int(project.total_funding):,}")

    # Use len() to avoid bypassing prefetch cache
    team_count = len(project.team_members.all())
    if team_count > 0:
        parts.append(f"involving a team of {team_count + 1}")

    if not parts:
        return f"Research project: {project.title}."

    summary = ", ".join(parts) + "."
    return summary[0].upper() + summary[1:]


@login_required
@researcher_required
@require_http_methods(["POST"])
def api_update_project_funding(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project = Project.objects.get(id=project_id, researcher=researcher)
        data = json.loads(request.body)

        def to_decimal(val):
            try:
                return Decimal(str(val)) if val not in (None, '', 'null') else None
            except InvalidOperation:
                return None

        def to_date(val):
            try:
                return datetime.strptime(val, '%Y-%m-%d').date() if val else None
            except ValueError:
                return None

        project.funding_organization = data.get('funding_organization') or None
        project.funding_type         = data.get('funding_type') or None
        project.total_funding        = to_decimal(data.get('total_funding'))
        project.funding_received     = to_decimal(data.get('funding_received'))
        project.start_date           = to_date(data.get('start_date'))
        project.end_date             = to_date(data.get('end_date'))
        project.currency = data.get('currency') or 'CAD'

        project.manually_overridden = True

        project.save()

        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} updated funding for "{project.title}"')
        return JsonResponse({'success': True})

    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@researcher_required
def api_project_funding(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)

        records = Funding.objects.filter(
            researcher=researcher,
            project=project,
        ).order_by('start_date')

        from .utils import mask_financial_fields

        result = []
        for f in records:
            result.append({
                'organization':   f.organization  or '',
                'funding_type':   f.funding_type   or '',
                'program_name':   f.program_name   or '',
                'amount':         float(f.amount or 0),
                'grant_total':    float(f.grant_total or f.amount or 0),  # ← ADD
                'amount_to_ibme': float(f.amount_to_ibme) if f.amount_to_ibme else None,
                'start_date':     str(f.start_date) if f.start_date else None,
                'end_date':       str(f.end_date)   if f.end_date   else None,
                'currency':       f.currency or 'CAD',
            })

        return JsonResponse({
            'funding_records': mask_financial_fields(result, request.user),
            'computed_total':  sum(r['grant_total'] for r in result),   # ← use grant_total
            'computed_ibme':   sum(r['amount_to_ibme'] or 0 for r in result),
        })

    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_http_methods(["POST"])
@researcher_required
def api_update_project_conception(request, id):
    """
    Researcher edits conception on their own projects via the modal.
    Admin can edit any project (also has Django admin access).
    Students are blocked.
    """
    if request.user.user_type == 'student':
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)
    try:
        if request.user.user_type == 'admin':
            project = get_object_or_404(Project, id=id)
        else:
            researcher = ResearcherProfile.objects.get(user=request.user)
            project    = get_object_or_404(Project, id=id, researcher=researcher)

        project.conception = json.loads(request.body).get('conception', '')
        project.save()
        log_action(request, 'project_updated', target=project,
                    summary=f'{request.user.get_full_name()} updated conception for "{project.title}"')
        return JsonResponse({'success': True})
    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@researcher_required
def api_update_project_next_steps(request, id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    try:
        researcher    = ResearcherProfile.objects.get(user=request.user)
        project       = Project.objects.get(id=id, researcher=researcher)
        project.next_steps = json.loads(request.body).get('next_steps', '')
        project.save()
        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} updated next steps for "{project.title}"')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
@researcher_required
def update_project_ip(request, project_id):
    try:
        project = Project.objects.get(id=project_id, researcher__user=request.user)
        project.ip_activities = json.loads(request.body).get('ip_activities', '')
        project.save()
        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} updated IP activities for "{project.title}"')
        return JsonResponse({'success': True})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@researcher_required
@require_http_methods(["POST"])
def api_add_project_member(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
        data       = json.loads(request.body)

        name         = data.get('name', '').strip()
        role         = data.get('role', 'other')
        partner_type = data.get('partner_type', 'other')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)

        member = ProjectMember.objects.create(
            project      = project,
            name         = name,
            role         = role,
            partner_type = partner_type,
            is_academic_collaborator = False,
        )

        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} added external member "{name}" to "{project.title}"')

        return JsonResponse({
            'success': True,
            'member': {
                'id':           member.id,
                'name':         member.name,
                'role':         member.role,
                'role_display': member.get_role_display(),
                'partner_type': member.partner_type,
                'partner_type_display': member.get_partner_type_display(),
            }
        })
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@researcher_required
@require_http_methods(["POST"])
def api_remove_project_member(request, project_id, member_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
        member     = ProjectMember.objects.get(id=member_id, project=project)
        name       = member.name
        member.delete()

        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} removed "{name}" from "{project.title}"')

        return JsonResponse({'success': True})
    except (Project.DoesNotExist, ProjectMember.DoesNotExist):
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@researcher_required
def export_projects_csv(request):
    researcher = request.user.researcherprofile
    projects   = Project.objects.filter(
        researcher=researcher,
        is_deleted=False,
    ).prefetch_related('team_members').order_by('-start_date')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="grants_projects.csv"'
    writer = csv.writer(response)
    writer.writerow(['Title','Status','Your Role','Funding Organization','Funding Type',
                     'Start Date','End Date', 'Currency', 'Total Funding','Amount Awarded to IBME', 'Amount Kept by IBME',
                     'Team Members','Academic Collaborators','Community / Industry Partners',
                     'Description','Next Steps','Linked Publications (count)','Linked KM Activities (count)'])
    for project in projects:
        team      = project.team_members.all()
        pub_count = project.linked_publications.filter(is_active=True).count()
        km_count  = project.linked_activities.filter(
            is_active=True, category='knowledge_mobilization').count()
        writer.writerow([
            project.title, project.get_status_display(), project.get_role_display(),
            project.funding_organization or '', project.funding_type or '',
            project.start_date or '', project.end_date or '',
            project.currency or 'CAD',
            project.total_funding or '', 
            project.funding_received or '',
            project.funding_kept_by_unb or '',
            '; '.join(m.name for m in team),
            '; '.join(m.name for m in team.filter(partner_type='academic')),
            '; '.join(m.name for m in team.exclude(partner_type='academic').exclude(partner_type=None)),
            project.description or '', project.next_steps or '',
            pub_count, km_count,
        ])
    return response


# ─────────────────────────────────────────────
# Project ↔ Publication / Activity Linking
# ─────────────────────────────────────────────

@login_required
@researcher_required
def api_search_publications(request, project_id):
    query  = request.GET.get('q', '').strip()
    recent = request.GET.get('recent') == '1'
    project    = get_object_or_404(Project, id=project_id, researcher=request.user.researcherprofile)
    researcher = project.researcher
    if not query and recent:
        pubs = Publication.objects.filter(researcher=researcher, is_active=True).order_by('-publication_date')[:10]
    elif len(query) >= 2:
        pubs = Publication.objects.filter(researcher=researcher, is_active=True,
                                          title__icontains=query).order_by('-publication_date')[:10]
    else:
        return JsonResponse({'results': []})
    linked_ids = set(project.linked_publications.values_list('id', flat=True))
    return JsonResponse({'results': [{
        'id': p.id, 'title': p.title, 'type': p.get_publication_type_display(),
        'year': str(p.publication_date)[:4] if p.publication_date else '',
        'journal': p.journal or '', 'linked': p.id in linked_ids,
    } for p in pubs]})


@login_required
@researcher_required
def api_search_activities(request, project_id):
    query  = request.GET.get('q', '').strip()
    recent = request.GET.get('recent') == '1'
    project    = get_object_or_404(Project, id=project_id, researcher=request.user.researcherprofile)
    researcher = project.researcher
    if not query and recent:
        acts = Activity.objects.filter(researcher=researcher, is_active=True).order_by('-date')[:10]
    elif len(query) >= 2:
        acts = Activity.objects.filter(researcher=researcher, is_active=True,
                                       title__icontains=query).order_by('-date')[:10]
    else:
        return JsonResponse({'results': []})
    linked_ids = set(project.linked_activities.values_list('id', flat=True))
    return JsonResponse({'results': [{
        'id': a.id, 'title': a.title, 'type': a.get_activity_type_display(),
        'category': a.get_category_display() if hasattr(a, 'get_category_display') else a.category,
        'date': a.date.strftime('%b %d, %Y') if a.date else '',
        'linked': a.id in linked_ids,
    } for a in acts]})


@login_required
@researcher_required
@require_http_methods(["POST"])
def api_link_publication(request, project_id, pub_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        Project.objects.get(id=project_id, researcher=researcher).linked_publications.add(
            Publication.objects.get(id=pub_id, researcher=researcher))
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)


@login_required
@require_http_methods(["POST"])
@researcher_required
def api_unlink_publication(request, project_id, pub_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        Project.objects.get(id=project_id, researcher=researcher).linked_publications.remove(
            Publication.objects.get(id=pub_id, researcher=researcher))
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)


@login_required
@require_http_methods(["POST"])
@researcher_required
def api_link_activity(request, project_id, act_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        Project.objects.get(id=project_id, researcher=researcher).linked_activities.add(
            Activity.objects.get(id=act_id, researcher=researcher))
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)


@login_required
@require_http_methods(["POST"])
@researcher_required
def api_unlink_activity(request, project_id, act_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        Project.objects.get(id=project_id, researcher=researcher).linked_activities.remove(
            Activity.objects.get(id=act_id, researcher=researcher))
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)


@login_required
@researcher_required
def api_get_linked_items(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
    except (ResearcherProfile.DoesNotExist, Project.DoesNotExist):
        return JsonResponse({'error': 'Not found'}, status=404)

    return JsonResponse({
        'publications': [{
            'id': p.id, 'title': p.title, 'type': p.get_publication_type_display(),
            'year': str(p.publication_date)[:4] if p.publication_date else '',
            'journal': p.journal or '',
        } for p in project.linked_publications.filter(is_active=True)],
        'activities': [{
            'id': a.id, 'title': a.title, 'type': a.get_activity_type_display(),
            'category': a.get_category_display() if hasattr(a, 'get_category_display') else a.category,
            'date': a.date.strftime('%b %d, %Y') if a.date else '',
        } for a in project.linked_activities.filter(is_active=True)],
    })


# ─────────────────────────────────────────────
# Peer Search
# ─────────────────────────────────────────────

@login_required
def api_peer_search(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'users': []})
    users = CustomUser.objects.filter(
        Q(first_name__icontains=query) | Q(last_name__icontains=query),
        user_type__in=['student', 'researcher'], is_active=True
    ).exclude(id=request.user.id)[:10]
    return JsonResponse({'users': [{
        'id': u.id, 'name': f"{u.first_name} {u.last_name}".strip() or u.username,
        'user_type': u.user_type, 'email': u.email,
    } for u in users]})


# ─────────────────────────────────────────────
# Notifications
# ─────────────────────────────────────────────

@login_required
def api_pending_activities(request):
    if request.user.user_type != 'admin':
        return JsonResponse({'success': False, 'error': 'Admin only'}, status=403)
    reviews = ActivityReview.objects.filter(
        status='pending').select_related('activity__researcher__user').order_by('-created_at')
    return JsonResponse({'activities': [{
        'review_id':     r.id,
        'activity_id':   r.activity.id,
        'title':         r.activity.title,
        'type':          r.activity.get_activity_type_display(),
        'date':          r.activity.date.strftime('%b %d, %Y') if r.activity.date else '',
        'student':       r.activity.researcher.user.get_full_name(),
        'student_email': r.activity.researcher.user.email,
        'description':   r.activity.description or '',
        'submitted':     r.created_at.strftime('%b %d, %Y'),
        'location':      r.activity.location or '',
        'invited':       r.activity.invited,
        'keynote':       r.activity.keynote,
        'co_presenters': r.activity.co_presenters or '',
        'audience':      r.activity.audience or '',
        'tagged_users': [
            f"{u.first_name} {u.last_name}".strip()
            for u in r.activity.tagged_users.all()
        ]
    } for r in reviews]})


@login_required
@require_http_methods(["POST"])
def api_mark_notifications_read(request):
    try:
        StudentNotification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST"])
def api_clear_notifications(request):
    try:
        deleted_count, _ = StudentNotification.objects.filter(user=request.user).delete()
        return JsonResponse({'success': True, 'deleted': deleted_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_http_methods(["POST"])
def api_dismiss_notification(request, notif_id):
    try:
        deleted_count, _ = StudentNotification.objects.filter(id=notif_id, user=request.user).delete()
        if deleted_count == 0:
            return JsonResponse({'success': False, 'error': 'Notification not found'}, status=404)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ── views.py additions ────────────────────────────────────────

@login_required
def profile_view(request):
    student_profile    = None
    researcher_profile = None
    education          = []
    all_award_records  = []
    researchers        = []
    current_students   = []
    all_students       = []
    linked_student_ids = set()

    if request.user.user_type == 'student':
        student_profile = StudentProfile.objects.filter(user=request.user).first()
        researchers = ResearcherProfile.objects.select_related("user").filter(
            user__user_type="researcher"
        ).order_by("user__last_name")

    elif request.user.user_type == 'researcher':
        researcher_profile = ResearcherProfile.objects.filter(user=request.user).first()
        if researcher_profile:
            education = Education.objects.filter(
                researcher=researcher_profile
            ).order_by('-expected_date')

            all_award_records = Recognition.objects.filter(
                researcher=researcher_profile
            ).order_by('-start_date')

            current_students = SupervisionRecord.objects.filter(
                researcher=researcher_profile,
                status='in_progress',
            ).select_related(
                'linked_student__user'
            ).order_by('student_name')

            # IDs of students already linked to a record for this researcher
            linked_student_ids = set(
                SupervisionRecord.objects.filter(
                    researcher=researcher_profile,
                    linked_student__isnull=False,
                ).values_list('linked_student_id', flat=True)
            )

            all_supervised_names = SupervisionRecord.objects.filter(
                researcher=researcher_profile,
            ).values_list('student_name', flat=True)

            name_filters = Q()
            for name in all_supervised_names:
                parts = name.strip().split()
                if len(parts) >= 2:
                    name_filters |= Q(
                        user__first_name__iexact=parts[0],
                        user__last_name__iexact=parts[-1],
                    ) | Q(
                        user__first_name__iexact=parts[-1],
                        user__last_name__iexact=parts[0],
                    )

            all_students = StudentProfile.objects.select_related('user').filter(
                name_filters,
                user__approval_status='approved',
                user__user_type='student',
            ).order_by(
                'user__last_name', 'user__first_name'
            ) if name_filters else StudentProfile.objects.none()

    return render(request, 'Pages/editable/profile.html', {
        'student_profile':    student_profile,
        'researcher_profile': researcher_profile,
        'education':          education,
        'all_award_records':  all_award_records,
        'researchers':        researchers,
        'current_students':   current_students,
        'all_students':       all_students,
        'linked_student_ids': linked_student_ids,
    })


@login_required
@require_http_methods(["POST"])
def api_update_student_academic_info(request):
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        data    = json.loads(request.body)
        profile = StudentProfile.objects.get(user=request.user)

        from django.utils.dateparse import parse_date

        def safe_date(val):
            if not val:
                return None
            parsed = parse_date(val)
            if not parsed:
                raise ValueError(f'Invalid date: {val}')
            return parsed

        valid_degrees = [c[0] for c in StudentProfile.DEGREE_CHOICES]

        degree_level = data.get('degree_level')
        if degree_level and degree_level not in valid_degrees:
            return JsonResponse({'success': False, 'error': 'Invalid degree level.'}, status=400)

        if 'degree_level' in data:
            profile.degree_level = degree_level or None
        if 'department' in data:
            profile.department = (data.get('department') or '').strip() or None
        if 'start_date' in data:
            profile.start_date = safe_date(data.get('start_date'))
        if 'expected_end_date' in data:
            profile.expected_end_date = safe_date(data.get('expected_end_date'))
        if 'thesis_title' in data:
            profile.thesis_title = (data.get('thesis_title') or '').strip() or None
        if 'graduation_date' in data:
            profile.graduation_date = safe_date(data.get('graduation_date'))

        profile.manually_overridden = True
        profile.save(update_fields=[
            'degree_level',
            'department',
            'start_date',
            'expected_end_date',
            'graduation_date',
            'thesis_title',
            'manually_overridden',
        ])

        log_action(request, 'profile_updated', target=profile,
                   summary=f'{request.user.get_full_name()} updated academic info')

        return JsonResponse({'success': True})

    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except StudentProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_update_basic_info(request):
    try:
        data = json.loads(request.body)
        user = request.user

        def sanitize_name(value, fallback):
            value = (value or fallback or '').strip()
            # Strip anything that isn't letters, spaces, hyphens, apostrophes
            value = re.sub(r"[^a-zA-Z\s\-\']", '', value)
            return value[:50]  # hard length cap

        user.first_name = sanitize_name(data.get('first_name'), user.first_name)
        user.last_name  = sanitize_name(data.get('last_name'),  user.last_name)

        # Email is not user-editable — only admins can change it via admin panel
        if user.user_type == 'researcher':
            org = data.get('organization', user.organization or '').strip()
            user.organization = org[:50]

        user.save()
        log_action(request, 'profile_updated', target=user,
                   summary=f'{user.get_full_name()} updated their profile')
        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_http_methods(["POST"])
def api_update_research_interests(request):
    if request.user.user_type != 'researcher':
        return JsonResponse({'success': False, 'error': 'Not a researcher'}, status=403)
    try:
        data    = json.loads(request.body)
        profile = ResearcherProfile.objects.get(user=request.user)
        profile.research_interests = data.get('research_interests', '')
        profile.save()
        log_action(request, 'profile_updated', target=profile,
                   summary=f'{request.user.get_full_name()} updated research interests')
        return JsonResponse({'success': True})
    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_update_edi(request):
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        data    = json.loads(request.body)
        profile = StudentProfile.objects.get(user=request.user)
        profile.gender              = data.get('gender') or None
        profile.residency_status    = data.get('residency_status') or None
        profile.indigenous_identity = data.get('indigenous_identity') or None
        profile.race_ethnicity      = data.get('race_ethnicity') or None
        profile.edi_profile_completed = True
        profile.save()

        # ── Audit log — note what was provided (not the values) ──
        provided = [
            f for f, v in {
                'gender':              profile.gender,
                'residency_status':    profile.residency_status,
                'indigenous_identity': profile.indigenous_identity,
                'race_ethnicity':      profile.race_ethnicity,
            }.items() if v and v != 'prefer_not'
        ]
        summary = (
            f'{request.user.get_full_name()} updated EDI profile. '
            f'Fields provided: {", ".join(provided) if provided else "none (all prefer not to say)"}'
        )
        log_action(request, 'edi_updated', target=profile, summary=summary)
        return JsonResponse({'success': True})
    except StudentProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_update_consent(request):
    try:
        data = json.loads(request.body)
        request.user.consent_to_share = bool(data.get('consent_to_share', False))
        request.user.save()
        log_action(request, 'consent_updated', target=request.user,
                    summary=f'{request.user.get_full_name()} consent={request.user.consent_to_share}')
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@admin_required
def export_conference_equity_csv(request):
    students = StudentProfile.objects.select_related('user').filter(
        user__consent_to_share=True,
        degree_level__in=['msc', 'phd'],
    )
    count = students.count()

    if count < 5:
        messages.error(
            request,
            f"Only {count} consenting graduate student{'s' if count != 1 else ''} found. "
            "At least 5 are required to generate this EDI report and protect individual privacy."
        )
        return redirect('reports_list')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="conference_equity_report.csv"'
    writer = csv.writer(response)

    writer.writerow(['IBME Conference Equity Report'])
    writer.writerow([f'Generated: {date.today()}'])
    writer.writerow([f'Total consenting students: {count}'])
    writer.writerow([])

    # ── Section 1: Gender ────────────────────────────────────
    writer.writerow(['Section', 'Category', 'Count'])
    gender_map = {
        'female': 'Female', 'male': 'Male', 'non_binary': 'Non-binary',
        'self_describe': 'Prefer to self-describe', 'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    gender_counts = students.values('gender').annotate(n=Count('id'))
    for row in gender_counts:
        label = gender_map.get(row['gender'], row['gender'] or 'Not provided')
        writer.writerow(['Gender', label, row['n']])
    writer.writerow([])

    # ── Section 2: Indigenous identity ───────────────────────
    indigenous_map = {
        'yes': 'Yes — First Nations, Métis, or Inuit',
        'no': 'No',
        'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    indigenous_counts = students.values('indigenous_identity').annotate(n=Count('id'))
    for row in indigenous_counts:
        label = indigenous_map.get(row['indigenous_identity'], row['indigenous_identity'] or 'Not provided')
        writer.writerow(['Indigenous Identity', label, row['n']])
    writer.writerow([])

    # ── Section 3: Residency ─────────────────────────────────
    residency_map = {
        'citizen': 'Canadian Citizen', 'permanent_resident': 'Permanent Resident',
        'international': 'International Student', 'prefer_not': 'Prefer not to say',
        None: 'Not provided',
    }
    residency_counts = students.values('residency_status').annotate(n=Count('id'))
    for row in residency_counts:
        label = residency_map.get(row['residency_status'], row['residency_status'] or 'Not provided')
        writer.writerow(['Residency Status', label, row['n']])
    writer.writerow([])

    # ── Section 4: Degree level ──────────────────────────────
    degree_map = {'msc': 'MSc', 'phd': 'PhD'}
    degree_counts = students.values('degree_level').annotate(n=Count('id'))
    for row in degree_counts:
        label = degree_map.get(row['degree_level'], row['degree_level'])
        writer.writerow(['Degree Level', label, row['n']])
    writer.writerow([])

    # ── Annotate once for sections 5 & 6 ────────────────────
    current_year = date.today().year
    EQUITY_WINDOW_YEARS = 3
    window_start = current_year - EQUITY_WINDOW_YEARS

    students = students.annotate(
        conf_count=Count(
            'user__researcherprofile__activity',
            filter=Q(
                user__researcherprofile__activity__category='conference',
                user__researcherprofile__activity__is_active=True,
                user__researcherprofile__activity__date__year__gte=window_start,
            ),
            distinct=True,
        ),
        km_count=Count(
            'user__researcherprofile__activity',
            filter=Q(
                user__researcherprofile__activity__category='knowledge_mobilization',
                user__researcherprofile__activity__is_active=True,
                user__researcherprofile__activity__date__year__gte=window_start,
            ),
            distinct=True,
        ),
    )

    # ── Section 5: Conference attendance ────────────────────
    brackets = {'0 conferences': 0, '1–2 conferences': 0, '3–5 conferences': 0, '6+ conferences': 0}
    for s in students:
        c = s.conf_count  # ← fixed: no underscore, from annotation
        if c == 0:        brackets['0 conferences'] += 1
        elif c <= 2:      brackets['1–2 conferences'] += 1
        elif c <= 5:      brackets['3–5 conferences'] += 1
        else:             brackets['6+ conferences'] += 1

    for label, n in brackets.items():
        writer.writerow(['Conference Attendance (last 3 yrs)', label, n])
    writer.writerow([])

    # ── Section 6: KM activities ─────────────────────────────
    km_brackets = {'0 activities': 0, '1–2 activities': 0, '3–5 activities': 0, '6+ activities': 0}
    for s in students:
        k = s.km_count    # ← fixed: from annotation, no per-student query
        if k == 0:        km_brackets['0 activities'] += 1
        elif k <= 2:      km_brackets['1–2 activities'] += 1
        elif k <= 5:      km_brackets['3–5 activities'] += 1
        else:             km_brackets['6+ activities'] += 1

    for label, n in km_brackets.items():
        writer.writerow(['KM Activity Participation (last 3 yrs)', label, n])

    return response

@login_required
@admin_required
def conference_equity_report(request):
    from .models import Conference
    conferences = Conference.objects.annotate(
        activity_count=Count('activities')
    ).filter(activity_count__gt=0).order_by('-year', 'name')

    selected_conf = None
    result        = None
    suppressed    = False
    MIN_THRESHOLD = 5

    conf_id = request.GET.get('conference_id')
    if conf_id:
        try:
            selected_conf = Conference.objects.get(id=conf_id)

            EQUITY_WINDOW_YEARS = 3
            window_start = date.today().replace(year=date.today().year - EQUITY_WINDOW_YEARS)

            activities = Activity.objects.filter(
                conference=selected_conf,
                is_active=True,
                date__gte=window_start,
            ).prefetch_related('tagged_users')

            # collect all participants
            participant_ids = set()
            for act in activities:
                participant_ids.add(act.researcher.user.id)
                for u in act.tagged_users.all():
                    participant_ids.add(u.id)

            profiles = StudentProfile.objects.filter(
                user__id__in=participant_ids,
                user__consent_to_share=True,
                degree_level__in=['msc', 'phd'],
            )

            count = profiles.count()

            if count < MIN_THRESHOLD:
                suppressed = True
            else:
                gender_map = {
                    'female': 'Female', 'male': 'Male',
                    'non_binary': 'Non-binary',
                    'self_describe': 'Prefer to self-describe',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                indigenous_map = {
                    'yes': 'Yes', 'no': 'No',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                residency_map = {
                    'citizen': 'Canadian Citizen',
                    'permanent_resident': 'Permanent Resident',
                    'international': 'International Student',
                    'prefer_not': 'Prefer not to say',
                    None: 'Not provided',
                }
                degree_map = {'msc': 'MSc', 'phd': 'PhD'}

                def get_counts(field, label_map):
                    from collections import Counter
                    counts = Counter(getattr(p, field) for p in profiles)
                    return [
                        {'label': label_map.get(k, k or 'Not provided'), 'count': v}
                        for k, v in sorted(counts.items(), key=lambda x: -x[1])
                    ]

                race_map = {
                    'white':           'White / European descent',
                    'black':           'Black / African descent',
                    'east_asian':      'East Asian',
                    'south_asian':     'South Asian',
                    'southeast_asian': 'Southeast Asian',
                    'latin':           'Latin American / Hispanic',
                    'middle_eastern':  'Middle Eastern / North African',
                    'indigenous':      'Indigenous (other)',
                    'mixed':           'Two or more ethnicities',
                    'other':           'Other',
                    'prefer_not':      'Prefer not to say',
                    None:              'Not provided',
                }

                result = {
                    'total':          count,
                    'gender':         get_counts('gender', gender_map),
                    'indigenous':     get_counts('indigenous_identity', indigenous_map),
                    'residency':      get_counts('residency_status', residency_map),
                    'degree':         get_counts('degree_level', degree_map),
                    'race_ethnicity': get_counts('race_ethnicity', race_map),
                }

        except Conference.DoesNotExist:
            pass

    return render(request, 'Pages/reports/conference_equity_report.html', {
        'conferences':     conferences,
        'selected_conf':   selected_conf,
        'result':          result,
        'suppressed':      suppressed,
        'conf_id':         conf_id,
        'MIN_THRESHOLD':   MIN_THRESHOLD,
    })



@login_required
@admin_required
def conference_equity_summary(request):
    from collections import defaultdict
    from datetime import date, timedelta

    WINDOW_YEARS    = 3
    three_years_ago = date.today() - timedelta(days=WINDOW_YEARS * 365)

    filter_pi   = request.GET.get('pi', '')
    filter_dept = request.GET.get('dept', '')

    # ── Step 1: All conference activities in the 3yr window ──
    conf_activities = Activity.objects.filter(
        category='conference',
        is_active=True,
        date__gte=three_years_ago,
    ).prefetch_related('tagged_users').select_related('researcher__user')

    # ── Step 2: Build user_id → conference count ──────────────
    user_conf_count = defaultdict(int)

    for act in conf_activities:
        if act.researcher.user.user_type == 'student':
            user_conf_count[act.researcher.user_id] += 1
        for u in act.tagged_users.all():
            if u.user_type == 'student':
                user_conf_count[u.id] += 1

    # ── Step 3: Get supervision records with linked students ──
    sup_qs = SupervisionRecord.objects.filter(
        linked_student__isnull=False,
        status='in_progress',
    ).select_related(
        'researcher__user',
        'linked_student__user',
        'linked_student',
    )

    if filter_pi:
        sup_qs = sup_qs.filter(researcher__user__id=filter_pi)
    if filter_dept:
        sup_qs = sup_qs.filter(department__icontains=filter_dept)

    # ── Step 4: Group by PI + department ─────────────────────
    pi_dept_map = defaultdict(lambda: {
        'pi':           '',
        'pi_id':        None,
        'department':   '',
        'msc':          0,
        'phd':          0,
        'other':        0,
        'total':        0,
        'conf_total':   0,
        'students':     [],
        'students_set': set(),
    })

    degree_to_level = {
        'masters_thesis':     'msc',
        'masters_non_thesis': 'msc',
        'doctorate':          'phd',
    }

    for sup in sup_qs:
        dept  = (sup.department or '').strip() or 'Unknown'
        key   = f"{sup.researcher.user_id}|{dept.lower()}"
        entry = pi_dept_map[key]

        entry['pi']         = sup.researcher.user.get_full_name()
        entry['pi_id']      = sup.researcher.user_id
        entry['department'] = dept

        student    = sup.linked_student
        degree_lvl = degree_to_level.get(sup.degree_type, 'other')
        conf_count = user_conf_count.get(student.user_id, 0)

        if student.user_id in entry['students_set']:
            continue
        entry['students_set'].add(student.user_id)

        entry[degree_lvl]   += 1
        entry['total']      += 1
        entry['conf_total'] += conf_count

        entry['students'].append({
            'name':       student.user.get_full_name(),
            'degree':     sup.get_degree_type_display() or '—',
            'conf_count': conf_count,
        })

    # ── Step 5: Compute averages ──────────────────────────────
    rows = []

    for entry in pi_dept_map.values():
        total = entry['total']
        if not total:
            continue

        avg_conf          = round(entry['conf_total'] / total, 1)
        zero_conf         = sum(1 for s in entry['students'] if s['conf_count'] == 0)
        zero_conf_pct     = round(zero_conf / total * 100)
        participation_pct = 100 - zero_conf_pct

        rows.append({
            'pi':                entry['pi'],
            'pi_id':             entry['pi_id'],
            'department':        entry['department'],
            'msc':               entry['msc'],
            'phd':               entry['phd'],
            'other':             entry['other'],
            'total':             total,
            'avg_conf':          avg_conf,
            'conf_total':        entry['conf_total'],
            'zero_conf':         zero_conf,
            'zero_conf_pct':     zero_conf_pct,
            'participation_pct': participation_pct,
        })

    rows.sort(key=lambda x: x['avg_conf'], reverse=True)

    # ── Step 6: Overall summary + highlight below average ─────
    total_students = sum(r['total'] for r in rows)
    total_conf     = sum(r['conf_total'] for r in rows)
    overall_avg    = round(total_conf / total_students, 1) if total_students else 0

    for r in rows:
        r['below_avg'] = r['avg_conf'] < overall_avg

    # ── Dropdown options ──────────────────────────────────────
    all_departments = list(
        SupervisionRecord.objects.filter(
            department__isnull=False,
            linked_student__isnull=False,
        ).exclude(department='').values_list(
            'department', flat=True
        ).distinct().order_by('department')
    )

    all_supervisors = CustomUser.objects.filter(
        user_type='researcher'
    ).order_by('last_name')

    # ── CSV export ────────────────────────────────────────────
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="conference_equity_summary.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'PI', 'Department', 'MSc Students', 'PhD Students', 'Other',
            'Total Students', 'Total Conferences Attended',
            'Avg Conferences/Student', 'Students with 0 Conferences', '% with 0 Conferences',
        ])
        for r in rows:
            writer.writerow([
                r['pi'], r['department'],
                r['msc'], r['phd'], r['other'],
                r['total'], r['conf_total'], r['avg_conf'],
                r['zero_conf'], r['zero_conf_pct'],
            ])
        return response

    return render(request, 'Pages/reports/conference_equity_summary.html', {
        'rows':            rows,
        'total_students':  total_students,
        'total_conf':      total_conf,
        'overall_avg':     overall_avg,
        'WINDOW_YEARS':    WINDOW_YEARS,
        'filter_pi':       filter_pi,
        'filter_dept':     filter_dept,
        'all_supervisors': all_supervisors,
        'all_departments': all_departments,
    })


@login_required
@require_http_methods(["POST"])
def api_request_supervisor(request):
    """Student requests a researcher as their supervisor."""
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        data = json.loads(request.body)
        supervisor_id = data.get('supervisor_id')
        if not supervisor_id:
            return JsonResponse({'success': False, 'error': 'supervisor_id required'}, status=400)

        supervisor = ResearcherProfile.objects.get(id=supervisor_id)
        student_profile = StudentProfile.objects.get(user=request.user)

        # Check if already has an approved supervisor
        if student_profile.supervisor:
            return JsonResponse({
                'success': False,
                'error': f'You already have an assigned supervisor: {student_profile.supervisor.user.get_full_name()}'
            }, status=400)

        # Check for existing pending request to this supervisor
        existing = SupervisorRequest.objects.filter(
            student=student_profile, supervisor=supervisor, status='pending',
        ).first()
        if existing:
            return JsonResponse({
                'success': False,
                'error': 'You already have a pending request to this supervisor.'
            }, status=400)

        sup_request = SupervisorRequest.objects.create(
            student=student_profile, supervisor=supervisor, status='pending',
        )

        # Notify admins
        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=f'{request.user.get_full_name()} requested {supervisor.user.get_full_name()} as their supervisor.',
                request_id=sup_request.id,    # fixed: was missing
            )

        # Notify the supervisor directly
        StudentNotification.objects.create(
            user=supervisor.user,
            message=f'{request.user.get_full_name()} has requested you as their supervisor.',
            request_id=sup_request.id,
        )

        # Audit log
        log_action(
            request, 'supervisor_requested', target=sup_request,
            summary=f'{request.user.get_full_name()} requested {supervisor.user.get_full_name()} as supervisor',
            details={'supervisor_id': supervisor.id, 'student_id': student_profile.id},
        )

        return JsonResponse({
            'success': True,
            'message': f'Supervisor request sent to {supervisor.user.get_full_name()}.',
            'request_id': sup_request.id,
        })
    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Researcher not found'}, status=404)
    except StudentProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_request_co_supervisor(request):
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        data = json.loads(request.body)
        supervisor_id = data.get('supervisor_id')
        if not supervisor_id:
            return JsonResponse({'success': False, 'error': 'supervisor_id required'}, status=400)

        supervisor      = ResearcherProfile.objects.get(id=supervisor_id)
        student_profile = StudentProfile.objects.get(user=request.user)

        # Can't add primary supervisor as co-supervisor
        if student_profile.supervisor and student_profile.supervisor.id == supervisor.id:
            return JsonResponse({
                'success': False,
                'error': f'{supervisor.user.get_full_name()} is already your primary supervisor.'
            }, status=400)

        # Already a co-supervisor
        if student_profile.co_supervisors.filter(id=supervisor.id).exists():
            return JsonResponse({
                'success': False,
                'error': f'{supervisor.user.get_full_name()} is already your co-supervisor.'
            }, status=400)

        # Instantly add — no approval needed
        student_profile.co_supervisors.add(supervisor)

        # Notify the researcher
        StudentNotification.objects.create(
            user=supervisor.user,
            message=f'{request.user.get_full_name()} has added you as their co-supervisor.',
        )

        # Notify admins
        for admin_user in CustomUser.objects.filter(user_type='admin'):
            StudentNotification.objects.create(
                user=admin_user,
                message=f'{request.user.get_full_name()} added {supervisor.user.get_full_name()} as co-supervisor.',
            )

        log_action(
            request, 'supervisor_requested', target=student_profile,
            summary=f'{request.user.get_full_name()} added {supervisor.user.get_full_name()} as co-supervisor',
            details={'supervisor_id': supervisor.id, 'student_id': student_profile.id},
        )

        return JsonResponse({
            'success': True,
            'message': f'{supervisor.user.get_full_name()} added as co-supervisor.',
        })

    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Researcher not found'}, status=404)
    except StudentProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_remove_co_supervisor(request):
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        data = json.loads(request.body)
        supervisor_id = data.get('supervisor_id')
        if not supervisor_id:
            return JsonResponse({'success': False, 'error': 'supervisor_id required'}, status=400)

        supervisor      = ResearcherProfile.objects.get(id=supervisor_id)
        student_profile = StudentProfile.objects.get(user=request.user)

        if not student_profile.co_supervisors.filter(id=supervisor.id).exists():
            return JsonResponse({'success': False, 'error': 'Not a co-supervisor.'}, status=400)

        student_profile.co_supervisors.remove(supervisor)

        StudentNotification.objects.create(
            user=supervisor.user,
            message=f'{request.user.get_full_name()} has removed you as their co-supervisor.',
        )

        log_action(
            request, 'profile_updated', target=student_profile,
            summary=f'{request.user.get_full_name()} removed {supervisor.user.get_full_name()} as co-supervisor',
        )

        return JsonResponse({'success': True, 'message': f'{supervisor.user.get_full_name()} removed as co-supervisor.'})

    except ResearcherProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Researcher not found'}, status=404)
    except StudentProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_notifications(request):
    notifications = StudentNotification.objects.filter(
        user=request.user
    ).order_by('-created_at')

    unread_count = notifications.filter(is_read=False).count()

    data = []
    for n in notifications:
        item = {
            'id':          n.id,
            'message':     n.message,
            'is_read':     n.is_read,
            'created_at':  n.created_at.strftime('%b %d, %Y'),
            'request_id':  None,
            'activity_id': None,
        }

        # fixed: removed fragile message string check, just use request_id directly
        if n.request_id:
            try:
                req = SupervisorRequest.objects.get(id=n.request_id, status='pending')
                item['request_id'] = req.id
            except SupervisorRequest.DoesNotExist:
                pass

        data.append(item)

    notifications.filter(is_read=False).update(is_read=True)

    return JsonResponse({'notifications': data, 'unread_count': unread_count})

@login_required
@require_POST
def api_review_supervisor_request(request, request_id):

    # ── Authorization: only the target supervisor or an admin may review ──
    if request.user.user_type == 'admin':
        req = get_object_or_404(SupervisorRequest, id=request_id)
    elif request.user.user_type == 'researcher':
        try:
            profile = ResearcherProfile.objects.get(user=request.user)
        except ResearcherProfile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)
        req = get_object_or_404(SupervisorRequest, id=request_id, supervisor=profile)
    else:
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)
    # ── End auth check ────────────────────────────────────────

    if req.status != "pending":
        return JsonResponse({
            "success": False,
            "error": "This request has already been reviewed."
        }, status=400)

    data   = json.loads(request.body)
    action = data.get("action")

    if action == "approve":
        req.status = "approved"
        student = req.student
        student.supervisor = req.supervisor
        student.save()
        StudentNotification.objects.create(
            user=student.user,
            message=f"{request.user.get_full_name()} approved your supervisor request."
        )
        log_action(request, 'supervisor_approved', target=req,
                summary=f'{request.user.get_full_name()} approved {req.student.user.get_full_name()} → {req.supervisor.user.get_full_name()}')

    elif action == "reject":
        req.status = "rejected"
        student = req.student
        StudentNotification.objects.create(
            user=student.user,
            message=f"{request.user.get_full_name()} rejected your supervisor request."
        )
        log_action(request, 'supervisor_rejected', target=req,
                summary=f'{request.user.get_full_name()} rejected {req.student.user.get_full_name()} → {req.supervisor.user.get_full_name()}')
    else:
        return JsonResponse({"success": False, "error": "Invalid action"}, status=400)

    req.save()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["GET"])
def api_pending_supervisor_requests(request):
    """Researcher sees their pending requests. Admin sees all."""
    if request.user.user_type == 'admin':
        qs = SupervisorRequest.objects.filter(status='pending')
    elif request.user.user_type == 'researcher':
        try:
            profile = ResearcherProfile.objects.get(user=request.user)
            qs = SupervisorRequest.objects.filter(supervisor=profile, status='pending')
        except ResearcherProfile.DoesNotExist:
            return JsonResponse({'requests': []})
    else:
        return JsonResponse({'success': False, 'error': 'Not authorized'}, status=403)

    qs = qs.select_related('student__user', 'supervisor__user').order_by('-created_at')
    return JsonResponse({'requests': [{
        'id': r.id,
        'student_name': r.student.user.get_full_name(),
        'student_email': r.student.user.email,
        'student_degree': r.student.get_degree_level_display() if r.student.degree_level else '',
        'student_department': r.student.department or '',
        'supervisor_name': r.supervisor.user.get_full_name(),
        'supervisor_id': r.supervisor.id,
        'submitted': r.created_at.strftime('%b %d, %Y'),
    } for r in qs]})

@login_required
@require_http_methods(["GET"])
def api_my_supervisor_requests(request):
    """Student sees their own supervisor requests and current status."""
    if request.user.user_type != 'student':
        return JsonResponse({'success': False, 'error': 'Students only'}, status=403)
    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        return JsonResponse({'requests': [], 'current_supervisor': None})

    qs = SupervisorRequest.objects.filter(
        student=student_profile
    ).select_related('supervisor__user').order_by('-created_at')

    current_sup = None
    if student_profile.supervisor:
        current_sup = {
            'name': student_profile.supervisor.user.get_full_name(),
            'email': student_profile.supervisor.user.email,
        }
    return JsonResponse({
        'current_supervisor': current_sup,
        'requests': [{
            'id': r.id,
            'supervisor_name': r.supervisor.user.get_full_name(),
            'status': r.status,
            'submitted': r.created_at.strftime('%b %d, %Y'),
        } for r in qs],
    })


@login_required
@require_http_methods(["GET"])
def api_search_supervisors(request):
    """Search researchers available as supervisors."""
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})
    researchers = ResearcherProfile.objects.filter(
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(user__email__icontains=query) |
        Q(research_interests__icontains=query),
        user__user_type='researcher',
        user__approval_status='approved',
    ).select_related('user')[:10]
    return JsonResponse({'results': [{
        'id': r.id,
        'name': r.user.get_full_name(),
        'email': r.user.email,
        'title': r.title or '',
        'interests': (r.research_interests or '')[:120],
        'student_count': r.students.count(),
    } for r in researchers]})





@login_required
def notifications_view(request):
    context = {}
    if request.user.user_type == 'student':
        context['student'] = getattr(request.user, 'student_profile', None)
    return render(request, 'Pages/notifications.html', context)


@login_required
@require_http_methods(["GET"])
def api_unread_count(request):
    count = StudentNotification.objects.filter(
        user=request.user,
        is_read=False
    ).count()
    return JsonResponse({'unread': count})

@login_required
@require_http_methods(["POST"])
def api_update_project_kept_by_unb(request, project_id):
    try:
        project = Project.objects.get(id=project_id, researcher__user=request.user)
        data    = json.loads(request.body)
        val     = data.get('funding_kept_by_unb')
        
        project.funding_kept_by_unb = Decimal(str(val)) if val else None
        project.manually_overridden = True  # ← add this
        project.save()
        log_action(request, 'project_updated', target=project,
                   summary=f'{request.user.get_full_name()} updated kept by UNB for "{project.title}"')
        return JsonResponse({'success': True})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@researcher_required
@require_http_methods(["POST"])
def api_tag_project_member(request, project_id, user_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
        user       = CustomUser.objects.get(id=user_id)

        project.tagged_members.add(user)

        StudentNotification.objects.create(
            user=user,
            message=f'{request.user.get_full_name()} has added you as HQP on the project "{project.title}".'
        )

        log_action(request, 'hqp_tagged', target=project,
                   summary=f'{request.user.get_full_name()} tagged {user.get_full_name()} as HQP on "{project.title}"',
                   details={'user_id': user.id, 'user_name': user.get_full_name(), 'project_id': project.id})

        return JsonResponse({'success': True})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@researcher_required
@require_http_methods(["POST"])
def api_untag_project_member(request, project_id, user_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
        user       = CustomUser.objects.get(id=user_id)

        project.tagged_members.remove(user)

        StudentNotification.objects.create(
            user=user,
            message=f'{request.user.get_full_name()} has removed you from the project "{project.title}".'
        )

        log_action(request, 'hqp_untagged', target=project,
                   summary=f'{request.user.get_full_name()} removed {user.get_full_name()} from "{project.title}"',
                   details={'user_id': user.id, 'user_name': user.get_full_name(), 'project_id': project.id})

        return JsonResponse({'success': True})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
def api_get_team_members(request, project_id):
    try:
        project = get_object_or_404(
            Project,
            id=project_id,
            researcher__user=request.user
        )
        members = [{
            'id':                       m.id,
            'name':                     m.name,
            'role':                     m.role,
            'role_display':             m.get_role_display(),
            'partner_type':             m.partner_type or 'academic',
            'partner_type_display':     m.get_partner_type_display() if m.partner_type else 'Academic',
            'is_academic_collaborator': m.is_academic_collaborator,
            'manually_added':           not m.is_academic_collaborator,  # CCV = is_academic_collaborator=True
        } for m in project.team_members.all()]
        return JsonResponse({'members': members})
    except Project.DoesNotExist:
        return JsonResponse({'members': []})


@login_required
@researcher_required
def api_get_tagged_members(request, project_id):
    try:
        researcher = ResearcherProfile.objects.get(user=request.user)
        project    = Project.objects.get(id=project_id, researcher=researcher)
        members    = []

        for u in project.tagged_members.all():
            try:
                sp             = u.student_profile
                degree_level   = sp.degree_level or ''
                degree_display = sp.degree_label or ''
                department     = sp.department or ''
            except Exception:
                degree_level = degree_display = department = ''

            members.append({
                'id':             u.id,
                'name':           u.get_full_name(),
                'user_type':      u.user_type,
                'is_hqp':         u.user_type == 'student',
                'degree_level':   degree_level,
                'degree_display': degree_display,
                'department':     department,
            })

        return JsonResponse({'members': members})
    except Project.DoesNotExist:
        return JsonResponse({'error': 'Project not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@researcher_required
def pi_report_view(request):
    from django.shortcuts import get_object_or_404
    from django.db.models import Prefetch, Case, When, IntegerField, Count, Sum, Q

    researcher = get_object_or_404(ResearcherProfile, user=request.user)

    # ── XLSX Export ───────────────────────────────────────────
    if request.GET.get('export') == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        RED   = 'C8102E'
        GREY  = 'F3F4F6'
        WHITE = 'FFFFFF'

        wb = Workbook()
        wb.remove(wb.active)

        def make_sheet(wb, title, headers):
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font      = Font(bold=True, color=WHITE, name='Arial', size=11)
                cell.fill      = PatternFill('solid', start_color=RED)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 20
            return ws

        def auto_width(ws, min_w=12, max_w=60):
            for col in ws.columns:
                length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in col
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

        def style_data_rows(ws):
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                bg = GREY if row_idx % 2 == 0 else WHITE
                for cell in row:
                    cell.fill      = PatternFill('solid', start_color=bg)
                    cell.font      = Font(name='Arial', size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # ── Sheet 1: Projects & Grants ────────────────────────
        ws = make_sheet(wb, 'Projects & Grants', [
            'Title', 'Organization', 'Program', 'Role', 'Status',
            'Currency', 'Total Funding', 'Awarded to IBME', 'Kept by IBME',
            'Start Date', 'End Date', 'Funding Type',
            'Description', 'Conception', 'Next Steps', 'IP Activities',
            'HQP', 'Collaborators & Partners',
        ])
        for p in Project.objects.filter(
            researcher=researcher, is_deleted=False
        ).exclude(status='rejected').prefetch_related(
            Prefetch(
                'tagged_members',
                queryset=CustomUser.objects.filter(user_type='student'),
                to_attr='hqp_students',
            ),
            'team_members',
        ).order_by('-start_date'):
            hqp     = '; '.join(u.get_full_name() for u in p.hqp_students)
            collabs = '; '.join(
                f"{m.name} ({m.get_partner_type_display()})" for m in p.team_members.all()
            )
            ws.append([
                p.title,
                p.funding_organization or '',
                p.program_name or '',
                p.get_role_display() or '',
                p.get_status_display() or '',
                p.currency or 'CAD',
                float(p.total_funding)       if p.total_funding       else '',
                float(p.funding_received)    if p.funding_received    else '',
                float(p.funding_kept_by_unb) if p.funding_kept_by_unb else '',
                str(p.start_date) if p.start_date else '',
                str(p.end_date)   if p.end_date   else '',
                p.funding_type  or '',
                p.description   or '',
                p.conception    or '',
                p.next_steps    or '',
                p.ip_activities or '',
                hqp,
                collabs,
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 2: Publications ─────────────────────────────
        ws = make_sheet(wb, 'Publications', [
            'Title', 'Type', 'Status', 'Authors', 'Journal / Venue',
            'Date', 'Year', 'Volume', 'Issue', 'Pages', 'DOI',
            'Refereed', 'Open Access', 'Invited', 'Language', 'Abstract',
        ])
        for p in Publication.objects.filter(
            researcher=researcher, is_active=True
        ).order_by('-publication_date'):
            ws.append([
                p.title,
                p.get_publication_type_display(),
                p.get_status_display() or '',
                p.authors  or '',
                p.journal  or '',
                str(p.publication_date)     if p.publication_date else '',
                str(p.publication_date)[:4] if p.publication_date else '',
                p.volume   or '',
                p.issue    or '',
                p.pages    or '',
                p.doi      or '',
                'Yes' if p.refereed    else ('No' if p.refereed    is False else ''),
                'Yes' if p.open_access else ('No' if p.open_access is False else ''),
                'Yes' if p.invited     else ('No' if p.invited     is False else ''),
                p.language or '',
                p.abstract or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 3: Activities ───────────────────────────────
        ws = make_sheet(wb, 'Activities', [
            'Title', 'Category', 'Type', 'Date', 'Location',
            'Invited', 'Keynote', 'Co-Presenters', 'Audience', 'Description',
        ])
        for a in Activity.objects.filter(
            researcher=researcher, is_active=True
        ).order_by('-date'):
            ws.append([
                a.title,
                a.get_category_display()      or '',
                a.get_activity_type_display() or '',
                str(a.date) if a.date else '',
                a.location      or '',
                'Yes' if a.invited else ('No' if a.invited is False else ''),
                'Yes' if a.keynote else ('No' if a.keynote is False else ''),
                a.co_presenters or '',
                a.audience      or '',
                a.description   or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 4: Supervision ──────────────────────────────
        ws = make_sheet(wb, 'Supervision', [
            'Student', 'Degree', 'Department', 'Role', 'Institution',
            'Start', 'End', 'Status', 'Thesis Title',
        ])
        for s in SupervisionRecord.objects.filter(
            researcher=researcher
        ).order_by('-start_date'):
            ws.append([
                s.student_name,
                s.get_degree_type_display() or '',
                s.department       or '',
                s.supervision_role or '',
                s.institution      or '',
                str(s.start_date) if s.start_date else '',
                str(s.end_date)   if s.end_date   else '',
                s.get_status_display() or '',
                s.thesis_title or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        # ── Sheet 5: Funding Breakdown ────────────────────────
        # grant_total = full project Total Funding from XML (blue bar)
        # amount      = researcher's allocated portion (may be 0 for Co-I)
        # amount_to_ibme = what actually came to IBME
        ws = make_sheet(wb, 'Funding Breakdown', [
            'Grant Title', 'Organization', 'Program', 'Role', 'Status',
            'Currency', 'Grant Total', 'Researcher Portion', 'Amount to IBME',
            'Start Date', 'End Date', 'Funding Type',
        ])
        for f in Funding.objects.filter(
            researcher=researcher
        ).exclude(status='rejected').order_by('-start_date'):
            ws.append([
                f.title,
                f.organization  or '',
                f.program_name  or '',
                f.get_role_display()   or '',
                f.get_status_display() or '',
                f.currency or 'CAD',
                float(f.grant_total) if f.grant_total else (float(f.amount) if f.amount else ''),
                float(f.amount)         if f.amount         else '',
                float(f.amount_to_ibme) if f.amount_to_ibme else '',
                str(f.start_date) if f.start_date else '',
                str(f.end_date)   if f.end_date   else '',
                f.funding_type or '',
            ])
        style_data_rows(ws)
        auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="pi_report_{researcher.user.last_name}.xlsx"'
        )
        return response

    # ── Date filters ──────────────────────────────────────────
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    date_from     = None
    date_to       = None

    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except ValueError:
        pass
    try:
        if date_to_str:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        pass

    # ── Projects ──────────────────────────────────────────────
    projects_qs = Project.objects.filter(
        researcher=researcher,
        is_deleted=False,
    ).exclude(
        status='rejected',
    ).prefetch_related(
        'team_members',
        Prefetch(
            'tagged_members',
            queryset=CustomUser.objects.select_related('student_profile'),
        ),
        'linked_publications',
        'linked_activities',
        'funding_breakdown',
    )

    if date_from:
        projects_qs = projects_qs.filter(start_date__gte=date_from)
    if date_to:
        projects_qs = projects_qs.filter(start_date__lte=date_to)

    projects = projects_qs.order_by(
        Case(When(status='awarded', then=0), default=1, output_field=IntegerField()),
        '-start_date',
    )

    # ── Funding records ───────────────────────────────────────
    funding_records = Funding.objects.filter(
        researcher=researcher,
    ).exclude(
        status='rejected',
    ).select_related(
        'project',
    ).prefetch_related(
        'project__team_members',
    ).order_by('-start_date')

    if date_from:
        funding_records = funding_records.filter(start_date__gte=date_from)
    if date_to:
        funding_records = funding_records.filter(start_date__lte=date_to)

    # ── Publications ──────────────────────────────────────────
    publications_qs = Publication.objects.filter(
        researcher=researcher, is_active=True
    ).order_by('-publication_date')

    if date_from:
        publications_qs = publications_qs.filter(publication_date__gte=date_from)
    if date_to:
        publications_qs = publications_qs.filter(publication_date__lte=date_to)

    # ── Activities ────────────────────────────────────────────
    activities_qs = Activity.objects.filter(
        researcher=researcher, is_active=True
    ).select_related('conference').prefetch_related(
        'tagged_users', 'objectives'
    ).order_by('-date')

    if date_from:
        activities_qs = activities_qs.filter(date__gte=date_from)
    if date_to:
        activities_qs = activities_qs.filter(date__lte=date_to)

    # ── Supervision ───────────────────────────────────────────
    supervision = SupervisionRecord.objects.filter(researcher=researcher)

    if date_from:
        supervision = supervision.filter(start_date__gte=date_from)
    if date_to:
        supervision = supervision.filter(start_date__lte=date_to)

    supervision = supervision.order_by(
        Case(
            When(status='in_progress', then=0),
            When(status='completed',   then=1),
            default=2,
            output_field=IntegerField(),
        ),
        '-start_date',
    )

    # ── KPI aggregation ───────────────────────────────────────
    # total_funding: PI-role awarded grants only (researcher controls budget)
    # total_funding_all: all roles awarded (full association value)
    proj_stats = projects_qs.aggregate(
        total=Count('id'),
        awarded=Count('id', filter=Q(status='awarded')),
        submitted=Count('id', filter=Q(status='submitted')),
        completed=Count('id', filter=Q(status='completed')),
        pi_funding=Sum(                    # ← renamed
            'total_funding',
            filter=Q(status='awarded', role__in=['pi', 'pa'])
        ),
        all_funding=Sum(                   # ← renamed
            'total_funding',
            filter=Q(status='awarded')
        ),
    )

    return render(request, 'Pages/reports/pi_report.html', {
        'researcher':         researcher,
        'projects':           projects,
        'funding_records':    funding_records,
        'publications':       publications_qs,
        'activities':         activities_qs,
        'supervision':        supervision,
        'date_from':          date_from_str,
        'date_to':            date_to_str,
        'proj_count':         proj_stats['total']      or 0,
        'pub_count':          publications_qs.count(),
        'act_count':          activities_qs.count(),
        'sup_count':          supervision.count(),
        'funding_count':      funding_records.count(),
        'awarded_count':      proj_stats['awarded']    or 0,
        'submitted_count':    proj_stats['submitted']  or 0,
        'completed_count':    proj_stats['completed']  or 0,
        'total_funding':      proj_stats['pi_funding']  or 0,   # ← updated
        'total_funding_all':  proj_stats['all_funding'] or 0,   # ← updated
    })